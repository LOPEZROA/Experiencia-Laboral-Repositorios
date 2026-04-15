import re
import time
import json
import csv
import io
import os
import gc
import sys
import unicodedata
import threading
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Callable
from collections import defaultdict
import numpy as np
import pandas as pd
from dateutil import parser as dateparser
from flask import session
import openpyxl
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList

try:
    from flask_session import Session
except Exception:
    Session = None

try:
    import redis
except Exception:
    redis = None

try:
    import psycopg2
except Exception:
    psycopg2 = None

from .auth_db import is_admin_user as _auth_is_admin_user
from .config import (
    ADMIN_CORES_SQL_PATH,
    ADMIN_CORES_TRIGGER_TABLES,
    ALLOWED_C_SALIDA_VALUES,
    ALLOWED_EXTENSIONS,
    APP_OUTPUT_PREFIXES,
    BULK_SQL_CROSSES,
    BACKGROUND_CLEANUP_ENABLED,
    BACKGROUND_CLEANUP_STARTUP_DELAY_SECONDS,
    C_SALIDA_LABELS,
    COMPACT_DB_INDEX,
    CLEANUP_INTERVAL_SECONDS,
    DB_DIR,
    ESTAB_DEST_FILTER,
    FILE_RETENTION_SECONDS,
    JOB_RETENTION_SECONDS,
    JOB_RUNNING_TTL_SECONDS,
    JOBS_BACKEND,
    JOB_STORE_PREFIX,
    MAX_CONCURRENT_JOBS,
    MAX_STORED_JOBS,
    NOMINA_NAME_RE,
    NOMINA_VERIFY_LAZY_LOAD,
    NOMINA_STATE_ALIASES,
    NOMINA_TYPE_ALIASES,
    LOW_MEMORY_DB_REFRESH,
    MEMORY_TRIM_ROUNDS,
    MEMORY_TRIM_SLEEP_MS,
    OUTPUT_DIR,
    PROCESS_MAX_CPU_PERCENT,
    PG_BASE_TABLES,
    PG_DATABASE,
    PG_DSN,
    PG_HOST,
    PG_NOMINA_TABLES,
    PG_PASSWORD,
    PG_PORT,
    PG_SCHEMA,
    PG_USER,
    REDIS_URL,
    REQUESTED_JOBS_BACKEND,
    REQUESTED_SESSION_BACKEND,
    ROW_LEVEL_SQL_LOOKUPS,
    SESSION_BACKEND,
    SQL_STREAM_BATCH_SIZE,
    STRICT_REDIS_BACKEND,
    PRELOAD_MAX_CPU_PERCENT,
    TABLE_PREVIEW_LIMIT,
    UPLOAD_DIR,
    VERIFY_FIELDS,
)

def canon(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    s = re.sub(r"[\s\-_]+", "", s)
    s = re.sub(r"[^a-z0-9]", "", s)
    return s

def normalize_run(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, (float, np.floating)):
        if float(v).is_integer():
            return str(int(v))
    s = str(v).strip()
    s = s.replace(".", "").replace(" ", "")
    s = s.lstrip("0") or "0"
    return s.upper()

def normalize_dv(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, (float, np.floating)):
        if float(v).is_integer():
            return str(int(v))
    s = str(v).strip().replace(".", "").replace(" ", "").replace("-", "")
    return s.upper()


def _release_memory_pressure(rounds: Optional[int] = None, sleep_ms: Optional[int] = None) -> None:
    trim_rounds = MEMORY_TRIM_ROUNDS if rounds is None else rounds
    trim_sleep_ms = MEMORY_TRIM_SLEEP_MS if sleep_ms is None else sleep_ms
    try:
        trim_rounds = max(1, min(10, int(trim_rounds or 1)))
    except Exception:
        trim_rounds = 1
    try:
        sleep_secs = max(0.0, min(0.5, float(trim_sleep_ms or 0) / 1000.0))
    except Exception:
        sleep_secs = 0.0

    for _ in range(trim_rounds):
        try:
            gc.collect()
        except Exception:
            pass
        try:
            if os.name == "nt":
                import ctypes
                handle = ctypes.windll.kernel32.GetCurrentProcess()
                ctypes.windll.psapi.EmptyWorkingSet(handle)
            else:
                import ctypes
                libc = None
                for name in ("libc.so.6", "libc.so", "libSystem.B.dylib"):
                    try:
                        libc = ctypes.CDLL(name)
                        break
                    except Exception:
                        continue
                if libc is not None and hasattr(libc, "malloc_trim"):
                    libc.malloc_trim(0)
        except Exception:
            pass
        if sleep_secs > 0:
            try:
                time.sleep(sleep_secs)
            except Exception:
                pass


def normalize_rut_concat(v: Any) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    s = str(v).strip().upper().replace(".", "").replace(" ", "")
    if not s:
        return ""
    if "-" in s:
        parts = s.split("-")
        if len(parts) >= 2:
            run = normalize_run(parts[0])
            dv = normalize_dv(parts[1][:1])
            if run and dv:
                return f"{run}-{dv}"
    if len(s) >= 2 and re.fullmatch(r"[0-9]+[0-9K]", s):
        run = normalize_run(s[:-1])
        dv = normalize_dv(s[-1])
        if run and dv:
            return f"{run}-{dv}"
    return s


def normalize_rut_compact(v: Any) -> str:
    rut = normalize_rut_concat(v)
    if not rut:
        return ""
    compact = re.sub(r"[^0-9K]", "", str(rut).upper())
    if re.fullmatch(r"[0-9]+[0-9K]", compact):
        return compact
    return ""


def _intern_str(value: Any) -> str:
    if value is None:
        return ""
    s = str(value)
    if not s:
        return ""
    if len(s) <= 48:
        try:
            return sys.intern(s)
        except Exception:
            return s
    return s


def normalize_text(v: Any) -> str:
    s = str(v).strip()
    s = s.replace(".", "").replace(" ", "")
    s = s.replace("-", "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.upper()

def normalize_presta(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, (float, np.floating)):
        if float(v).is_integer():
            return str(int(v))
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return normalize_text(s)

def normalize_id(v: Any) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, (float, np.floating)):
        if float(v).is_integer():
            return str(int(v))
        s = str(v).strip()
        s = s.rstrip("0").rstrip(".")
        return s
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.strip()


_OBSERVACION_SALUD_MENTAL_PRESTA_CODES = ("07-048", "07-049", "07-117-0")
_OBSERVACION_SALUD_MENTAL_PRESTA_CODES_NORM = {
    normalize_presta(code) for code in _OBSERVACION_SALUD_MENTAL_PRESTA_CODES
}


def classify_observacion_caso(
    sigte_id: Any,
    presta_min: Any,
    db: Optional["DBIndex"],
) -> str:
    sigte_norm = normalize_id(sigte_id)
    if not sigte_norm:
        return "Sin observaciones"
    try:
        special_ids = db.comges_especiales_sigte_ids if db is not None else set()
    except Exception:
        special_ids = set()
    if sigte_norm not in special_ids:
        return "Sin observaciones"
    presta_norm = normalize_presta(presta_min)
    if presta_norm in _OBSERVACION_SALUD_MENTAL_PRESTA_CODES_NORM:
        return "Salud mental"
    return "Sename"


def normalize_compare_value(field: str, v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""

    if field in ("RUN", "RUN_PROF_SOL", "RUN_PROF_RESOL"):
        return normalize_run(v)
    if field in ("DV", "DV_PROF_SOL", "DV_PROF_RESOL"):
        return normalize_dv(v)
    if field in ("FECHA_NAC", "F_ENTRADA", "F_SALIDA"):
        dt = parse_excel_date(v)
        return dt.date().isoformat() if dt else ""
    return normalize_text(v)

def parse_excel_date(v: Any) -> Optional[datetime]:
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, datetime):
        return v
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day") and not isinstance(v, (int, float, str)):
        try:
            return datetime(v.year, v.month, v.day)
        except Exception:
            pass
    if isinstance(v, (np.integer, np.floating)) and not pd.isna(v):
        base = datetime(1899, 12, 30)
        try:
            return base + timedelta(days=int(v))
        except Exception:
            return None
    if isinstance(v, (int, float)) and not pd.isna(v):
        base = datetime(1899, 12, 30)
        try:
            return base + timedelta(days=int(v))
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None

    m_iso = re.fullmatch(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})(?:[ T].*)?", s)
    if m_iso:
        try:
            y, m, d = (int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3)))
            return datetime(y, m, d)
        except Exception:
            pass

    m_lat = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})(?:[ T].*)?", s)
    if m_lat:
        try:
            d, m, y = (int(m_lat.group(1)), int(m_lat.group(2)), int(m_lat.group(3)))
            return datetime(y, m, d)
        except Exception:
            pass

    s_num = s.replace(",", ".")
    if re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", s_num):
        try:
            serial = float(s_num)
            if 10000 <= serial <= 100000:
                base = datetime(1899, 12, 30)
                return base + timedelta(days=int(serial))
        except Exception:
            pass
    s = s.replace("-", "/")
    try:
        dt = dateparser.parse(s, dayfirst=True)
        if dt is None:
            return None
        return dt
    except Exception:
        return None


def normalize_date(v: Any) -> Optional[datetime]:
    return parse_excel_date(v)


def _normalize_prevision_value(v: Any) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, (float, np.floating)):
        try:
            if float(v).is_integer():
                return str(int(v))
            return str(v).strip()
        except Exception:
            return str(v).strip()
    s = str(v).strip()
    if not s:
        return ""
    if re.fullmatch(r"[0-9]+(?:\.0+)?", s):
        try:
            return str(int(float(s)))
        except Exception:
            return s
    return s.upper()


def _append_issue_to_missing_report(
    report_series: pd.Series,
    issue_mask: pd.Series,
    issue_text: str,
) -> pd.Series:
    if report_series is None:
        return pd.Series([], dtype=object)
    out = report_series.fillna("").astype(str).copy()
    if out.empty:
        return out
    mask = pd.Series(issue_mask, index=out.index).fillna(False).astype(bool)
    if not bool(mask.any()):
        return out
    issue_norm = canon(issue_text)
    sin_datos_norm = canon("Sin datos faltantes")
    falta_norm = canon("Falta")
    for idx in out.index[mask]:
        raw = str(out.at[idx]).strip()
        if issue_norm and issue_norm in canon(raw):
            continue
        raw_norm = canon(raw)
        if (not raw) or (raw_norm == sin_datos_norm):
            out.at[idx] = f"Falta: {issue_text}"
            continue
        if raw_norm.startswith(falta_norm):
            out.at[idx] = f"{raw}, {issue_text}"
            continue
        out.at[idx] = f"{raw} | {issue_text}"
    return out


def _format_short_date_value(v: Any) -> Any:
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    dt = parse_excel_date(v)
    if dt is None:
        return None
    return datetime(dt.year, dt.month, dt.day)


def _is_date_like_column_name(col_name: str) -> bool:
    c = canon(col_name)
    if not c:
        return False
    if "fecha" in c:
        return True
    return c in {"fnac", "fentrada", "fsalida", "fdefuncion"}


def format_dates_for_export(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    if df.empty:
        return df.copy()
    out = df.copy()
    for col in out.columns:
        s = out[col]
        if pd.api.types.is_datetime64_any_dtype(s):
            out[col] = pd.to_datetime(s, errors="coerce")
            continue
        if _is_date_like_column_name(str(col)):
            out[col] = pd.to_datetime(s.map(_format_short_date_value), errors="coerce")
    return out


def apply_short_date_format_to_workbook(workbook: Any) -> None:
    if workbook is None:
        return
    try:
        worksheets = list(workbook.worksheets)
    except Exception:
        worksheets = []
    for ws in worksheets:
        try:
            max_row = int(ws.max_row or 0)
            max_col = int(ws.max_column or 0)
        except Exception:
            continue
        if max_row < 2 or max_col < 1:
            continue
        date_cols: List[int] = []
        for col_idx in range(1, max_col + 1):
            header = ws.cell(row=1, column=col_idx).value
            if _is_date_like_column_name(str(header or "")):
                date_cols.append(col_idx)
        if not date_cols:
            continue
        for col_idx in date_cols:
            for row_idx in range(2, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None:
                    continue
                if isinstance(val, datetime):
                    cell.value = val.date()
                    cell.number_format = "DD-MM-YYYY"
                elif isinstance(val, date):
                    cell.number_format = "DD-MM-YYYY"


def format_duration(seconds: float) -> str:
    if seconds is None:
        return "0 min 00.00 s"
    if seconds < 0:
        seconds = 0.0
    total_minutes = int(seconds // 60)
    secs = seconds - (total_minutes * 60)
    return f"{total_minutes} min {secs:05.2f} s"


def to_excel_serial(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        base = datetime(1899, 12, 30)
        return str((dt - base).days)
    if hasattr(dt, "year") and hasattr(dt, "month") and hasattr(dt, "day"):
        try:
            dd = datetime(dt.year, dt.month, dt.day)
            base = datetime(1899, 12, 30)
            return str((dd - base).days)
        except Exception:
            pass
    parsed = parse_excel_date(dt)
    if parsed is None:
        return ""
    base = datetime(1899, 12, 30)
    return str((parsed - base).days)


def allowed_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def ensure_dirs() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _pg_quote_ident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


def _pg_qualified_table(table_name: str) -> str:
    raw_name = str(table_name or "").strip()
    if "." in raw_name:
        schema_name, rel_name = raw_name.split(".", 1)
        schema_name = schema_name.strip() or (PG_SCHEMA if PG_SCHEMA else "public")
        rel_name = rel_name.strip()
        return f"{_pg_quote_ident(schema_name)}.{_pg_quote_ident(rel_name)}"
    schema = PG_SCHEMA if PG_SCHEMA else "public"
    return f"{_pg_quote_ident(schema)}.{_pg_quote_ident(raw_name)}"


def _sql_query_dataframe(conn: Any, query: str, params: Optional[List[Any]] = None) -> pd.DataFrame:
    with conn.cursor() as cur:
        cur.execute(query, params or [])
        rows = cur.fetchall()
        cols = [str(d[0]).strip() for d in (cur.description or [])]
    if not cols:
        return pd.DataFrame()
    if not rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame.from_records(rows, columns=cols)


def _sql_table_exists(conn: Any, table_name: str) -> bool:
    try:
        qualified = _pg_qualified_table(table_name)
        with conn.cursor() as cur:
            cur.execute(f"SELECT 1 FROM {qualified} LIMIT 1")
        return True
    except Exception:
        return False


def _pg_connect():
    if psycopg2 is None:
        raise RuntimeError(
            "No se encontró 'psycopg2'. Instala 'psycopg2-binary' para conectar a PostgreSQL."
        )
    if PG_DSN:
        return psycopg2.connect(PG_DSN)
    return psycopg2.connect(
        host=PG_HOST,
        port=PG_PORT,
        dbname=PG_DATABASE,
        user=PG_USER,
        password=PG_PASSWORD,
        connect_timeout=10,
    )


_ESTAB_DEST_FILTER_TABLES = {
    *(v for v in [PG_BASE_TABLES.get("historico", "")] if v),
    *(v for v in PG_NOMINA_TABLES.values() if v),
    "core.nomina_ic",
    "core.nomina_iq",
    "core.nomina_proc",
    "core.historico",
}


def _sql_should_apply_estab_dest_filter(table_name: str) -> bool:
    return str(table_name or "").strip() in _ESTAB_DEST_FILTER_TABLES


def _sql_fetch_table_df(table_name: str, required_columns: Optional[List[str]] = None) -> pd.DataFrame:
    try:
        with _pg_connect() as conn:
            cols = _sql_fetch_table_columns(conn, table_name)
            if not cols:
                return pd.DataFrame()
            selected_cols: List[str] = []
            if required_columns:
                for req in required_columns:
                    picked = _sql_pick_column(cols, [str(req)])
                    if picked and picked not in selected_cols:
                        selected_cols.append(picked)
            estab_filter = str(ESTAB_DEST_FILTER) if _sql_should_apply_estab_dest_filter(table_name) else ""
            query, params = _sql_build_filtered_search_query(
                table_name=table_name,
                columns=cols,
                rut_filter="",
                id_filter="",
                estab_dest_filter=estab_filter,
                select_columns=selected_cols,
            )
            df = _sql_query_dataframe(conn, query, params)
    except Exception as e:
        schema = PG_SCHEMA if PG_SCHEMA else "public"
        detail = str(e).strip()
        if not detail:
            detail = repr(e)
        raise RuntimeError(
            f"No se pudo consultar {schema}.{table_name} en PostgreSQL ({type(e).__name__}): {detail}"
        ) from e
    if df is None or df.empty:
        return pd.DataFrame()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _sql_build_table_select(
    table_name: str,
    columns: List[str],
    required_columns: Optional[List[str]] = None,
) -> Tuple[List[str], str, List[Any]]:
    selected_cols: List[str] = []
    if required_columns:
        for req in required_columns:
            picked = _sql_pick_column(columns, [str(req)])
            if picked and picked not in selected_cols:
                selected_cols.append(picked)
    estab_filter = str(ESTAB_DEST_FILTER) if _sql_should_apply_estab_dest_filter(table_name) else ""
    query, params = _sql_build_filtered_search_query(
        table_name=table_name,
        columns=columns,
        rut_filter="",
        id_filter="",
        estab_dest_filter=estab_filter,
        select_columns=selected_cols,
    )
    selected = selected_cols if selected_cols else list(columns)
    return selected, query, params


def _sql_stream_table_rows_conn(
    conn: Any,
    table_name: str,
    required_columns: Optional[List[str]] = None,
    batch_size: int = SQL_STREAM_BATCH_SIZE,
) -> Tuple[List[str], Iterable[Tuple[Any, ...]]]:
    cols = _sql_fetch_table_columns(conn, table_name)
    if not cols:
        return [], []
    selected_cols, query, params = _sql_build_table_select(
        table_name=table_name,
        columns=cols,
        required_columns=required_columns,
    )
    size = max(100, int(batch_size or SQL_STREAM_BATCH_SIZE or 5000))
    cursor_name = f"le_stream_{uuid.uuid4().hex[:12]}"
    cur = conn.cursor(name=cursor_name)
    cur.itersize = size
    cur.execute(query, params)

    def _iter_rows() -> Iterable[Tuple[Any, ...]]:
        try:
            while True:
                chunk = cur.fetchmany(size)
                if not chunk:
                    break
                for row in chunk:
                    yield row
        finally:
            try:
                cur.close()
            except Exception:
                pass

    header_vals = [str(c).strip() for c in selected_cols]
    return header_vals, _iter_rows()


def _sql_fetch_table_columns(conn: Any, table_name: str) -> List[str]:
    query = f"SELECT * FROM {_pg_qualified_table(table_name)} LIMIT 0"
    with conn.cursor() as cur:
        cur.execute(query)
        if not cur.description:
            return []
        return [str(d[0]).strip() for d in cur.description]


def _sql_build_filtered_search_query(
    table_name: str,
    columns: List[str],
    rut_filter: str = "",
    id_filter: str = "",
    estab_dest_filter: str = "",
    select_columns: Optional[List[str]] = None,
) -> Tuple[str, List[Any]]:
    colmap = {canon(c): c for c in columns}

    def _pick(cands: List[str]) -> str:
        for cand in cands:
            c = canon(cand)
            if c in colmap:
                return colmap[c]
            for k, real in colmap.items():
                if c and c in k:
                    return real
        return ""

    def _txt(col: str) -> str:
        return f"trim(COALESCE(CAST({_pg_quote_ident(col)} AS text), ''))"

    def _norm_no_spaces(col: str) -> str:
        return f"regexp_replace({_txt(col)}, '\\s+', '', 'g')"

    def _norm_run(col: str) -> str:
        base = _norm_no_spaces(col)
        return f"COALESCE(NULLIF(regexp_replace({base}, '^0+', ''), ''), '0')"

    def _norm_dv(col: str) -> str:
        base = _norm_no_spaces(col)
        return f"upper(regexp_replace({base}, '[^0-9Kk]', '', 'g'))"

    def _norm_rut(col: str) -> str:
        base = _norm_no_spaces(col)
        return f"upper(regexp_replace({base}, '[^0-9Kk]', '', 'g'))"

    def _norm_id(col: str) -> str:
        base = _norm_no_spaces(col)
        return f"regexp_replace({base}, '\\.0+$', '')"

    params: List[Any] = []
    where_parts: List[str] = []

    if rut_filter:
        run_col = _pick(["RUN", "run"])
        dv_col = _pick(["DV", "dv"])
        rut_col = _pick(["RUT CONCATENADO", "rut_concatenado", "RUTCONCATENADO", "RUT", "rut"])

        rut_joined = str(rut_filter).replace("-", "").strip().upper()
        run_val = ""
        dv_val = ""
        if "-" in rut_filter:
            run_raw, dv_raw = rut_filter.split("-", 1)
            run_val = normalize_run(run_raw)
            dv_val = normalize_dv(dv_raw[:1])
        elif len(rut_joined) >= 2:
            run_val = normalize_run(rut_joined[:-1])
            dv_val = normalize_dv(rut_joined[-1])

        rut_clause = ""
        if run_col and dv_col and run_val and dv_val:
            rut_clause = f"({_norm_run(run_col)} = %s AND {_norm_dv(dv_col)} = %s)"
            params.extend([run_val, dv_val])
        elif rut_col and rut_joined:
            rut_clause = f"({_norm_rut(rut_col)} = %s)"
            params.append(rut_joined)
        else:
            # No hay columnas suficientes para aplicar filtro de RUT en esta tabla.
            return f"SELECT * FROM {_pg_qualified_table(table_name)} WHERE 1=0", []
        where_parts.append(rut_clause)

    if id_filter:
        id_col = _pick(["ID_LOCAL_NORM", "id_local_norm", "ID_LOCAL", "id_local"])
        sigte_col = _pick(["SIGTE_ID_NORM", "sigte_id_norm", "SIGTE_ID", "sigte_id"])
        id_clauses: List[str] = []
        if id_col:
            if canon(id_col) == "idlocalnorm":
                id_clauses.append(f"{_txt(id_col)} = %s")
            else:
                id_clauses.append(f"{_norm_id(id_col)} = %s")
            params.append(id_filter)
        if sigte_col:
            if canon(sigte_col) == "sigteidnorm":
                id_clauses.append(f"{_txt(sigte_col)} = %s")
            else:
                id_clauses.append(f"{_norm_id(sigte_col)} = %s")
            params.append(id_filter)
        if not id_clauses:
            # No hay columnas suficientes para aplicar filtro de ID en esta tabla.
            return f"SELECT * FROM {_pg_qualified_table(table_name)} WHERE 1=0", []
        where_parts.append("(" + " OR ".join(id_clauses) + ")")

    if estab_dest_filter:
        estab_col = _pick(["ESTAB_DEST", "estab_dest"])
        if estab_col:
            where_parts.append(f"{_norm_id(estab_col)} = %s")
            params.append(str(estab_dest_filter))

    selected = [str(c).strip() for c in (select_columns or []) if str(c).strip()]
    selected_safe: List[str] = []
    if selected:
        known = {str(c).strip() for c in columns}
        for col in selected:
            if col in known and col not in selected_safe:
                selected_safe.append(col)
    select_sql = "*"
    if selected_safe:
        select_sql = ", ".join(_pg_quote_ident(c) for c in selected_safe)

    query = f"SELECT {select_sql} FROM {_pg_qualified_table(table_name)}"
    if where_parts:
        query += " WHERE " + " AND ".join(f"({p})" for p in where_parts)
    return query, params


def _sql_fetch_filtered_table_df(
    conn: Any,
    table_name: str,
    rut_filter: str = "",
    id_filter: str = "",
) -> pd.DataFrame:
    cols = _sql_fetch_table_columns(conn, table_name)
    if not cols:
        return pd.DataFrame()
    estab_filter = str(ESTAB_DEST_FILTER) if _sql_should_apply_estab_dest_filter(table_name) else ""
    query, params = _sql_build_filtered_search_query(
        table_name=table_name,
        columns=cols,
        rut_filter=rut_filter,
        id_filter=id_filter,
        estab_dest_filter=estab_filter,
    )
    df = _sql_query_dataframe(conn, query, params)
    if df is None or df.empty:
        return pd.DataFrame()
    df.columns = [str(c).strip() for c in df.columns]
    return df


_SQL_TIMELINE_COLUMNS_CACHE: Dict[str, List[str]] = {}
_SQL_TIMELINE_COLUMNS_LOCK = threading.Lock()


def _clear_sql_timeline_columns_cache() -> None:
    with _SQL_TIMELINE_COLUMNS_LOCK:
        _SQL_TIMELINE_COLUMNS_CACHE.clear()


def _sql_get_table_columns_cached(conn: Any, table_name: str) -> List[str]:
    cached = _SQL_TIMELINE_COLUMNS_CACHE.get(table_name)
    if cached is not None:
        return cached
    cols = _sql_fetch_table_columns(conn, table_name)
    with _SQL_TIMELINE_COLUMNS_LOCK:
        _SQL_TIMELINE_COLUMNS_CACHE[table_name] = cols
    return cols


def _sql_pick_column(columns: List[str], candidates: List[str]) -> str:
    colmap = {canon(c): c for c in columns}
    for cand in candidates:
        key = canon(cand)
        if key in colmap:
            return colmap[key]
        for norm_col, real_col in colmap.items():
            if key and key in norm_col:
                return real_col
    return ""


def _sql_expr_trim(col: str) -> str:
    return f"trim(COALESCE(CAST({_pg_quote_ident(col)} AS text), ''))"


def _sql_expr_norm_run(col: str) -> str:
    raw = _sql_expr_trim(col)
    digits = f"regexp_replace({raw}, '[^0-9]', '', 'g')"
    return f"COALESCE(NULLIF(regexp_replace({digits}, '^0+', ''), ''), '0')"


def _sql_expr_norm_dv(col: str) -> str:
    raw = _sql_expr_trim(col)
    return f"upper(regexp_replace({raw}, '[^0-9Kk]', '', 'g'))"


def _sql_expr_norm_text_legacy(col: str) -> str:
    raw = _sql_expr_trim(col)
    translated = f"translate({raw}, 'ÃÃ‰ÃÃ“ÃšÃœÃ‘Ã¡Ã©Ã­Ã³ÃºÃ¼Ã±', 'AEIOUUNAEIOUUN')"
    return f"upper(regexp_replace({translated}, '[\\s\\.\\-]+', '', 'g'))"


def _sql_expr_norm_text(col: str) -> str:
    raw = _sql_expr_trim(col)
    return f"upper(regexp_replace({raw}, '[\\s\\.\\-]+', '', 'g'))"


def _sql_expr_norm_presta(col: str) -> str:
    return _sql_expr_norm_text(col)


def _sql_expr_norm_id(col: str) -> str:
    raw = _sql_expr_trim(col)
    no_spaces = f"regexp_replace({raw}, '\\s+', '', 'g')"
    return f"regexp_replace({no_spaces}, '\\.0+$', '')"


def _sql_expr_date(col: str) -> str:
    raw = _sql_expr_trim(col)
    raw_num = f"replace({raw}, ',', '.')"
    raw_token = f"split_part({raw}, ' ', 1)"
    iso_token = f"replace({raw_token}, '/', '-')"
    lat_token = f"replace({raw_token}, '-', '/')"
    return (
        "CASE "
        f"WHEN {raw} = '' THEN NULL::date "
        f"WHEN {raw_num} ~ '^[0-9]+(?:\\.[0-9]+)?$' THEN (DATE '1899-12-30' + FLOOR(({raw_num})::numeric)::int) "
        f"WHEN {raw_token} ~ '^\\d{{4}}[/-]\\d{{1,2}}[/-]\\d{{1,2}}$' THEN ({iso_token})::date "
        f"WHEN {raw_token} ~ '^\\d{{1,2}}[/-]\\d{{1,2}}[/-]\\d{{4}}$' THEN to_date({lat_token}, 'DD/MM/YYYY') "
        "ELSE NULL::date END"
    )


def _sql_expr_norm_compare(col: Optional[str], field: str) -> str:
    if not col:
        return "''::text"
    if field in {"RUN", "RUN_PROF_SOL", "RUN_PROF_RESOL"}:
        return _sql_expr_norm_run(col)
    if field in {"DV", "DV_PROF_SOL", "DV_PROF_RESOL"}:
        return _sql_expr_norm_dv(col)
    if field in {"FECHA_NAC", "F_ENTRADA", "F_SALIDA"}:
        return f"COALESCE(to_char({_sql_expr_date(col)}, 'YYYY-MM-DD'), '')"
    return _sql_expr_norm_text(col)


_CORE_NOMINA_TABLES = [
    "core.nomina_ic",
    "core.nomina_iq",
    "core.nomina_proc",
]
_CORE_HISTORICO_TABLE = "core.historico"
_CORE_COMGES_ESPECIALES_TABLE = "core.comges_especiales"


def _sql_nomina_table_names(conn: Optional[Any] = None) -> List[str]:
    if conn is not None:
        core_tables = [t for t in _CORE_NOMINA_TABLES if _sql_table_exists(conn, t)]
        if core_tables:
            return sorted(core_tables)
    return sorted({str(v).strip() for v in PG_NOMINA_TABLES.values() if str(v).strip()})


def _sql_historico_table_name(conn: Optional[Any] = None) -> str:
    if conn is not None and _sql_table_exists(conn, _CORE_HISTORICO_TABLE):
        return _CORE_HISTORICO_TABLE
    return str(PG_BASE_TABLES.get("historico", "") or "").strip()


def _sql_comges_especiales_table_name(conn: Optional[Any] = None) -> str:
    if conn is not None and _sql_table_exists(conn, _CORE_COMGES_ESPECIALES_TABLE):
        return _CORE_COMGES_ESPECIALES_TABLE
    return str(PG_BASE_TABLES.get("comges_especiales", "") or "").strip()


def _sql_fetch_nomina_lookup_for_ids(
    conn: Any,
    ids: Iterable[str],
    include_records: bool = True,
) -> Tuple[Dict[str, str], Dict[str, str], Dict[str, Dict[str, str]]]:
    normalized_ids: List[str] = []
    seen_ids: set = set()
    for raw in ids:
        key = normalize_id(raw)
        if not key or key in seen_ids:
            continue
        seen_ids.add(key)
        normalized_ids.append(key)
    if not normalized_ids:
        return {}, {}, {}

    tmp_table = f"tmp_nomina_ids_{uuid.uuid4().hex[:10]}"
    q_tmp = _pg_quote_ident(tmp_table)
    with conn.cursor() as cur:
        cur.execute(f"CREATE TEMP TABLE {q_tmp} (id_norm text PRIMARY KEY) ON COMMIT DROP;")
    buf = io.StringIO()
    buf.write("id_norm\n")
    for key in normalized_ids:
        buf.write(f"{key}\n")
    buf.seek(0)
    with conn.cursor() as cur:
        cur.copy_expert(
            f"COPY {q_tmp} (id_norm) FROM STDIN WITH (FORMAT csv, HEADER true);",
            buf,
        )

    out_by_id: Dict[str, str] = {}
    out_by_source: Dict[str, str] = {}
    out_by_record: Dict[str, Dict[str, str]] = {}

    table_names = _sql_nomina_table_names(conn)
    for table_name in table_names:
        cols = _sql_fetch_table_columns(conn, table_name)
        if not cols:
            continue
        col_id = _sql_pick_column(cols, ["ID_LOCAL_NORM", "id_local_norm", "ID_LOCAL", "id_local"])
        col_sig = _sql_pick_column(cols, ["SIGTE_ID_NORM", "sigte_id_norm", "SIGTE_ID", "sigte_id"])
        if not col_id and not col_sig:
            continue
        id_expr = _sql_expr_trim(col_id) if canon(col_id) == "idlocalnorm" else (_sql_expr_norm_id(col_id) if col_id else "''::text")
        sig_expr = _sql_expr_trim(col_sig) if canon(col_sig) == "sigteidnorm" else (_sql_expr_norm_id(col_sig) if col_sig else "''::text")

        where_parts: List[str] = []
        if col_id:
            where_parts.append(f"{id_expr} IN (SELECT id_norm FROM {q_tmp})")
        if col_sig:
            where_parts.append(f"{sig_expr} IN (SELECT id_norm FROM {q_tmp})")
        if not where_parts:
            continue

        verify_aliases: List[Tuple[str, str]] = []
        verify_selects: List[str] = []
        if include_records:
            for field, candidates in VERIFY_FIELDS.items():
                col_field = _sql_pick_column(cols, candidates)
                alias = f"vf_{field.lower()}"
                verify_aliases.append((field, alias))
                verify_selects.append(
                    f"{_sql_expr_norm_compare(col_field, field)} AS {_pg_quote_ident(alias)}"
                )

        select_parts: List[str] = [
            f"{id_expr} AS id_local_norm",
            f"{sig_expr} AS sigte_id_norm",
        ]
        if verify_selects:
            select_parts.extend(verify_selects)
        query = (
            "SELECT "
            + ", ".join(select_parts)
            + f" FROM {_pg_qualified_table(table_name)} WHERE "
            + "(" + " OR ".join(where_parts) + ")"
        )

        df = _sql_query_dataframe(conn, query, [])
        if df is None or df.empty:
            continue

        for row in df.to_dict(orient="records"):
            id_local_norm = normalize_id(row.get("id_local_norm"))
            sigte_id_norm = normalize_id(row.get("sigte_id_norm"))
            sigte_val = sigte_id_norm or id_local_norm

            rec: Dict[str, str] = {}
            if include_records:
                for field, alias in verify_aliases:
                    rec[field] = str(row.get(alias, "") or "")

            for key in (id_local_norm, sigte_id_norm):
                if not key:
                    continue
                out_by_id[key] = sigte_val or key
                out_by_source[key] = table_name
                if include_records:
                    out_by_record[key] = rec

    return out_by_id, out_by_source, out_by_record


def _sql_fetch_historico_lookup_for_ids(
    conn: Any,
    ids: Iterable[str],
) -> Dict[str, str]:
    normalized_ids: List[str] = []
    seen_ids: set = set()
    for raw in ids:
        key = normalize_id(raw)
        if not key or key in seen_ids:
            continue
        seen_ids.add(key)
        normalized_ids.append(key)
    if not normalized_ids:
        return {}

    hist_table = _sql_historico_table_name(conn)
    if not hist_table:
        return {}

    cols = _sql_fetch_table_columns(conn, hist_table)
    if not cols:
        return {}
    col_id = _sql_pick_column(cols, ["ID_LOCAL_NORM", "id_local_norm", "ID_LOCAL", "id_local"])
    col_sig = _sql_pick_column(cols, ["SIGTE_ID_NORM", "sigte_id_norm", "SIGTE_ID", "sigte_id"])
    if not col_id and not col_sig:
        return {}

    id_expr = _sql_expr_trim(col_id) if canon(col_id) == "idlocalnorm" else (_sql_expr_norm_id(col_id) if col_id else "''::text")
    sig_expr = _sql_expr_trim(col_sig) if canon(col_sig) == "sigteidnorm" else (_sql_expr_norm_id(col_sig) if col_sig else "''::text")

    tmp_table = f"tmp_hist_ids_{uuid.uuid4().hex[:10]}"
    q_tmp = _pg_quote_ident(tmp_table)
    with conn.cursor() as cur:
        cur.execute(f"CREATE TEMP TABLE {q_tmp} (id_norm text PRIMARY KEY) ON COMMIT DROP;")
    buf = io.StringIO()
    buf.write("id_norm\n")
    for key in normalized_ids:
        buf.write(f"{key}\n")
    buf.seek(0)
    with conn.cursor() as cur:
        cur.copy_expert(
            f"COPY {q_tmp} (id_norm) FROM STDIN WITH (FORMAT csv, HEADER true);",
            buf,
        )

    query = (
        "WITH src AS ("
        "SELECT "
        f"{id_expr} AS id_local_norm, "
        f"{sig_expr} AS sigte_id_norm, "
        f"COALESCE(NULLIF({sig_expr}, ''), NULLIF({id_expr}, '')) AS sigte_final "
        f"FROM {_pg_qualified_table(hist_table)}"
        ") "
        "SELECT k.id_norm, MIN(src.sigte_final) AS sigte_final "
        f"FROM {q_tmp} k "
        "JOIN src ON k.id_norm = src.id_local_norm OR k.id_norm = src.sigte_id_norm "
        "WHERE src.sigte_final <> '' "
        "GROUP BY k.id_norm"
    )
    df = _sql_query_dataframe(conn, query, [])
    if df is None or df.empty:
        return {}
    out: Dict[str, str] = {}
    for row in df.to_dict(orient="records"):
        key = normalize_id(row.get("id_norm"))
        sigte = normalize_id(row.get("sigte_final"))
        if key and sigte:
            out[key] = sigte
    return out


def _sql_fetch_comges_special_ids(
    conn: Any,
    sigte_ids: Iterable[str],
) -> set:
    normalized_ids: List[str] = []
    seen_ids: set = set()
    for raw in sigte_ids:
        key = normalize_id(raw)
        if not key or key in seen_ids:
            continue
        seen_ids.add(key)
        normalized_ids.append(key)
    if not normalized_ids:
        return set()

    table_name = _sql_comges_especiales_table_name(conn)
    if not table_name:
        return set()
    cols = _sql_fetch_table_columns(conn, table_name)
    if not cols:
        return set()

    sigte_candidates = {
        canon("SIGTE_ID"),
        canon("sigte_id"),
        canon("SIGTE_ID2"),
        canon("sigte_id2"),
        canon("SIGTE_ID.1"),
        canon("sigte_id.1"),
        canon("SIGTE_ID_NORM"),
        canon("sigte_id_norm"),
    }
    sigte_cols = [str(col) for col in cols if canon(col) in sigte_candidates]
    if not sigte_cols:
        return set()

    tmp_table = f"tmp_comges_ids_{uuid.uuid4().hex[:10]}"
    q_tmp = _pg_quote_ident(tmp_table)
    with conn.cursor() as cur:
        cur.execute(f"CREATE TEMP TABLE {q_tmp} (id_norm text PRIMARY KEY) ON COMMIT DROP;")
    buf = io.StringIO()
    buf.write("id_norm\n")
    for key in normalized_ids:
        buf.write(f"{key}\n")
    buf.seek(0)
    with conn.cursor() as cur:
        cur.copy_expert(
            f"COPY {q_tmp} (id_norm) FROM STDIN WITH (FORMAT csv, HEADER true);",
            buf,
        )

    unions: List[str] = []
    for col in sigte_cols:
        expr = _sql_expr_trim(col) if canon(col) == "sigteidnorm" else _sql_expr_norm_id(col)
        unions.append(
            "SELECT "
            + f"{expr} AS sigte_norm "
            + f"FROM {_pg_qualified_table(table_name)}"
        )
    union_sql = " UNION ALL ".join(unions)
    query = (
        "WITH src AS ("
        + union_sql
        + ") "
        "SELECT DISTINCT k.id_norm "
        f"FROM {q_tmp} k "
        "JOIN src ON src.sigte_norm = k.id_norm "
        "WHERE src.sigte_norm <> ''"
    )
    df = _sql_query_dataframe(conn, query, [])
    if df is None or df.empty:
        return set()
    out: set = set()
    for row in df.to_dict(orient="records"):
        key = normalize_id(row.get("id_norm"))
        if key:
            out.add(key)
    return out


def _sql_fetch_defunciones_lookup_for_ruts(
    conn: Any,
    rut_values: Iterable[str],
) -> Dict[str, date]:
    normalized_ruts: List[str] = []
    seen_ruts: set = set()
    for raw in rut_values:
        key = normalize_rut_compact(raw)
        if not key or key in seen_ruts:
            continue
        seen_ruts.add(key)
        normalized_ruts.append(key)
    if not normalized_ruts:
        return {}

    table_name = str(PG_BASE_TABLES.get("defunciones", "") or "").strip()
    if not table_name:
        return {}
    cols = _sql_fetch_table_columns(conn, table_name)
    if not cols:
        return {}

    col_rut = _sql_pick_column(cols, ["RUT_DV", "rut_dv", "RUTCONCATENADO", "rutconcatenado", "RUT", "rut"])
    col_run = _sql_pick_column(cols, ["RUN", "run"])
    col_dv = _sql_pick_column(cols, ["DV", "dv"])
    col_fecha = _sql_pick_column(cols, ["FECHA_DEF", "fecha_def", "FECHA DEF", "FECHA_DEFUNCION", "F_DEFUNCION"])
    if not col_rut and not (col_run and col_dv):
        return {}

    rut_expr_parts: List[str] = []
    if col_run and col_dv:
        q_run = _pg_quote_ident(col_run)
        q_dv = _pg_quote_ident(col_dv)
        run_expr = f"regexp_replace(trim(COALESCE({q_run}::text, '')), '[^0-9]', '', 'g')"
        dv_expr = f"upper(regexp_replace(trim(COALESCE({q_dv}::text, '')), '[^0-9Kk]', '', 'g'))"
        rut_expr_parts.append(
            f"CASE WHEN {run_expr} <> '' AND {dv_expr} <> '' THEN {run_expr} || left({dv_expr}, 1) ELSE '' END"
        )
    if col_rut:
        q_rut = _pg_quote_ident(col_rut)
        rut_expr_parts.append(
            f"upper(regexp_replace(trim(COALESCE({q_rut}::text, '')), '[^0-9Kk]', '', 'g'))"
        )
    if not rut_expr_parts:
        return {}
    if len(rut_expr_parts) == 1:
        rut_expr = rut_expr_parts[0]
    else:
        rut_expr = f"CASE WHEN {rut_expr_parts[0]} <> '' THEN {rut_expr_parts[0]} ELSE {rut_expr_parts[1]} END"

    fecha_expr = "NULL::date"
    if col_fecha:
        q_fecha = _pg_quote_ident(col_fecha)
        fecha_expr = (
            "CASE "
            f"WHEN {q_fecha} IS NULL OR trim(COALESCE({q_fecha}::text, '')) = '' THEN NULL::date "
            f"WHEN trim(COALESCE({q_fecha}::text, '')) ~ '^\\d+$' THEN (DATE '1899-12-30' + trim(COALESCE({q_fecha}::text, ''))::int)::date "
            f"WHEN trim(COALESCE({q_fecha}::text, '')) ~ '^\\d{{2}}/\\d{{2}}/\\d{{4}}$' THEN to_date(trim(COALESCE({q_fecha}::text, '')), 'DD/MM/YYYY') "
            f"WHEN trim(COALESCE({q_fecha}::text, '')) ~ '^\\d{{4}}-\\d{{2}}-\\d{{2}}$' THEN trim(COALESCE({q_fecha}::text, ''))::date "
            "ELSE NULL::date END"
        )

    tmp_table = f"tmp_def_ruts_{uuid.uuid4().hex[:10]}"
    q_tmp = _pg_quote_ident(tmp_table)
    with conn.cursor() as cur:
        cur.execute(f"CREATE TEMP TABLE {q_tmp} (rut_norm text PRIMARY KEY) ON COMMIT DROP;")
    buf = io.StringIO()
    buf.write("rut_norm\n")
    for key in normalized_ruts:
        buf.write(f"{key}\n")
    buf.seek(0)
    with conn.cursor() as cur:
        cur.copy_expert(
            f"COPY {q_tmp} (rut_norm) FROM STDIN WITH (FORMAT csv, HEADER true);",
            buf,
        )

    query = (
        "WITH src AS ("
        "SELECT "
        f"{rut_expr} AS rut_norm, "
        f"{fecha_expr} AS fecha_def "
        f"FROM {_pg_qualified_table(table_name)}"
        ") "
        "SELECT src.rut_norm, MAX(src.fecha_def) AS fecha_def "
        f"FROM src JOIN {q_tmp} t ON t.rut_norm = src.rut_norm "
        "WHERE src.rut_norm <> '' "
        "GROUP BY src.rut_norm"
    )
    df = _sql_query_dataframe(conn, query, [])
    if df is None or df.empty:
        return {}
    out: Dict[str, date] = {}
    for row in df.to_dict(orient="records"):
        key = normalize_rut_compact(row.get("rut_norm"))
        if not key:
            continue
        dt = parse_excel_date(row.get("fecha_def"))
        if dt is not None:
            out[key] = dt.date()
    return out


def _sql_fetch_patient_timeline_records(
    conn: Any,
    run: str,
    dv: str,
    presta: str,
    include_nomina: bool = True,
    include_historico: bool = False,
) -> List["TimelineRec"]:
    run_norm = normalize_run(run)
    dv_norm = normalize_dv(dv)
    presta_norm = normalize_presta(presta)
    if not (run_norm and dv_norm and presta_norm):
        return []

    table_names: List[str] = []
    if include_nomina:
        table_names.extend(_sql_nomina_table_names(conn))
    if include_historico:
        hist = _sql_historico_table_name(conn)
        if hist:
            table_names.append(hist)

    out: List["TimelineRec"] = []
    for table_name in table_names:
        cols = _sql_get_table_columns_cached(conn, table_name)
        if not cols:
            continue

        col_run = _sql_pick_column(cols, ["RUN", "run"])
        col_dv = _sql_pick_column(cols, ["DV", "dv"])
        col_presta = _sql_pick_column(cols, ["PRESTA_NORM", "presta_norm", "PRESTA_MIN", "presta_min"])
        col_fin = _sql_pick_column(cols, ["F_ENTRADA", "f_entrada"])
        col_fout = _sql_pick_column(cols, ["F_SALIDA", "f_salida"])
        col_sigte = _sql_pick_column(cols, ["SIGTE_ID_NORM", "sigte_id_norm", "SIGTE_ID", "sigte_id"])
        col_idlocal = _sql_pick_column(cols, ["ID_LOCAL_NORM", "id_local_norm", "ID_LOCAL", "id_local"])
        col_ext = _sql_pick_column(cols, ["EXT_NORM", "ext_norm", "EXTREMIDAD", "extremidad"])
        col_estab = _sql_pick_column(cols, ["ESTAB_DEST", "estab_dest"])
        if not (col_run and col_dv and col_presta):
            continue

        fin_expr = _sql_expr_date(col_fin) if col_fin else "NULL::date"
        fout_expr = _sql_expr_date(col_fout) if col_fout else "NULL::date"
        presta_expr = _sql_expr_trim(col_presta) if canon(col_presta) == "prestanorm" else _sql_expr_norm_presta(col_presta)
        sigte_expr = _sql_expr_trim(col_sigte) if canon(col_sigte) == "sigteidnorm" else (_sql_expr_norm_id(col_sigte) if col_sigte else "''::text")
        idlocal_expr = _sql_expr_trim(col_idlocal) if canon(col_idlocal) == "idlocalnorm" else (_sql_expr_norm_id(col_idlocal) if col_idlocal else "''::text")
        ext_expr = _sql_expr_trim(col_ext) if canon(col_ext) == "extnorm" else (_sql_expr_norm_text(col_ext) if col_ext else "''::text")

        query = (
            f"SELECT {fin_expr} AS f_in, "
            f"{fout_expr} AS f_out, "
            f"{sigte_expr} AS sigte_id, "
            f"{idlocal_expr} AS id_local, "
            f"{ext_expr} AS extremidad "
            f"FROM {_pg_qualified_table(table_name)} "
            f"WHERE {_sql_expr_norm_run(col_run)} = %s "
            f"AND {_sql_expr_norm_dv(col_dv)} = %s "
            f"AND {presta_expr} = %s "
        )
        params: List[Any] = [run_norm, dv_norm, presta_norm]
        if _sql_should_apply_estab_dest_filter(table_name) and col_estab:
            query += f"AND {_sql_expr_norm_id(col_estab)} = %s "
            params.append(str(ESTAB_DEST_FILTER))
        query += f"AND ({fin_expr} IS NOT NULL OR {fout_expr} IS NOT NULL)"

        with conn.cursor() as cur:
            cur.execute(query, params)
            rows = cur.fetchall()

        for f_in_raw, f_out_raw, sigte_raw, idlocal_raw, ext_raw in rows:
            f_in = parse_excel_date(f_in_raw)
            f_out = parse_excel_date(f_out_raw)
            sigte_id = normalize_id(sigte_raw)
            id_local = normalize_id(idlocal_raw)
            if not sigte_id:
                sigte_id = id_local
            out.append(
                TimelineRec(
                    f_in=f_in,
                    f_out=f_out,
                    sigte_id=sigte_id,
                    id_local=id_local,
                    source=table_name,
                    extremidad=normalize_text(ext_raw),
                )
            )

    out.sort(key=lambda r: (r.f_in or datetime.min))
    return out


def _sql_fetch_patient_timeline_records_cached(
    conn: Optional[Any],
    run: str,
    dv: str,
    presta: str,
    include_nomina: bool = True,
    include_historico: bool = False,
    cache: Optional[Dict[Tuple[str, str, str, bool, bool], List["TimelineRec"]]] = None,
) -> List["TimelineRec"]:
    if conn is None:
        return []
    run_norm = normalize_run(run)
    dv_norm = normalize_dv(dv)
    presta_norm = normalize_presta(presta)
    if not (run_norm and dv_norm and presta_norm):
        return []
    if cache is None:
        return _sql_fetch_patient_timeline_records(
            conn=conn,
            run=run_norm,
            dv=dv_norm,
            presta=presta_norm,
            include_nomina=include_nomina,
            include_historico=include_historico,
        )
    cache_key = (run_norm, dv_norm, presta_norm, bool(include_nomina), bool(include_historico))
    if cache_key in cache:
        return cache[cache_key]
    recs = _sql_fetch_patient_timeline_records(
        conn=conn,
        run=run_norm,
        dv=dv_norm,
        presta=presta_norm,
        include_nomina=include_nomina,
        include_historico=include_historico,
    )
    cache[cache_key] = recs
    return recs


def _sql_work_series(df: pd.DataFrame, col: Optional[str]) -> pd.Series:
    if not col or col not in df.columns:
        return pd.Series([""] * len(df), index=df.index)
    return df[col].fillna("").astype(str).map(lambda v: str(v).strip())


def _sql_build_missing_expr(
    alias: str,
    present_fields: List[Tuple[str, str]],
    contact_expr: str,
    include_yz: bool,
) -> str:
    if not present_fields:
        return "'Sin datos faltantes'"
    parts: List[str] = []
    for label, col_name in present_fields:
        parts.append(
            "CASE WHEN COALESCE(trim({a}.{c}), '') = '' THEN '{label}' END".format(
                a=alias,
                c=_pg_quote_ident(col_name),
                label=str(label).replace("'", "''"),
            )
        )
    if include_yz:
        parts.append(
            "CASE WHEN COALESCE(trim({a}.\"sospecha_diag\"), '') = '' "
            "AND COALESCE(trim({a}.\"confir_diag\"), '') = '' THEN 'Info en Y o Z' END".format(
                a=alias
            )
        )
    parts.append(
        "CASE WHEN ({contact}) <> 'Posee datos de contacto' THEN 'Medio de Contacto' END".format(
            contact=contact_expr
        )
    )
    arr = "ARRAY[{parts}]".format(parts=", ".join(parts))
    clean = "array_remove({arr}, NULL)".format(arr=arr)
    return (
        "CASE WHEN cardinality({clean}) = 0 THEN 'Sin datos faltantes' "
        "ELSE 'Falta: ' || array_to_string({clean}, ', ') END".format(clean=clean)
    )


def _sql_build_timeline_union_query(conn: Any, table_names: List[str]) -> str:
    unions: List[str] = []
    for table_name in table_names:
        cols = _sql_fetch_table_columns(conn, table_name)
        if not cols:
            continue
        col_run = _sql_pick_column(cols, ["RUN", "run"])
        col_dv = _sql_pick_column(cols, ["DV", "dv"])
        col_presta = _sql_pick_column(cols, ["PRESTA_NORM", "presta_norm", "PRESTA_MIN", "presta_min"])
        col_fin = _sql_pick_column(cols, ["F_ENTRADA", "f_entrada"])
        col_fout = _sql_pick_column(cols, ["F_SALIDA", "f_salida"])
        col_sigte = _sql_pick_column(cols, ["SIGTE_ID_NORM", "sigte_id_norm", "SIGTE_ID", "sigte_id"])
        col_idlocal = _sql_pick_column(cols, ["ID_LOCAL_NORM", "id_local_norm", "ID_LOCAL", "id_local"])
        col_ext = _sql_pick_column(cols, ["EXT_NORM", "ext_norm", "EXTREMIDAD", "extremidad"])
        col_estab = _sql_pick_column(cols, ["ESTAB_DEST", "estab_dest"])
        if not (col_run and col_dv and col_presta):
            continue
        fin_expr = _sql_expr_date(col_fin) if col_fin else "NULL::date"
        fout_expr = _sql_expr_date(col_fout) if col_fout else "NULL::date"
        id_expr = _sql_expr_trim(col_idlocal) if canon(col_idlocal) == "idlocalnorm" else (_sql_expr_norm_id(col_idlocal) if col_idlocal else "''::text")
        sigte_expr = _sql_expr_trim(col_sigte) if canon(col_sigte) == "sigteidnorm" else (_sql_expr_norm_id(col_sigte) if col_sigte else "''::text")
        ext_expr = _sql_expr_trim(col_ext) if canon(col_ext) == "extnorm" else (_sql_expr_norm_text(col_ext) if col_ext else "''::text")
        run_expr = _sql_expr_norm_run(col_run)
        dv_expr = _sql_expr_norm_dv(col_dv)
        presta_expr = _sql_expr_trim(col_presta) if canon(col_presta) == "prestanorm" else _sql_expr_norm_presta(col_presta)
        q = (
            "SELECT "
            f"{run_expr} AS run_norm, "
            f"{dv_expr} AS dv_norm, "
            f"{presta_expr} AS presta_norm, "
            f"{fin_expr} AS f_in, "
            f"{fout_expr} AS f_out, "
            f"{id_expr} AS id_local_norm, "
            f"{sigte_expr} AS sigte_id_norm, "
            f"{ext_expr} AS ext_norm "
            f"FROM {_pg_qualified_table(table_name)} "
            f"WHERE ({fin_expr} IS NOT NULL OR {fout_expr} IS NOT NULL)"
        )
        if _sql_should_apply_estab_dest_filter(table_name) and col_estab:
            estab_literal = str(ESTAB_DEST_FILTER).replace("'", "''")
            q += f" AND {_sql_expr_norm_id(col_estab)} = '{estab_literal}'"
        unions.append(q)
    if not unions:
        return (
            "SELECT ''::text AS run_norm, ''::text AS dv_norm, ''::text AS presta_norm, "
            "NULL::date AS f_in, NULL::date AS f_out, ''::text AS id_local_norm, "
            "''::text AS sigte_id_norm, ''::text AS ext_norm WHERE 1=0"
        )
    return " UNION ALL ".join(unions)


def _sql_bulk_enrich_work_df(
    conn: Any,
    work_df: pd.DataFrame,
    selected: Dict[str, bool],
    ingreso_fields_present: List[Tuple[str, str]],
    egreso_fields_present: List[Tuple[str, str]],
    contact_columns_present: List[str],
    include_yz: bool = False,
) -> pd.DataFrame:
    tmp_table = f"tmp_work_enrich_{uuid.uuid4().hex[:10]}"
    q_tmp = _pg_quote_ident(tmp_table)
    create_sql = (
        f"CREATE TEMP TABLE {q_tmp} ("
        "row_id integer PRIMARY KEY, "
        "run_norm text, dv_norm text, presta_norm text, "
        "tipo_norm text, plano_norm text, ext_norm text, "
        "f_in date, f_out date, "
        "estab_norm text, estab_codigo text, id_local_norm text, rut_norm text, "
        "serv_salud text, run_txt text, dv_txt text, nombres text, primer_apellido text, segundo_apellido text, "
        "fecha_nac text, sexo text, prevision text, presta_est text, estab_orig text, estab_dest text, "
        "c_salida text, e_otor_at text, presta_min_salida text, prais text, region text, comuna text, ciudad text, "
        "cond_ruralidad text, nom_calle text, run_prof_sol text, dv_prof_sol text, run_prof_resol text, dv_prof_resol text, "
        "sospecha_diag text, confir_diag text, "
        "phone_a text, phone_b text, phone_c text, phone_d text, phone_e text, "
        "email_a text, email_b text, email_c text"
        ") ON COMMIT DROP;"
    )
    with conn.cursor() as cur:
        cur.execute(create_sql)

    buf = io.StringIO()
    work_df.to_csv(buf, index=False, header=True, sep="\t", na_rep="", lineterminator="\n")
    buf.seek(0)
    cols = ",".join(_pg_quote_ident(c) for c in work_df.columns)
    with conn.cursor() as cur:
        cur.copy_expert(
            f"COPY {q_tmp} ({cols}) FROM STDIN WITH (FORMAT csv, HEADER true, DELIMITER E'\\t', NULL '');",
            buf,
        )
    if len(work_df) >= 15000:
        idx_keys = _pg_quote_ident(f"{tmp_table}_keys")
        idx_id = _pg_quote_ident(f"{tmp_table}_id")
        idx_rut = _pg_quote_ident(f"{tmp_table}_rut")
        with conn.cursor() as cur:
            cur.execute(f"CREATE INDEX {idx_keys} ON {q_tmp} (run_norm, dv_norm, presta_norm);")
            cur.execute(f"CREATE INDEX {idx_id} ON {q_tmp} (id_local_norm);")
            cur.execute(f"CREATE INDEX {idx_rut} ON {q_tmp} (rut_norm);")
            cur.execute(f"ANALYZE {q_tmp};")

    contact_parts: List[str] = []
    for col in contact_columns_present:
        contact_parts.append(f"COALESCE(trim(w.{_pg_quote_ident(col)}), '') = ''")
    if contact_parts:
        contact_expr = (
            "CASE WHEN {all_blank} THEN 'Sin datos de contacto' ELSE 'Posee datos de contacto' END".format(
                all_blank=" AND ".join(contact_parts)
            )
        )
    else:
        contact_expr = "'Sin datos de contacto'"

    falt_ing_expr = _sql_build_missing_expr("w", ingreso_fields_present, contact_expr, include_yz=include_yz)
    falt_egr_expr = _sql_build_missing_expr("w", egreso_fields_present, contact_expr, include_yz=include_yz)

    nom_union = _sql_build_timeline_union_query(conn, _sql_nomina_table_names(conn))
    hist_name = _sql_historico_table_name(conn)
    hist_union = _sql_build_timeline_union_query(conn, [hist_name] if hist_name else [])

    q_cgr = _pg_qualified_table(PG_BASE_TABLES["cgr"])
    q_def = _pg_qualified_table(PG_BASE_TABLES["defunciones"])
    q_estab = _pg_qualified_table(PG_BASE_TABLES["establecimientos"])

    sql = f"""
WITH w AS (
    SELECT
      *,
      COALESCE(LEAST(f_in, f_out), f_in, f_out) AS r_start,
      CASE WHEN f_in IS NOT NULL AND f_out IS NULL THEN NULL::date
           ELSE COALESCE(GREATEST(f_in, f_out), f_in, f_out) END AS r_end
    FROM {q_tmp}
),
w_keys AS (
    SELECT DISTINCT run_norm, dv_norm, presta_norm
    FROM w
    WHERE run_norm <> '' AND dv_norm <> '' AND presta_norm <> ''
),
w_ids AS (
    SELECT DISTINCT id_local_norm
    FROM w
    WHERE id_local_norm <> ''
),
w_ruts AS (
    SELECT DISTINCT rut_norm
    FROM w
    WHERE rut_norm <> ''
),
nom AS (
    SELECT
      n.*,
      COALESCE(LEAST(n.f_in, n.f_out), n.f_in, n.f_out) AS r_start,
      CASE WHEN n.f_in IS NOT NULL AND n.f_out IS NULL THEN NULL::date
           ELSE COALESCE(GREATEST(n.f_in, n.f_out), n.f_in, n.f_out) END AS r_end
    FROM ({nom_union}) n
    JOIN w_keys k
      ON k.run_norm = n.run_norm
     AND k.dv_norm = n.dv_norm
     AND k.presta_norm = n.presta_norm
),
hist AS (
    SELECT
      h.*,
      COALESCE(LEAST(h.f_in, h.f_out), h.f_in, h.f_out) AS r_start,
      CASE WHEN h.f_in IS NOT NULL AND h.f_out IS NULL THEN NULL::date
           ELSE COALESCE(GREATEST(h.f_in, h.f_out), h.f_in, h.f_out) END AS r_end
    FROM ({hist_union}) h
    JOIN w_keys k
      ON k.run_norm = h.run_norm
     AND k.dv_norm = h.dv_norm
     AND k.presta_norm = h.presta_norm
),
defu AS (
    SELECT
      upper(regexp_replace(trim(COALESCE(rut_dv, '')), '[^0-9Kk]', '', 'g')) AS rut_norm,
      CASE
        WHEN fecha_def IS NULL OR trim(fecha_def) = '' THEN NULL::date
        WHEN trim(fecha_def) ~ '^\\d+$' THEN (DATE '1899-12-30' + trim(fecha_def)::int)::date
        WHEN trim(fecha_def) ~ '^\\d{{2}}/\\d{{2}}/\\d{{4}}$' THEN to_date(trim(fecha_def), 'DD/MM/YYYY')
        WHEN trim(fecha_def) ~ '^\\d{{4}}-\\d{{2}}-\\d{{2}}$' THEN trim(fecha_def)::date
        ELSE NULL::date
      END AS fecha_def
    FROM {q_def}
    WHERE upper(regexp_replace(trim(COALESCE(rut_dv, '')), '[^0-9Kk]', '', 'g')) IN (
      SELECT rut_norm FROM w_ruts
    )
),
cgr AS (
    SELECT
      regexp_replace(trim(COALESCE(id_local, '')), '\\.0+$', '') AS id_local_norm,
      trim(COALESCE(origen, '')) AS origen
    FROM {q_cgr}
    WHERE regexp_replace(trim(COALESCE(id_local, '')), '\\.0+$', '') IN (
      SELECT id_local_norm FROM w_ids
    )
),
estab AS (
    SELECT trim(COALESCE(codigo, '')) AS codigo_norm
    FROM {q_estab}
)
SELECT
  w.row_id,
  {contact_expr} AS contacto,
  {falt_ing_expr} AS falt_ingreso,
  {falt_egr_expr} AS falt_egreso,
  CASE
    WHEN w.run_norm = '' OR w.dv_norm = '' OR w.presta_norm = '' THEN 'Sin evaluacion (faltan datos clave)'
    WHEN w.f_in IS NULL AND w.f_out IS NULL THEN 'Sin evaluacion (faltan fechas)'
    WHEN near_row.min_diff IS NOT NULL AND near_row.min_diff <= 365
      THEN 'Alerta: Caso cercano en fechas del registro (' || COALESCE(NULLIF(near_row.near_id, ''), 'sin ID') || ')'
    ELSE 'Sin alerta'
  END AS alerta_cercano,
  CASE
    WHEN w.run_norm = '' OR w.dv_norm = '' OR w.presta_norm = '' OR w.f_in IS NULL THEN 'Sin evaluacion (faltan datos clave)'
    WHEN tras_row.exact_exists THEN 'Sin traslape (caso existe en nominas)'
    WHEN NOT tras_row.any_exists THEN 'Sin traslape (sin registros en nominas de la especialidad)'
    WHEN tras_row.future_closed_id <> '' THEN 'Caso traslape, traslape con SIGTE_ID: ' || tras_row.future_closed_id
    ELSE 'Sin traslape (Ultimo caso registrado de la especialidad)'
  END AS traslape,
  CASE
    WHEN w.run_norm = '' OR w.dv_norm = '' OR w.presta_norm = '' THEN 'Sin evaluacion (faltan datos clave)'
    WHEN w.r_start IS NULL THEN 'Sin evaluacion (faltan fechas)'
    WHEN COALESCE(NULLIF(dup_ext.other_id, ''), NULLIF(dup_int.other_id, '')) IS NOT NULL
      THEN 'Caso duplicado, duplicidad con ID_LOCAL: ' || COALESCE(NULLIF(dup_ext.other_id, ''), dup_int.other_id)
    ELSE 'Sin duplicidad'
  END AS duplicidad,
  CASE WHEN hist_hit.found_hist THEN 'Se encuentra en historico' ELSE 'No se encuentra en historico' END AS cruce_historico,
  CASE WHEN cgr_hit.has_399 THEN cgr_hit.origen ELSE 'No se encuentra en CGR 399' END AS cruce_cgr_399,
  CASE WHEN cgr_hit.has_84 AND NOT cgr_hit.has_399 THEN cgr_hit.origen ELSE 'No se encuentra en CGR 84' END AS cruce_cgr_84,
  CASE WHEN def_hit.fecha_def IS NOT NULL THEN 'Paciente fallecido' ELSE 'Paciente vivo' END AS cruce_defunciones,
  COALESCE(to_char(def_hit.fecha_def, 'DD/MM/YYYY'), '') AS fecha_defuncion,
  CASE
    WHEN def_hit.fecha_def IS NOT NULL AND w.f_out IS NOT NULL AND w.f_out > def_hit.fecha_def
      THEN 'Alerta: paciente con egreso posterior a la fecha de fallecimiento'
    ELSE 'Sin alertas'
  END AS alerta_fallecimiento,
  CASE
    WHEN def_hit.fecha_def IS NOT NULL
      AND ((w.f_in IS NOT NULL AND def_hit.fecha_def < w.f_in) OR (w.f_out IS NOT NULL AND def_hit.fecha_def < w.f_out))
      THEN 'Alerta: fecha de defuncion anterior a F_ENTRADA/F_SALIDA'
    ELSE 'Sin alertas'
  END AS alerta_fecha_defuncion,
  CASE
    WHEN w.estab_codigo = '' THEN 'Macro red'
    WHEN EXISTS (SELECT 1 FROM estab e WHERE e.codigo_norm = w.estab_codigo) THEN 'Corresponde establecimiento'
    ELSE 'Macro red'
  END AS cruce_establecimientos
FROM w
LEFT JOIN LATERAL (
  SELECT
    MIN(ABS((wd.d - nd.d)))::int AS min_diff,
    (
      ARRAY_AGG(COALESCE(n.id_local_norm, n.sigte_id_norm, '') ORDER BY ABS((wd.d - nd.d)), COALESCE(n.id_local_norm, n.sigte_id_norm, ''))
    )[1] AS near_id
  FROM nom n
  JOIN LATERAL (VALUES (w.f_in), (w.f_out)) wd(d) ON wd.d IS NOT NULL
  JOIN LATERAL (VALUES (n.f_in), (n.f_out)) nd(d) ON nd.d IS NOT NULL
  WHERE n.run_norm = w.run_norm
    AND n.dv_norm = w.dv_norm
    AND n.presta_norm = w.presta_norm
    AND (w.id_local_norm = '' OR COALESCE(n.id_local_norm, n.sigte_id_norm, '') <> w.id_local_norm)
) near_row ON TRUE
LEFT JOIN LATERAL (
  SELECT
    EXISTS (
      SELECT 1 FROM nom n
      WHERE n.run_norm = w.run_norm
        AND n.dv_norm = w.dv_norm
        AND n.presta_norm = w.presta_norm
        AND n.f_in = w.f_in
    ) AS exact_exists,
    EXISTS (
      SELECT 1 FROM nom n
      WHERE n.run_norm = w.run_norm
        AND n.dv_norm = w.dv_norm
        AND n.presta_norm = w.presta_norm
    ) AS any_exists,
    COALESCE((
      SELECT COALESCE(n.sigte_id_norm, n.id_local_norm, '')
      FROM nom n
      WHERE n.run_norm = w.run_norm
        AND n.dv_norm = w.dv_norm
        AND n.presta_norm = w.presta_norm
        AND w.f_in IS NOT NULL
        AND n.f_in > w.f_in
        AND n.f_out IS NOT NULL
      ORDER BY n.f_in DESC
      LIMIT 1
    ), '') AS future_closed_id
) tras_row ON TRUE
LEFT JOIN LATERAL (
  SELECT COALESCE(n.id_local_norm, n.sigte_id_norm, '') AS other_id
  FROM (
    SELECT run_norm, dv_norm, presta_norm, ext_norm, r_start, r_end, id_local_norm, sigte_id_norm FROM nom
    UNION ALL
    SELECT run_norm, dv_norm, presta_norm, ext_norm, r_start, r_end, id_local_norm, sigte_id_norm FROM hist
  ) n
  WHERE n.run_norm = w.run_norm
    AND n.dv_norm = w.dv_norm
    AND n.presta_norm = w.presta_norm
    AND (w.ext_norm = '' OR n.ext_norm = '' OR w.ext_norm = n.ext_norm)
    AND w.r_start IS NOT NULL
    AND n.r_start IS NOT NULL
    AND w.r_start <= COALESCE(n.r_end, DATE '9999-12-31')
    AND n.r_start <= COALESCE(w.r_end, DATE '9999-12-31')
    AND (
      COALESCE(n.id_local_norm, n.sigte_id_norm, '') = ''
      OR w.id_local_norm = ''
      OR COALESCE(n.id_local_norm, n.sigte_id_norm, '') <> w.id_local_norm
    )
  ORDER BY COALESCE(n.id_local_norm, n.sigte_id_norm, '')
  LIMIT 1
) dup_ext ON TRUE
LEFT JOIN LATERAL (
  SELECT COALESCE(w2.id_local_norm, '') AS other_id
  FROM w w2
  WHERE w2.row_id < w.row_id
    AND w2.run_norm = w.run_norm
    AND w2.dv_norm = w.dv_norm
    AND w2.presta_norm = w.presta_norm
    AND (w.ext_norm = '' OR w2.ext_norm = '' OR w.ext_norm = w2.ext_norm)
    AND w.r_start IS NOT NULL
    AND w2.r_start IS NOT NULL
    AND w.r_start <= COALESCE(w2.r_end, DATE '9999-12-31')
    AND w2.r_start <= COALESCE(w.r_end, DATE '9999-12-31')
    AND (w.id_local_norm = '' OR w2.id_local_norm = '' OR w2.id_local_norm <> w.id_local_norm)
  ORDER BY w2.row_id DESC
  LIMIT 1
) dup_int ON TRUE
LEFT JOIN LATERAL (
  SELECT EXISTS (
    SELECT 1
    FROM hist h
    WHERE h.run_norm = w.run_norm
      AND h.dv_norm = w.dv_norm
      AND h.presta_norm = w.presta_norm
      AND (
        (w.id_local_norm <> '' AND COALESCE(h.id_local_norm, h.sigte_id_norm, '') = w.id_local_norm)
        OR (w.f_in IS NOT NULL AND h.f_in = w.f_in)
      )
  ) AS found_hist
) hist_hit ON TRUE
LEFT JOIN LATERAL (
  SELECT
    COALESCE(c.origen, '') AS origen,
    (regexp_replace(lower(COALESCE(c.origen, '')), '[^a-z0-9]', '', 'g') LIKE '%399%') AS has_399,
    (regexp_replace(lower(COALESCE(c.origen, '')), '[^a-z0-9]', '', 'g') LIKE '%84%') AS has_84
  FROM cgr c
  WHERE c.id_local_norm = w.id_local_norm
  LIMIT 1
) cgr_hit ON TRUE
LEFT JOIN LATERAL (
  SELECT d.fecha_def
  FROM defu d
  WHERE d.rut_norm = w.rut_norm
  ORDER BY d.fecha_def DESC NULLS LAST
  LIMIT 1
) def_hit ON TRUE
ORDER BY w.row_id
"""
    return _sql_query_dataframe(conn, sql, [])


def _admin_norm_col(value: Any) -> str:
    s = str(value if value is not None else "").strip().lower()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    return s


_ADMIN_DEFAULT_FILE_EXTS = {".csv"}
_ADMIN_TABLE_FILE_EXTS: Dict[str, set] = {
    "comges_especiales": {".csv", ".xlsx", ".xlsb", ".xlsm"},
    "defunciones": {".csv", ".xlsx", ".xlsb", ".xlsm"},
}
_ADMIN_COMGES_COLUMN_ALIASES: Dict[str, List[str]] = {
    "sigte_id": ["sigte_id", "sigte_id2", "sigte_id.1", "sigte_id22"],
    "run": ["run", "run_persona"],
    "dv": ["dv", "dv_persona"],
    "sexo": ["sexo", "genero"],
    "estab_dest": ["estab_dest", "establecimiento de destino", "estab destino"],
    "f_entrada": ["f_entrada", "fechaingreso", "fecha_entrada", "fentrada"],
    "f_salida": ["f_salida", "fecha_salida", "fsalida"],
    "c_salida": ["c_salida", "csalida", "codigo_salida"],
}


def _admin_allowed_extensions_for_table(table_name: str) -> List[str]:
    key = str(table_name or "").strip().lower()
    allowed = _ADMIN_TABLE_FILE_EXTS.get(key)
    if not allowed:
        allowed = _ADMIN_DEFAULT_FILE_EXTS
    return sorted({str(ext).strip().lower() for ext in allowed if str(ext).strip()})


def _admin_allowed_ext_text_for_table(table_name: str) -> str:
    return ", ".join(_admin_allowed_extensions_for_table(table_name))


def _admin_accept_attr_for_table(table_name: str) -> str:
    exts = _admin_allowed_extensions_for_table(table_name)
    mime_map = {
        ".csv": "text/csv",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
        ".xlsb": "application/vnd.ms-excel.sheet.binary.macroEnabled.12",
    }
    parts: List[str] = []
    for ext in exts:
        parts.append(ext)
        mime = mime_map.get(ext)
        if mime:
            parts.append(mime)
    return ",".join(parts)


def _admin_is_allowed_file_for_table(table_name: str, filename: str) -> bool:
    suffix = Path(filename or "").suffix.lower()
    if not suffix:
        return False
    return suffix in set(_admin_allowed_extensions_for_table(table_name))


def _admin_is_csv_file(filename: str) -> bool:
    return Path(filename).suffix.lower() == ".csv"


def _admin_credentials() -> Dict[str, str]:
    return {}


def _is_admin_user(rut: Optional[str]) -> bool:
    target = str(rut or "").strip().upper()
    if not target:
        return False
    return bool(_auth_is_admin_user(target))


def _read_text_fallback(path: Path) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return path.read_text(encoding=enc)
        except Exception:
            continue
    raise RuntimeError(f"No se pudo leer archivo de texto: {path}")


def _admin_get_table_columns(conn: Any, table_name: str) -> List[str]:
    schema = PG_SCHEMA if PG_SCHEMA else "public"
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema=%s AND table_name=%s
            ORDER BY ordinal_position;
            """,
            (schema, table_name),
        )
        cols = [str(r[0]) for r in cur.fetchall()]
    if not cols:
        raise RuntimeError(f"No se encontraron columnas para {schema}.{table_name}.")
    return cols


def _admin_best_read_csv(
    path: Path,
    expected_cols_norm: List[str],
) -> Tuple[pd.DataFrame, int, int, str, str]:
    expected_set = {c for c in expected_cols_norm if c}
    best_df: Optional[pd.DataFrame] = None
    best_header = 0
    best_score = -1
    best_delim = ","
    best_encoding = "utf-8-sig"

    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as fh:
                sample = fh.read(65536)
        except UnicodeDecodeError:
            continue
        except Exception:
            continue
        delim = _guess_csv_delimiter(sample)
        for header in (0, 1, 2, 3, 4):
            try:
                df = pd.read_csv(
                    path,
                    header=header,
                    dtype=object,
                    sep=delim,
                    engine="python",
                    encoding=encoding,
                    on_bad_lines="skip",
                )
            except Exception:
                continue
            cols_norm = [_admin_norm_col(c) for c in list(df.columns)]
            score = len({c for c in cols_norm if c} & expected_set)
            if score > best_score:
                best_df = df
                best_header = header
                best_score = score
                best_delim = delim
                best_encoding = encoding

    if best_df is None:
        try:
            with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
                sample = fh.read(65536)
            delim = _guess_csv_delimiter(sample)
            for header in (0, 1, 2, 3, 4):
                try:
                    df = pd.read_csv(
                        path,
                        header=header,
                        dtype=object,
                        sep=delim,
                        engine="python",
                        encoding="utf-8-sig",
                        on_bad_lines="skip",
                    )
                except Exception:
                    continue
                cols_norm = [_admin_norm_col(c) for c in list(df.columns)]
                score = len({c for c in cols_norm if c} & expected_set)
                if score > best_score:
                    best_df = df
                    best_header = header
                    best_score = score
                    best_delim = delim
                    best_encoding = "utf-8-sig"
        except Exception:
            pass

    if best_df is None:
        raise RuntimeError(f"No se pudo leer {path.name} como CSV con encabezados 0..4.")
    return best_df, best_header, max(best_score, 0), best_delim, best_encoding


def _admin_align_df(df: pd.DataFrame, table_cols: List[str]) -> pd.DataFrame:
    expected_norm = [_admin_norm_col(c) for c in table_cols]
    norm_to_idx: Dict[str, int] = {}
    for idx, col in enumerate(list(df.columns)):
        key = _admin_norm_col(col)
        if key and key not in norm_to_idx:
            norm_to_idx[key] = idx

    aligned = pd.DataFrame()
    for col_name, norm_name in zip(table_cols, expected_norm):
        idx = norm_to_idx.get(norm_name)
        if idx is None:
            aligned[col_name] = None
        else:
            aligned[col_name] = df.iloc[:, idx]

    def _to_text_or_none(v: Any) -> Optional[str]:
        try:
            if pd.isna(v):
                return None
        except Exception:
            pass
        return str(v).strip()

    for col_name in aligned.columns:
        aligned[col_name] = aligned[col_name].map(_to_text_or_none)
    return aligned


def _admin_read_csv_for_table(
    path: Path,
    table_cols: List[str],
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    expected_cols_norm = [_admin_norm_col(c) for c in table_cols]
    summary: List[Dict[str, Any]] = []
    df_raw, header_used, score, delim, encoding = _admin_best_read_csv(path, expected_cols_norm)
    df_aligned = _admin_align_df(df_raw, table_cols)
    summary.append(
        {
            "hoja": "CSV",
            "header": int(header_used),
            "match": int(score),
            "filas": int(len(df_aligned)),
            "delimitador": delim,
            "encoding": encoding,
        }
    )
    return df_aligned, summary


def _admin_pick_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    norm_to_real: Dict[str, str] = {}
    for col in list(df.columns):
        key = _admin_norm_col(col)
        if key and key not in norm_to_real:
            norm_to_real[key] = str(col)
    for cand in candidates:
        key = _admin_norm_col(cand)
        if key in norm_to_real:
            return norm_to_real[key]
    for cand in candidates:
        key = _admin_norm_col(cand)
        if not key:
            continue
        for norm_name, real in norm_to_real.items():
            if key in norm_name:
                return real
    return None


def _admin_best_read_excel_sheet(
    path: Path,
    sheet_name: str,
    expected_cols_norm: List[str],
) -> Tuple[pd.DataFrame, int, int]:
    expected_set = {c for c in expected_cols_norm if c}
    best_df: Optional[pd.DataFrame] = None
    best_header = 0
    best_score = -1
    engine = "pyxlsb" if path.suffix.lower() == ".xlsb" else "openpyxl"

    for header in (0, 1, 2, 3, 4):
        try:
            df = pd.read_excel(path, sheet_name=sheet_name, header=header, dtype=object, engine=engine)
        except Exception:
            continue
        cols_norm = [_admin_norm_col(c) for c in list(df.columns)]
        score = len({c for c in cols_norm if c} & expected_set)
        if score > best_score:
            best_df = df
            best_header = header
            best_score = score
    if best_df is None:
        raise RuntimeError(f"No se pudo leer la hoja '{sheet_name}' de {path.name}.")
    return best_df, best_header, max(best_score, 0)


def _admin_project_comges_sheet(df_raw: pd.DataFrame) -> pd.DataFrame:
    projected = pd.DataFrame(index=df_raw.index)
    for target_col, aliases in _ADMIN_COMGES_COLUMN_ALIASES.items():
        src_col = _admin_pick_column(df_raw, aliases)
        if src_col and src_col in df_raw.columns:
            projected[target_col] = df_raw[src_col]
        else:
            projected[target_col] = None
    return projected


def _admin_read_excel_for_table(
    path: Path,
    table_name: str,
    table_cols: List[str],
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    table_key = str(table_name or "").strip().lower()
    is_comges = table_key == "comges_especiales"
    expected_cols_norm = [_admin_norm_col(c) for c in table_cols]
    if is_comges:
        for aliases in _ADMIN_COMGES_COLUMN_ALIASES.values():
            expected_cols_norm.extend(_admin_norm_col(c) for c in aliases)

    engine = "pyxlsb" if path.suffix.lower() == ".xlsb" else "openpyxl"
    xls = pd.ExcelFile(path, engine=engine)
    if not xls.sheet_names:
        raise RuntimeError(f"El archivo {path.name} no contiene hojas.")

    summary: List[Dict[str, Any]] = []
    frames: List[pd.DataFrame] = []
    for sheet_name in xls.sheet_names:
        df_raw, header_used, score = _admin_best_read_excel_sheet(path, sheet_name, expected_cols_norm)
        if is_comges:
            df_projected = _admin_project_comges_sheet(df_raw)
        else:
            df_projected = df_raw
        df_aligned = _admin_align_df(df_projected, table_cols)
        if not df_aligned.empty:
            not_empty = (
                df_aligned.fillna("")
                .astype(str)
                .apply(lambda s: s.str.strip())
                .ne("")
                .any(axis=1)
            )
            df_aligned = df_aligned.loc[not_empty].copy()
        if is_comges:
            col_sigte = _admin_pick_column(df_aligned, ["sigte_id"])
            if col_sigte and col_sigte in df_aligned.columns:
                sigte_series = df_aligned[col_sigte].fillna("").astype(str).str.strip()
                df_aligned = df_aligned.loc[sigte_series != ""].copy()
        summary.append(
            {
                "hoja": str(sheet_name),
                "header": int(header_used),
                "match": int(score),
                "filas": int(len(df_aligned)),
                "delimitador": "",
                "encoding": engine,
            }
        )
        if not df_aligned.empty:
            frames.append(df_aligned)

    if not frames:
        return pd.DataFrame(columns=table_cols), summary
    df_final = pd.concat(frames, ignore_index=True, sort=False)
    return df_final, summary


def _admin_read_input_for_table(
    path: Path,
    table_name: str,
    table_cols: List[str],
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _admin_read_csv_for_table(path, table_cols)
    if suffix in {".xlsx", ".xlsm", ".xlsb"}:
        return _admin_read_excel_for_table(path, table_name, table_cols)
    raise RuntimeError(
        f"Formato no soportado para {table_name}: {suffix}. Permitidos: {_admin_allowed_ext_text_for_table(table_name)}"
    )


def _admin_count_table(conn: Any, table_name: str) -> int:
    with conn.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM {_pg_qualified_table(table_name)};")
        row = cur.fetchone()
    return int((row[0] if row else 0) or 0)


def _admin_table_allows_append(table_name: str) -> bool:
    key = str(table_name or "").strip().lower()
    return bool(key)


def _admin_resolve_load_mode(table_name: str, mode: str) -> str:
    key = str(mode or "").strip().lower()
    if key == "append" and _admin_table_allows_append(table_name):
        return "append"
    return "replace"


def _admin_copy_with_mode(conn: Any, table_name: str, df: pd.DataFrame, mode: str) -> None:
    resolved_mode = _admin_resolve_load_mode(table_name, mode)
    with conn.cursor() as cur:
        if resolved_mode == "replace":
            cur.execute(f"TRUNCATE TABLE {_pg_qualified_table(table_name)};")
        if df.empty:
            return
        buffer = io.StringIO()
        df.to_csv(buffer, index=False, header=True, sep="\t", na_rep="", lineterminator="\n")
        buffer.seek(0)
        col_list = ",".join(_pg_quote_ident(c) for c in df.columns)
        copy_sql = (
            f"COPY {_pg_qualified_table(table_name)} ({col_list}) "
            "FROM STDIN WITH (FORMAT csv, HEADER true, DELIMITER E'\\t', NULL '');"
        )
        cur.copy_expert(copy_sql, buffer)


def _admin_should_refresh_cores(selected_tables: Iterable[str]) -> bool:
    for table_name in selected_tables:
        if table_name in ADMIN_CORES_TRIGGER_TABLES:
            return True
    return False


def _admin_run_cores_and_indices(conn: Any) -> None:
    if not ADMIN_CORES_SQL_PATH.exists() or not ADMIN_CORES_SQL_PATH.is_file():
        raise RuntimeError(f"No existe script SQL de cores/indices: {ADMIN_CORES_SQL_PATH}")
    sql_script = _read_text_fallback(ADMIN_CORES_SQL_PATH)
    with conn.cursor() as cur:
        cur.execute(sql_script)


def _pg_table_from_pseudo_path(path: Path) -> str:
    stem = str(path.stem)
    if "." in stem:
        return stem.split(".", 1)[1]
    return stem


def _guess_csv_delimiter(sample: str) -> str:
    if not sample:
        return ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return str(dialect.delimiter or ",")
    except Exception:
        counts = {d: sample.count(d) for d in [",", ";", "\t", "|"]}
        delim, hits = max(counts.items(), key=lambda x: x[1])
        return delim if hits > 0 else ","


def _row_value(row: Tuple[Any, ...], idx: int) -> Any:
    if not idx:
        return None
    pos = idx - 1
    if pos < 0 or pos >= len(row):
        return None
    return row[pos]


def _read_csv_raw_df(path: Path) -> pd.DataFrame:
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as fh:
                sample = fh.read(65536)
            delim = _guess_csv_delimiter(sample)
            return pd.read_csv(
                path,
                header=None,
                dtype=object,
                sep=delim,
                engine="python",
                encoding=encoding,
                on_bad_lines="skip",
            )
        except UnicodeDecodeError:
            continue
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        sample = fh.read(65536)
    delim = _guess_csv_delimiter(sample)
    return pd.read_csv(
        path,
        header=None,
        dtype=object,
        sep=delim,
        engine="python",
        encoding="utf-8-sig",
        on_bad_lines="skip",
    )


def _parse_nomina_filename(path: Path) -> Optional[Tuple[str, Optional[str], bool]]:
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xlsb", ".csv", ".sql"}:
        return None
    stem = canon(path.stem)
    m = NOMINA_NAME_RE.match(stem)
    if m:
        tipo_raw = m.group("tipo")
        estado_raw = m.group("estado")
        tipo = NOMINA_TYPE_ALIASES.get(tipo_raw, tipo_raw)
        estado = NOMINA_STATE_ALIASES.get(estado_raw, estado_raw)
        return tipo, estado, True
    if not stem.startswith("nomina"):
        return None
    tipo = None
    for cand in ("proc", "iq", "cne", "ic"):
        if cand in stem:
            tipo = NOMINA_TYPE_ALIASES.get(cand, cand)
            break
    if not tipo:
        return None
    return tipo, None, False


def _sheet_matches_estado(sheet_name: str, estado: str) -> bool:
    s = canon(sheet_name)
    for alias, normalized in NOMINA_STATE_ALIASES.items():
        if normalized != estado:
            continue
        if canon(alias) in s:
            return True
    return False


def _canon_cell(v: Any) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return canon(v)


def detect_header_row_df(df: pd.DataFrame, max_scan_rows: int = 5) -> int:
    if df is None or df.empty:
        return 1
    keywords = ["run", "rut", "dv", "prest", "presta", "fentrada", "fsalida", "sigte", "id", "idlocal", "est"]
    best_row = 1
    best_score = -1
    limit = min(max_scan_rows, len(df))
    for r in range(limit):
        values = [_canon_cell(c) for c in df.iloc[r].tolist()]
        score = sum(any(k in v for k in keywords) for v in values if v)
        if score > best_score:
            best_score = score
            best_row = r + 1
    return best_row


def map_columns_values(header_vals: List[Any]) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for i, val in enumerate(header_vals, start=1):
        h = _canon_cell(val)
        if h:
            m[h] = i
    return m


def find_col_idx(colmap: Dict[str, int], candidates: List[str]) -> int:
    for cand in candidates:
        c = canon(cand)
        if c in colmap:
            return colmap[c]
        for k, idx in colmap.items():
            if c and c in k:
                return idx
    return 0


class _CpuLoadThrottle:
    def __init__(self, target_percent: float) -> None:
        self.cpu_count = max(1, int(os.cpu_count() or 1))
        pct = max(0.0, min(100.0, float(target_percent or 0.0)))
        self.target_cores = (pct / 100.0) * self.cpu_count
        self.enabled = self.target_cores > 0.0
        self.wall_start = time.perf_counter()
        self.cpu_start = time.process_time()
        self.counter = 0

    def tick(self) -> None:
        if not self.enabled:
            return
        self.counter += 1
        if self.counter % 2048 != 0:
            return
        wall = max(1e-6, time.perf_counter() - self.wall_start)
        cpu = max(0.0, time.process_time() - self.cpu_start)
        used_cores = cpu / wall
        if used_cores <= self.target_cores:
            return
        target_wall = cpu / max(self.target_cores, 1e-6)
        sleep_for = target_wall - wall
        if sleep_for > 0:
            time.sleep(min(sleep_for, 0.05))


class _PgConnGuard:
    def __init__(self, conn: Any) -> None:
        self._conn = conn

    def close(self) -> None:
        conn = self._conn
        self._conn = None
        if conn is None:
            return
        try:
            conn.close()
        except Exception:
            pass

    def __del__(self) -> None:
        self.close()


@dataclass(frozen=True)
class TimelineRec:
    f_in: Optional[datetime]
    f_out: Optional[datetime]
    sigte_id: str
    id_local: str
    source: str
    extremidad: str = ""


@dataclass
class NominaFileData:
    exact_keys: set
    by_patient_presta: Dict[str, List[TimelineRec]]
    by_id_meta: Dict[str, Tuple[str, str, Optional[Dict[str, str]]]]
    exact_keys_td: set
    by_patient_presta_td: Dict[str, List[TimelineRec]]


class DBIndex:
    def __init__(self, db_dir: Path):
        self.db_dir = db_dir
        self._load_throttle = _CpuLoadThrottle(PRELOAD_MAX_CPU_PERCENT)
        self._compact_mode = bool(COMPACT_DB_INDEX)
        self._stream_batch_size = max(100, int(SQL_STREAM_BATCH_SIZE or 5000))
        self._nomina_lazy_verify = bool(NOMINA_VERIFY_LAZY_LOAD)
        self._nomina_records_loaded = False
        self.historico_unico: set = set()
        self.historico_timeline: Dict[str, List[TimelineRec]] = defaultdict(list)
        self.historico_by_id: set = set()
        self.historico_by_id_map: Dict[str, str] = {}
        self.historico_core: set = set()
        self.cgr_399: Dict[str, str] = {}
        self.cgr_84: Dict[str, str] = {}
        self.defunciones_rut: set = set()
        self.defunciones_fecha: Dict[str, str] = {}
        self.defunciones_fecha_dt: Dict[str, datetime.date] = {}
        self.establecimientos: set = set()
        self.nomina_files: Dict[str, NominaFileData] = {}
        self.nomina_files_meta: Dict[str, Tuple[str, Optional[str]]] = {}
        self.nomina_exact_keys: set = set()
        self.nomina_by_patient_presta: Dict[str, List[TimelineRec]] = defaultdict(list)
        self.nomina_exact_keys_td: set = set()
        self.nomina_by_patient_presta_td: Dict[str, List[TimelineRec]] = defaultdict(list)
        self.nomina_by_id_meta: Dict[str, Tuple[str, str, Optional[Dict[str, str]]]] = {}
        self.comges_especiales_sigte_ids: set = set()

    def _tick_load_throttle(self) -> None:
        self._load_throttle.tick()

    def _intern(self, value: Any) -> str:
        if not self._compact_mode:
            return "" if value is None else str(value)
        return _intern_str(value)

    def _set_nomina_id_meta(
        self,
        mapping: Dict[str, Tuple[str, str, Optional[Dict[str, str]]]],
        key: str,
        sigte_id: str,
        source: str,
        record: Optional[Dict[str, str]] = None,
        keep_record: bool = False,
    ) -> None:
        k = self._intern(normalize_id(key))
        if not k:
            return
        sig = self._intern(normalize_id(sigte_id) or k)
        src = self._intern(source or "")
        rec = record if record is not None else None
        if (not keep_record) and self._nomina_lazy_verify:
            rec = None
        mapping[k] = (sig, src, rec)

    def has_nomina_ids(self) -> bool:
        return bool(self.nomina_by_id_meta)

    def has_nomina_records(self) -> bool:
        if not self.nomina_by_id_meta:
            return False
        if self._nomina_records_loaded:
            return True
        for _k, (_sig, _src, rec) in self.nomina_by_id_meta.items():
            if rec:
                return True
        return False

    def get_nomina_sigte(self, key: Any) -> str:
        k = normalize_id(key)
        if not k:
            return ""
        meta = self.nomina_by_id_meta.get(k)
        if not meta:
            return ""
        return meta[0] or ""

    def get_nomina_source(self, key: Any, default: str = "NOMINAS") -> str:
        k = normalize_id(key)
        if not k:
            return default
        meta = self.nomina_by_id_meta.get(k)
        if not meta:
            return default
        return meta[1] or default

    def get_nomina_record(self, key: Any) -> Optional[Dict[str, str]]:
        k = normalize_id(key)
        if not k:
            return None
        meta = self.nomina_by_id_meta.get(k)
        if not meta:
            return None
        rec = meta[2]
        if rec is not None:
            return rec
        if self._nomina_lazy_verify and (not self._nomina_records_loaded):
            self.ensure_nomina_records_loaded()
            meta = self.nomina_by_id_meta.get(k)
            if not meta:
                return None
            return meta[2]
        return None

    def load_all(self) -> None:
        self._load_historico()
        self._load_cgr()
        self._load_defunciones()
        self._load_establecimientos()
        self._load_comges_especiales()
        self._load_nominas(include_records=not self._nomina_lazy_verify)
        for k in list(self.nomina_by_patient_presta.keys()):
            self.nomina_by_patient_presta[k].sort(key=lambda r: (r.f_in or datetime.min))

    def _load_historico(self) -> None:
        def _consume_rows(header_vals: List[Any], rows: Iterable[Tuple[Any, ...]]) -> int:
            colmap = map_columns_values(header_vals)
            col_run = find_col_idx(colmap, ["RUN", "run"])
            col_dv = find_col_idx(colmap, ["DV", "dv"])
            col_tipo = find_col_idx(colmap, ["TIPO_PREST", "tipo_prest"])
            col_presta = find_col_idx(colmap, ["PRESTA_MIN", "presta_min"])
            col_plano = find_col_idx(colmap, ["PLANO", "plano"])
            col_ext = find_col_idx(colmap, ["EXTREMIDAD", "extremidad"])
            col_fin = find_col_idx(colmap, ["F_ENTRADA", "f_entrada"])
            col_fout = find_col_idx(colmap, ["F_SALIDA", "F_salida"])
            col_est = find_col_idx(colmap, ["ESTAB_DEST", "estab_dest"])
            col_idlocal = find_col_idx(colmap, ["ID_LOCAL", "id_local"])
            col_sigte = find_col_idx(colmap, ["SIGTE_ID", "sigte_id"])
            loaded = 0

            for r in rows:
                self._tick_load_throttle()
                try:
                    def _get(idx: int) -> Any:
                        return _row_value(r, idx)

                    run = self._intern(normalize_run(_get(col_run))) if col_run else ""
                    dv = self._intern(normalize_dv(_get(col_dv))) if col_dv else ""
                    tipo = self._intern(normalize_text(_get(col_tipo))) if col_tipo else ""
                    presta_raw = _get(col_presta)
                    presta = self._intern("" if presta_raw is None else str(presta_raw).strip())
                    presta_norm = self._intern(normalize_presta(presta))
                    plano = self._intern(normalize_text(_get(col_plano))) if col_plano else ""
                    ext = self._intern(normalize_text(_get(col_ext))) if col_ext else ""
                    f_in_dt = parse_excel_date(_get(col_fin)) if col_fin else None
                    f_in = to_excel_serial(f_in_dt)
                    est = self._intern(normalize_text(_get(col_est))) if col_est else ""
                    unico = self._intern(f"{run}{dv}{tipo}{presta}{plano}{ext}{f_in}{est}")
                    if unico:
                        self.historico_unico.add(unico)
                    if run and dv and presta_norm and f_in:
                        core_key = self._intern(f"{run}{dv}{presta_norm}{f_in}")
                        self.historico_core.add(core_key)

                    f_out_dt = parse_excel_date(_get(col_fout)) if col_fout else None
                    sigte_id = ""
                    if col_sigte:
                        val_s = _get(col_sigte)
                        if val_s is not None:
                            sigte_id = self._intern(normalize_id(val_s))
                    id_local = ""
                    if col_idlocal:
                        val = _get(col_idlocal)
                        if val is not None:
                            id_local = self._intern(normalize_id(val))
                    if not sigte_id:
                        sigte_id = id_local
                    if id_local:
                        self.historico_by_id.add(id_local)
                        self.historico_by_id_map.setdefault(id_local, sigte_id or id_local)
                    if sigte_id:
                        self.historico_by_id.add(sigte_id)
                        self.historico_by_id_map.setdefault(sigte_id, sigte_id)
                    key = self._intern(f"{run}|{dv}|{presta_norm}")
                    if run and dv and presta and (f_in_dt or f_out_dt):
                        self.historico_timeline[key].append(
                            TimelineRec(
                                f_in_dt,
                                f_out_dt,
                                sigte_id=sigte_id,
                                id_local=id_local,
                                source="HISTORICO",
                                extremidad=ext
                            )
                        )
                    loaded += 1
                except Exception:
                    continue
            return loaded

        required_historico_cols = [
            "RUN",
            "DV",
            "TIPO_PREST",
            "PRESTA_MIN",
            "PLANO",
            "EXTREMIDAD",
            "F_ENTRADA",
            "F_SALIDA",
            "ESTAB_DEST",
            "ID_LOCAL",
            "SIGTE_ID",
        ]
        candidates = [
            _CORE_HISTORICO_TABLE,
            str(PG_BASE_TABLES.get("historico", "") or "").strip(),
        ]
        seen_candidates: set = set()
        last_error: Optional[Exception] = None
        for table_name in candidates:
            name = str(table_name or "").strip()
            if (not name) or (name in seen_candidates):
                continue
            seen_candidates.add(name)
            try:
                with _pg_connect() as conn:
                    header_vals, rows_iter = _sql_stream_table_rows_conn(
                        conn=conn,
                        table_name=name,
                        required_columns=required_historico_cols,
                        batch_size=self._stream_batch_size,
                    )
                    if not header_vals:
                        continue
                    loaded = _consume_rows([str(c).upper() for c in header_vals], rows_iter)
            except Exception as e:
                last_error = e
                continue
            if loaded >= 0:
                return
        if last_error is not None:
            print(f"Advertencia: no se pudo cargar historico desde PostgreSQL: {last_error}")
        return

    def _load_cgr(self) -> None:
        with _pg_connect() as conn:
            header_vals, rows_iter = _sql_stream_table_rows_conn(
                conn=conn,
                table_name=PG_BASE_TABLES["cgr"],
                required_columns=["ID_LOCAL", "ANEXO", "ORIGEN", "OBS", "DETALLE"],
                batch_size=self._stream_batch_size,
            )
            if not header_vals:
                return
            colmap = map_columns_values([str(c).upper() for c in header_vals])
            col_key = find_col_idx(colmap, ["ID_LOCAL", "id_local"])
            col_val = find_col_idx(colmap, ["ANEXO", "ORIGEN", "OBS", "DETALLE"])
            for r in rows_iter:
                self._tick_load_throttle()
                k = _row_value(r, col_key) if col_key else None
                v = _row_value(r, col_val) if col_val else None
                if k is None:
                    continue
                kk = self._intern(normalize_id(k))
                vv = self._intern("" if v is None else str(v).strip())
                if not kk or not vv:
                    continue
                vv_norm = canon(vv)
                if "84" in vv_norm and "399" not in vv_norm:
                    self.cgr_84[kk] = vv
                elif "399" in vv_norm and "84" not in vv_norm:
                    self.cgr_399[kk] = vv
                else:
                    self.cgr_399[kk] = vv

    def _load_defunciones(self) -> None:
        def _consume_rows(header_vals: List[Any], rows: Iterable[Tuple[Any, ...]]) -> None:
            colmap = map_columns_values(header_vals)
            col_rut = find_col_idx(colmap, ["RUTCONCATENADO", "RUT"])
            col_run = find_col_idx(colmap, ["RUN", "run"])
            col_dv = find_col_idx(colmap, ["DV", "dv"])
            col_fecha_def = find_col_idx(colmap, ["FECHA_DEF", "fecha_def", "FECHA DEF", "FECHA_DEFUNCION", "F_DEFUNCION"])

            for r in rows:
                self._tick_load_throttle()
                rut_norm = ""
                if col_run and col_dv:
                    run_val = normalize_run(_row_value(r, col_run))
                    dv_val = normalize_dv(_row_value(r, col_dv))
                    if run_val and dv_val:
                        rut_norm = f"{run_val}-{dv_val}"
                if not rut_norm and col_rut:
                    v = _row_value(r, col_rut)
                    s = "" if v is None else str(v).strip().upper().replace(".", "").replace(" ", "")
                    if "-" in s:
                        parts = s.split("-")
                        if len(parts) >= 2:
                            run = normalize_run(parts[0])
                            dv = normalize_dv(parts[1][:1])
                            if run and dv:
                                rut_norm = f"{run}-{dv}"
                    elif re.fullmatch(r"[0-9]+[0-9K]", s):
                        run = normalize_run(s[:-1])
                        dv = normalize_dv(s[-1])
                        if run and dv:
                            rut_norm = f"{run}-{dv}"
                if not rut_norm:
                    continue
                rut_norm = self._intern(rut_norm)
                self.defunciones_rut.add(rut_norm)

                fecha_val = _row_value(r, col_fecha_def) if col_fecha_def else None
                fecha_def = ""
                dt = parse_excel_date(fecha_val)
                if dt is not None:
                    fecha_def = dt.strftime("%d-%m-%Y")
                    if rut_norm not in self.defunciones_fecha_dt:
                        self.defunciones_fecha_dt[rut_norm] = dt.date()
                elif fecha_val is not None:
                    fecha_def = str(fecha_val).strip()
                if fecha_def and rut_norm not in self.defunciones_fecha:
                    self.defunciones_fecha[rut_norm] = self._intern(fecha_def)

        with _pg_connect() as conn:
            header_vals, rows_iter = _sql_stream_table_rows_conn(
                conn=conn,
                table_name=PG_BASE_TABLES["defunciones"],
                required_columns=[
                    "RUTCONCATENADO",
                    "RUT",
                    "RUN",
                    "DV",
                    "FECHA_DEF",
                    "FECHA DEF",
                    "FECHA_DEFUNCION",
                    "F_DEFUNCION",
                ],
                batch_size=self._stream_batch_size,
            )
            if not header_vals:
                return
            _consume_rows([str(c).upper() for c in header_vals], rows_iter)

    def _load_establecimientos(self) -> None:
        def _consume_rows(header_vals: List[Any], rows: Iterable[Tuple[Any, ...]]) -> None:
            colmap = map_columns_values(header_vals)
            col_code = find_col_idx(colmap, ["CODIGO"])
            if col_code == 0:
                col_code = 1
            for r in rows:
                self._tick_load_throttle()
                v = _row_value(r, col_code) if col_code else None
                if v is None:
                    continue
                code = self._intern(str(v).strip())
                if code:
                    self.establecimientos.add(code)

        with _pg_connect() as conn:
            header_vals, rows_iter = _sql_stream_table_rows_conn(
                conn=conn,
                table_name=PG_BASE_TABLES["establecimientos"],
                required_columns=["CODIGO"],
                batch_size=self._stream_batch_size,
            )
            if not header_vals:
                return
            _consume_rows([str(c).upper() for c in header_vals], rows_iter)

    def _load_comges_especiales(self) -> None:
        def _consume_rows(header_vals: List[Any], rows: Iterable[Tuple[Any, ...]]) -> int:
            sigte_candidates = {
                canon("SIGTE_ID"),
                canon("sigte_id"),
                canon("SIGTE_ID2"),
                canon("sigte_id2"),
                canon("SIGTE_ID.1"),
                canon("sigte_id.1"),
            }
            colmap = map_columns_values(header_vals)
            sigte_cols = [idx for key, idx in colmap.items() if key in sigte_candidates]
            if not sigte_cols:
                return 0
            loaded = 0
            for row in rows:
                for col_sigte in sigte_cols:
                    self._tick_load_throttle()
                    sigte_norm = self._intern(normalize_id(_row_value(row, col_sigte)))
                    if not sigte_norm:
                        continue
                    self.comges_especiales_sigte_ids.add(sigte_norm)
                    loaded += 1
            return loaded

        table_candidates = [
            _sql_comges_especiales_table_name(None),
            _CORE_COMGES_ESPECIALES_TABLE,
        ]
        seen_tables: set = set()
        last_error: Optional[Exception] = None
        for table_name in table_candidates:
            name = str(table_name or "").strip()
            if not name or name in seen_tables:
                continue
            seen_tables.add(name)
            try:
                with _pg_connect() as conn:
                    header_vals, rows_iter = _sql_stream_table_rows_conn(
                        conn=conn,
                        table_name=name,
                        required_columns=[
                            "SIGTE_ID",
                            "sigte_id",
                            "SIGTE_ID2",
                            "sigte_id2",
                            "SIGTE_ID.1",
                            "sigte_id.1",
                        ],
                        batch_size=self._stream_batch_size,
                    )
                    if not header_vals:
                        continue
                    loaded = _consume_rows(header_vals, rows_iter)
                    if loaded > 0:
                        return
            except Exception as e:
                last_error = e
                continue

        def _consume_df(df_source: pd.DataFrame) -> int:
            if df_source is None or df_source.empty:
                return 0
            header_vals = [str(c).strip() for c in df_source.columns]
            rows_iter = df_source.itertuples(index=False, name=None)
            return _consume_rows(header_vals, rows_iter)

        workbook_candidates: List[Path] = []
        candidate_names = ("comges especiales.xlsx", "comges_especiales.xlsx")
        for fname in candidate_names:
            try:
                workbook_candidates.append(self.db_dir.parent / fname)
            except Exception:
                pass
            workbook_candidates.append(Path.cwd() / fname)
            workbook_candidates.append(self.db_dir / fname)

        seen_paths: set = set()
        for path in workbook_candidates:
            path_str = str(path.resolve()) if isinstance(path, Path) else str(path)
            if path_str in seen_paths:
                continue
            seen_paths.add(path_str)
            if not path.exists() or not path.is_file():
                continue
            try:
                df_local = load_work_df(path)
            except Exception as e:
                last_error = e
                continue
            if _consume_df(df_local) > 0:
                return

        if last_error is not None:
            print(f"Advertencia: no se pudo cargar comges especiales: {last_error}")

    def _load_nominas(self, include_records: bool = True) -> None:
        active_infos: List[Tuple[Path, str, Optional[str]]] = []
        for (tipo, estado), table_name in PG_NOMINA_TABLES.items():
            pseudo = Path(f"{PG_SCHEMA}.{table_name}.sql")
            active_infos.append((pseudo, tipo, estado))
        active_keys = {str(path) for path, _tipo, _estado in active_infos}
        self.nomina_files = {}
        self.sync_nomina_files(active_infos, active_keys, include_records=include_records)
        self._nomina_records_loaded = bool(include_records)

    def _parse_nomina_file(
        self,
        path: Path,
        estado: Optional[str],
        include_records: bool = True,
    ) -> NominaFileData:
        data = NominaFileData(
            exact_keys=set(),
            by_patient_presta=defaultdict(list),
            by_id_meta={},
            exact_keys_td=set(),
            by_patient_presta_td=defaultdict(list),
        )

        def consume_rows(
            rows: Iterable[Tuple[Any, ...]],
            col_run: int,
            col_dv: int,
            col_presta: int,
            col_fin: int,
            col_fout: int,
            col_sig: int,
            col_idlocal: int,
            col_ext: int,
            col_estab_dest: int,
            verify_cols: Dict[str, int],
            source: str,
        ) -> None:
            for row in rows:
                self._tick_load_throttle()
                run = self._intern(normalize_run(_row_value(row, col_run))) if col_run else ""
                dv = self._intern(normalize_dv(_row_value(row, col_dv))) if col_dv else ""
                presta = self._intern(normalize_presta(_row_value(row, col_presta))) if col_presta else ""
                f_in_dt = normalize_date(_row_value(row, col_fin)) if col_fin else None
                f_out_dt = normalize_date(_row_value(row, col_fout)) if col_fout else None
                ext = self._intern(normalize_text(_row_value(row, col_ext))) if col_ext else ""

                sigte_id = ""
                val_sig = _row_value(row, col_sig) if col_sig else None
                if val_sig is not None:
                    sigte_id = self._intern(normalize_id(val_sig))

                id_local = ""
                val_idlocal = _row_value(row, col_idlocal) if col_idlocal else None
                if val_idlocal is not None:
                    id_local = self._intern(normalize_id(val_idlocal))

                if not id_local:
                    id_local = sigte_id

                estab_dest_norm = ""
                val_estab = _row_value(row, col_estab_dest) if col_estab_dest else None
                if val_estab is not None:
                    estab_dest_norm = normalize_id(val_estab)
                is_td = bool(estab_dest_norm) and estab_dest_norm == ESTAB_DEST_FILTER

                if sigte_id and sigte_id not in data.by_id_meta:
                    self._set_nomina_id_meta(
                        data.by_id_meta,
                        sigte_id,
                        sigte_id,
                        source,
                        record=None,
                        keep_record=False,
                    )
                if id_local and id_local not in data.by_id_meta:
                    record: Optional[Dict[str, str]] = None
                    if include_records:
                        record = {}
                        for field, idx in verify_cols.items():
                            if idx:
                                record[field] = self._intern(normalize_compare_value(field, _row_value(row, idx)))
                            else:
                                record[field] = ""
                    self._set_nomina_id_meta(
                        data.by_id_meta,
                        id_local,
                        sigte_id or id_local,
                        source,
                        record=record,
                        keep_record=bool(include_records),
                    )

                if not (run and dv and presta):
                    continue

                if f_in_dt:
                    exact_key = self._intern(f"{run}|{dv}|{presta}|{to_excel_serial(f_in_dt)}")
                    data.exact_keys.add(exact_key)
                    if is_td:
                        data.exact_keys_td.add(exact_key)

                key_pp = self._intern(f"{run}|{dv}|{presta}")
                if f_in_dt or f_out_dt:
                    rec = TimelineRec(
                        f_in=f_in_dt,
                        f_out=f_out_dt,
                        sigte_id=sigte_id,
                        id_local=id_local,
                        source=source,
                        extremidad=ext
                    )
                    data.by_patient_presta[key_pp].append(rec)
                    if is_td:
                        data.by_patient_presta_td[key_pp].append(rec)

        if path.suffix.lower() == ".sql":
            table_name = _pg_table_from_pseudo_path(path)
            required_nomina_cols: List[str] = [
                "RUN",
                "DV",
                "PRESTA_MIN",
                "F_ENTRADA",
                "F_SALIDA",
                "SIGTE_ID",
                "ID_LOCAL",
                "EXTREMIDAD",
                "ESTAB_DEST",
            ]
            if include_records:
                for cands in VERIFY_FIELDS.values():
                    required_nomina_cols.extend(cands)
            source_label = estado or table_name
            source = self._intern(f"{path.name}:{source_label}")
            with _pg_connect() as conn:
                header_vals_raw, rows_iter = _sql_stream_table_rows_conn(
                    conn=conn,
                    table_name=table_name,
                    required_columns=required_nomina_cols,
                    batch_size=self._stream_batch_size,
                )
                if not header_vals_raw:
                    return data
                header_vals = [str(c).upper() for c in header_vals_raw]
                colmap = map_columns_values(header_vals)
                col_run = find_col_idx(colmap, ["RUN", "run"])
                col_dv = find_col_idx(colmap, ["DV", "dv"])
                col_presta = find_col_idx(colmap, ["PRESTA_MIN", "presta_min"])
                col_fin = find_col_idx(colmap, ["F_ENTRADA", "f_entrada"])
                col_fout = find_col_idx(colmap, ["F_SALIDA", "f_salida"])
                col_sig = find_col_idx(colmap, ["SIGTE_ID", "sigte_id"])
                col_idlocal = find_col_idx(colmap, ["ID_LOCAL", "id_local"])
                col_ext = find_col_idx(colmap, ["EXTREMIDAD", "extremidad"])
                col_estab_dest = find_col_idx(colmap, ["ESTAB_DEST", "estab_dest"])
                verify_cols = {f: find_col_idx(colmap, cands) for f, cands in VERIFY_FIELDS.items()} if include_records else {}
                consume_rows(
                    rows_iter,
                    col_run,
                    col_dv,
                    col_presta,
                    col_fin,
                    col_fout,
                    col_sig,
                    col_idlocal,
                    col_ext,
                    col_estab_dest,
                    verify_cols,
                    source,
                )
            return data

        raise RuntimeError(
            f"Modo SQL activo: origen de nomina no soportado ({path.suffix})."
        )

    def _rebuild_nomina_aggregates(self) -> None:
        self.nomina_exact_keys = set()
        self.nomina_by_patient_presta = defaultdict(list)
        self.nomina_exact_keys_td = set()
        self.nomina_by_patient_presta_td = defaultdict(list)
        self.nomina_by_id_meta = {}

        file_items: List[Tuple[int, int, str, str]] = []
        for key in self.nomina_files.keys():
            path = Path(key)
            try:
                st = path.stat()
                mtime = st.st_mtime_ns
                size = st.st_size
            except Exception:
                mtime = 0
                size = 0
            file_items.append((mtime, size, path.name.lower(), key))
        file_items.sort()

        for _mtime, _size, _name, key in file_items:
            data = self.nomina_files[key]
            self.nomina_exact_keys.update(data.exact_keys)
            for k, recs in data.by_patient_presta.items():
                self.nomina_by_patient_presta[k].extend(recs)
            self.nomina_exact_keys_td.update(data.exact_keys_td)
            for k, recs in data.by_patient_presta_td.items():
                self.nomina_by_patient_presta_td[k].extend(recs)
            for k, meta in data.by_id_meta.items():
                self.nomina_by_id_meta[k] = meta

        for k in list(self.nomina_by_patient_presta.keys()):
            self.nomina_by_patient_presta[k].sort(key=lambda r: (r.f_in or datetime.min))
        for k in list(self.nomina_by_patient_presta_td.keys()):
            self.nomina_by_patient_presta_td[k].sort(key=lambda r: (r.f_in or datetime.min))

    def sync_nomina_files(
        self,
        active_infos: List[Tuple[Path, str, Optional[str]]],
        changed_paths: Iterable[str],
        include_records: bool = True,
    ) -> None:
        active_keys = {str(path) for path, _tipo, _estado in active_infos}
        changed_set = set(changed_paths)
        failed_pairs: set = set()

        for path, _tipo, estado in active_infos:
            key = str(path)
            if key not in self.nomina_files or key in changed_set:
                try:
                    self.nomina_files[key] = self._parse_nomina_file(path, estado, include_records=include_records)
                except Exception as e:
                    print(f"Advertencia: no se pudo cargar nomina {path.name}: {e}")
                    failed_pairs.add((_tipo, estado))
                    if key not in self.nomina_files:
                        continue
            if key in self.nomina_files:
                self.nomina_files_meta[key] = (_tipo, estado)

        for key in list(self.nomina_files.keys()):
            if key not in active_keys:
                meta = self.nomina_files_meta.get(key)
                if meta and (meta[0], meta[1]) in failed_pairs:
                    continue
                del self.nomina_files[key]
        for key in list(self.nomina_files_meta.keys()):
            if key not in active_keys:
                meta = self.nomina_files_meta.get(key)
                if meta and (meta[0], meta[1]) in failed_pairs:
                    continue
                del self.nomina_files_meta[key]

        self._rebuild_nomina_aggregates()
        self._nomina_records_loaded = bool(include_records)

    def reload_historico(self) -> None:
        self.historico_unico = set()
        self.historico_timeline = defaultdict(list)
        self.historico_by_id = set()
        self.historico_by_id_map = {}
        self.historico_core = set()
        self._load_historico()

    def reload_cgr(self) -> None:
        self.cgr_399 = {}
        self.cgr_84 = {}
        self._load_cgr()

    def reload_defunciones(self) -> None:
        self.defunciones_rut = set()
        self.defunciones_fecha = {}
        self.defunciones_fecha_dt = {}
        self._load_defunciones()

    def reload_establecimientos(self) -> None:
        self.establecimientos = set()
        self._load_establecimientos()

    def reload_comges_especiales(self) -> None:
        self.comges_especiales_sigte_ids = set()
        self._load_comges_especiales()

    def ensure_nomina_records_loaded(self) -> None:
        if self._nomina_records_loaded:
            return
        active_infos: List[Tuple[Path, str, Optional[str]]] = []
        for (tipo, estado), table_name in PG_NOMINA_TABLES.items():
            pseudo = Path(f"{PG_SCHEMA}.{table_name}.sql")
            active_infos.append((pseudo, tipo, estado))
        active_keys = {str(path) for path, _tipo, _estado in active_infos}
        self.sync_nomina_files(active_infos, active_keys, include_records=True)
        self._nomina_records_loaded = True

    def reload_all_in_place(self) -> None:
        self.reload_historico()
        _release_memory_pressure(rounds=1, sleep_ms=0)
        self.reload_cgr()
        _release_memory_pressure(rounds=1, sleep_ms=0)
        self.reload_defunciones()
        _release_memory_pressure(rounds=1, sleep_ms=0)
        self.reload_establecimientos()
        _release_memory_pressure(rounds=1, sleep_ms=0)
        self.reload_comges_especiales()
        _release_memory_pressure(rounds=1, sleep_ms=0)
        self.nomina_files = {}
        self.nomina_files_meta = {}
        self.nomina_exact_keys = set()
        self.nomina_by_patient_presta = defaultdict(list)
        self.nomina_exact_keys_td = set()
        self.nomina_by_patient_presta_td = defaultdict(list)
        self.nomina_by_id_meta = {}
        self._nomina_records_loaded = False
        self._load_nominas(include_records=not self._nomina_lazy_verify)
        _release_memory_pressure(rounds=1, sleep_ms=0)


def _to_date(d: Optional[datetime]) -> Optional[datetime.date]:
    if d is None:
        return None
    try:
        if pd.isna(d):
            return None
    except Exception:
        pass
    try:
        return d.date()
    except Exception:
        return None


def normalize_range_date(f_in: Optional[datetime], f_out: Optional[datetime]) -> Optional[Tuple[datetime.date, Optional[datetime.date]]]:
    d_in = _to_date(f_in)
    d_out = _to_date(f_out)
    if d_in and d_out:
        if d_out < d_in:
            d_in, d_out = d_out, d_in
        return d_in, d_out
    if d_in:
        return d_in, None
    if d_out:
        return d_out, d_out
    return None


def ranges_overlap_date(a: Tuple[datetime.date, Optional[datetime.date]],
                        b: Tuple[datetime.date, Optional[datetime.date]]) -> bool:
    big = datetime(9999, 12, 31).date()
    a_start, a_end = a
    b_start, b_end = b
    ae = a_end or big
    be = b_end or big
    return (a_start <= be) and (b_start <= ae)


def days_diff(a: Optional[datetime], b: Optional[datetime]) -> Optional[int]:
    if a is None or b is None:
        return None
    try:
        if pd.isna(a) or pd.isna(b):
            return None
    except Exception:
        pass
    try:
        ad = a.date() if hasattr(a, "date") else a
        bd = b.date() if hasattr(b, "date") else b
        return abs((ad - bd).days)
    except Exception:
        return None


def any_same_day(a_in: Optional[datetime], a_out: Optional[datetime],
                 b_in: Optional[datetime], b_out: Optional[datetime]) -> bool:
    pairs = [
        (a_in, b_in),
        (a_in, b_out),
        (a_out, b_in),
        (a_out, b_out),
    ]
    for x, y in pairs:
        d = days_diff(x, y)
        if d is not None and d == 0:
            return True
    return False


def _legacy_compute_traslape(run: str, dv: str, presta: str, f_in: Optional[datetime], db: DBIndex) -> str:
    presta = normalize_presta(presta)

    if not (run and dv and presta and f_in):
        return "Sin evaluaciÃ³n (faltan datos clave)"

    exact_key = f"{run}|{dv}|{presta}|{to_excel_serial(f_in)}"
    if exact_key in db.nomina_exact_keys_td:
        return "Sin traslape (caso existe en nÃ³minas)"

    key_pp = f"{run}|{dv}|{presta}"
    if key_pp not in db.nomina_by_patient_presta_td:
        return "Sin traslape (sin registros en nÃ³minas de la especialidad)"

    recs = db.nomina_by_patient_presta_td[key_pp]
    for rec in reversed(recs):
        if rec.f_in and rec.f_in > f_in and rec.f_out is not None:
            sid = rec.sigte_id or "(sin SIGTE_ID)"
            return f"Caso traslape, traslape con SIGTE_ID: {sid}"
        if rec.f_in and rec.f_in <= f_in:
            break

    return "Sin traslape (Ultimo caso registrado de la especialidad)"


def _legacy_compute_duplicidad(run: str, dv: str, presta: str, f_in: Optional[datetime], f_out: Optional[datetime],
                       id_local: str, extremidad: str, db: DBIndex,
                       work_seen: Dict[str, List[TimelineRec]]) -> str:
    presta = normalize_presta(presta)
    if not (run and dv and presta):
        return "Sin evaluaciÃ³n (faltan datos clave)"

    current_range = normalize_range_date(f_in, f_out)
    if current_range is None:
        return "Sin evaluaciÃ³n (faltan fechas)"

    key_pp = f"{run}|{dv}|{presta}"
    candidates: List[TimelineRec] = []
    candidates.extend(db.nomina_by_patient_presta_td.get(key_pp, []))
    candidates.extend(db.historico_timeline.get(key_pp, []))
    candidates.extend(work_seen.get(key_pp, []))
    ext_cur = normalize_text(extremidad) if extremidad else ""
    for rec in candidates:
        rec_ext = normalize_text(rec.extremidad) if rec.extremidad else ""
        if ext_cur and rec_ext and ext_cur != rec_ext:
            continue
        rec_range = normalize_range_date(rec.f_in, rec.f_out)
        if rec_range is None:
            continue
        if ranges_overlap_date(current_range, rec_range) or any_same_day(f_in, f_out, rec.f_in, rec.f_out):
            other_id = rec.id_local or rec.sigte_id or ""
            if normalize_id(other_id) and normalize_id(id_local) and normalize_id(other_id) == normalize_id(id_local):
                continue
            if other_id:
                return f"Caso duplicado, duplicidad con ID_LOCAL: {other_id}"
            return "Caso duplicado (sin ID_LOCAL)"

    return "Sin duplicidad"


def _legacy_compute_caso_cercano(run: str, dv: str, presta: str, f_in: Optional[datetime], f_out: Optional[datetime],
                         id_local: str, db: DBIndex, max_days: int = 365) -> str:
    presta = normalize_presta(presta)
    if not (run and dv and presta):
        return "Sin evaluaciÃ³n (faltan datos clave)"
    if not (f_in or f_out):
        return "Sin evaluaciÃ³n (faltan fechas)"

    key_pp = f"{run}|{dv}|{presta}"
    recs = db.nomina_by_patient_presta_td.get(key_pp, [])
    current_id = (id_local or "").strip()

    best_diff: Optional[int] = None
    best_id = ""
    for rec in recs:
        rec_id = (rec.id_local or rec.sigte_id or "").strip()
        if current_id and rec_id and rec_id == current_id:
            continue
        diffs = [
            days_diff(f_in, rec.f_in),
            days_diff(f_in, rec.f_out),
            days_diff(f_out, rec.f_in),
            days_diff(f_out, rec.f_out),
        ]
        for d in diffs:
            if d is None:
                continue
            if best_diff is None or d < best_diff:
                best_diff = d
                best_id = rec.id_local or rec.sigte_id or ""

    if best_diff is not None and best_diff <= max_days:
        id_msg = best_id if best_id else "sin ID"
        return f"Alerta: Caso cercano en fechas del registro ({id_msg})"

    return "Sin alerta"


def compute_traslape(
    run: str,
    dv: str,
    presta: str,
    f_in: Optional[datetime],
    db: DBIndex,
    sql_conn: Optional[Any] = None,
    sql_cache: Optional[Dict[Tuple[str, str, str, bool, bool], List[TimelineRec]]] = None,
) -> str:
    presta = normalize_presta(presta)
    if not (run and dv and presta and f_in):
        return "Sin evaluación (faltan datos clave)"

    sql_ok = False
    recs: List[TimelineRec] = []
    if sql_conn is not None:
        try:
            recs = _sql_fetch_patient_timeline_records_cached(
                conn=sql_conn,
                run=run,
                dv=dv,
                presta=presta,
                include_nomina=True,
                include_historico=False,
                cache=sql_cache,
            )
            sql_ok = True
        except Exception:
            sql_ok = False

    if sql_ok:
        exact_serial = to_excel_serial(f_in)
        for rec in recs:
            if rec.f_in and to_excel_serial(rec.f_in) == exact_serial:
                return "Sin traslape (caso existe en nóminas)"
        if not recs:
            return "Sin traslape (sin registros en nóminas de la especialidad)"
        for rec in reversed(recs):
            if rec.f_in and rec.f_in > f_in and rec.f_out is not None:
                sid = rec.sigte_id or "(sin SIGTE_ID)"
                return f"Caso traslape, traslape con SIGTE_ID: {sid}"
            if rec.f_in and rec.f_in <= f_in:
                break
        return "Sin traslape (Ultimo caso registrado de la especialidad)"

    exact_key = f"{run}|{dv}|{presta}|{to_excel_serial(f_in)}"
    if exact_key in db.nomina_exact_keys_td:
        return "Sin traslape (caso existe en nóminas)"
    key_pp = f"{run}|{dv}|{presta}"
    if key_pp not in db.nomina_by_patient_presta_td:
        return "Sin traslape (sin registros en nóminas de la especialidad)"
    recs = db.nomina_by_patient_presta_td[key_pp]
    for rec in reversed(recs):
        if rec.f_in and rec.f_in > f_in and rec.f_out is not None:
            sid = rec.sigte_id or "(sin SIGTE_ID)"
            return f"Caso traslape, traslape con SIGTE_ID: {sid}"
        if rec.f_in and rec.f_in <= f_in:
            break
    return "Sin traslape (Ultimo caso registrado de la especialidad)"


def compute_duplicidad(
    run: str,
    dv: str,
    presta: str,
    f_in: Optional[datetime],
    f_out: Optional[datetime],
    id_local: str,
    extremidad: str,
    db: DBIndex,
    work_seen: Dict[str, List[TimelineRec]],
    sql_conn: Optional[Any] = None,
    sql_cache: Optional[Dict[Tuple[str, str, str, bool, bool], List[TimelineRec]]] = None,
) -> str:
    presta = normalize_presta(presta)
    if not (run and dv and presta):
        return "Sin evaluaciÃ³n (faltan datos clave)"

    current_range = normalize_range_date(f_in, f_out)
    if current_range is None:
        return "Sin evaluaciÃ³n (faltan fechas)"

    key_pp = f"{run}|{dv}|{presta}"
    candidates: List[TimelineRec] = []
    sql_ok = False
    if sql_conn is not None:
        try:
            candidates.extend(
                _sql_fetch_patient_timeline_records_cached(
                    conn=sql_conn,
                    run=run,
                    dv=dv,
                    presta=presta,
                    include_nomina=True,
                    include_historico=False,
                    cache=sql_cache,
                )
            )
            candidates.extend(
                _sql_fetch_patient_timeline_records_cached(
                    conn=sql_conn,
                    run=run,
                    dv=dv,
                    presta=presta,
                    include_nomina=False,
                    include_historico=True,
                    cache=sql_cache,
                )
            )
            sql_ok = True
        except Exception:
            sql_ok = False

    if not sql_ok:
        candidates.extend(db.nomina_by_patient_presta_td.get(key_pp, []))
        candidates.extend(db.historico_timeline.get(key_pp, []))

    candidates.extend(work_seen.get(key_pp, []))
    ext_cur = normalize_text(extremidad) if extremidad else ""
    for rec in candidates:
        rec_ext = normalize_text(rec.extremidad) if rec.extremidad else ""
        if ext_cur and rec_ext and ext_cur != rec_ext:
            continue
        rec_range = normalize_range_date(rec.f_in, rec.f_out)
        if rec_range is None:
            continue
        if ranges_overlap_date(current_range, rec_range) or any_same_day(f_in, f_out, rec.f_in, rec.f_out):
            other_id = rec.id_local or rec.sigte_id or ""
            if normalize_id(other_id) and normalize_id(id_local) and normalize_id(other_id) == normalize_id(id_local):
                continue
            if other_id:
                return f"Caso duplicado, duplicidad con ID_LOCAL: {other_id}"
            return "Caso duplicado (sin ID_LOCAL)"

    return "Sin duplicidad"


def compute_caso_cercano(
    run: str,
    dv: str,
    presta: str,
    f_in: Optional[datetime],
    f_out: Optional[datetime],
    id_local: str,
    db: DBIndex,
    max_days: int = 365,
    sql_conn: Optional[Any] = None,
    sql_cache: Optional[Dict[Tuple[str, str, str, bool, bool], List[TimelineRec]]] = None,
) -> str:
    presta = normalize_presta(presta)
    if not (run and dv and presta):
        return "Sin evaluaciÃ³n (faltan datos clave)"
    if not (f_in or f_out):
        return "Sin evaluaciÃ³n (faltan fechas)"

    key_pp = f"{run}|{dv}|{presta}"
    recs: List[TimelineRec] = []
    sql_ok = False
    if sql_conn is not None:
        try:
            recs = _sql_fetch_patient_timeline_records_cached(
                conn=sql_conn,
                run=run,
                dv=dv,
                presta=presta,
                include_nomina=True,
                include_historico=False,
                cache=sql_cache,
            )
            sql_ok = True
        except Exception:
            sql_ok = False

    if not sql_ok:
        recs = db.nomina_by_patient_presta_td.get(key_pp, [])

    current_id = (id_local or "").strip()
    best_diff: Optional[int] = None
    best_id = ""
    for rec in recs:
        rec_id = (rec.id_local or rec.sigte_id or "").strip()
        if current_id and rec_id and rec_id == current_id:
            continue
        diffs = [
            days_diff(f_in, rec.f_in),
            days_diff(f_in, rec.f_out),
            days_diff(f_out, rec.f_in),
            days_diff(f_out, rec.f_out),
        ]
        for d in diffs:
            if d is None:
                continue
            if best_diff is None or d < best_diff:
                best_diff = d
                best_id = rec.id_local or rec.sigte_id or ""

    if best_diff is not None and best_diff <= max_days:
        id_msg = best_id if best_id else "sin ID"
        return f"Alerta: Caso cercano en fechas del registro ({id_msg})"

    return "Sin alerta"


# =========================
# Procesamiento del archivo de trabajo
# =========================
def load_work_df(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = _load_df_any(path)
        if df is None or df.empty:
            return pd.DataFrame()
        df["__HOJA_ORIGEN__"] = "CSV"
        df.columns = [str(c).strip() if str(c).strip() else f"COL_{i}" for i, c in enumerate(df.columns)]
        return df

    engine = "pyxlsb" if path.suffix.lower() == ".xlsb" else "openpyxl"
    try:
        with pd.ExcelFile(path, engine=engine) as xls:
            sheet_names = list(xls.sheet_names or [])
            if not sheet_names:
                return pd.DataFrame()

            if len(sheet_names) == 1:
                df = pd.read_excel(xls, sheet_name=sheet_names[0], dtype=object)
                df["__HOJA_ORIGEN__"] = str(sheet_names[0])
            else:
                frames: List[pd.DataFrame] = []
                for sname in sheet_names:
                    sdf = pd.read_excel(xls, sheet_name=sname, dtype=object)
                    if sdf is None or sdf.empty:
                        continue
                    sdf = sdf.dropna(how="all")
                    if sdf.empty:
                        continue
                    sdf["__HOJA_ORIGEN__"] = str(sname)
                    frames.append(sdf)
                if not frames:
                    return pd.DataFrame()
                df = pd.concat(frames, ignore_index=True, sort=False)
    except ImportError as e:
        if engine == "pyxlsb":
            raise RuntimeError(
                "Para leer archivos .xlsb, instala la dependencia 'pyxlsb' (pip install pyxlsb)."
            ) from e
        raise
    df.columns = [str(c).strip() if str(c).strip() else f"COL_{i}" for i, c in enumerate(df.columns)]
    return df


def get_by_excel_letter(df: pd.DataFrame, col_letter: str) -> Optional[str]:
    col_letter = col_letter.upper().strip()
    num = 0
    for ch in col_letter:
        if not ("A" <= ch <= "Z"):
            return None
        num = num * 26 + (ord(ch) - ord("A") + 1)
    idx = num - 1
    if 0 <= idx < df.shape[1]:
        return df.columns[idx]
    return None


def pick_col(df: pd.DataFrame, candidates: List[str], fallback_letter: Optional[str] = None) -> Optional[str]:
    cmap = {canon(c): c for c in df.columns}
    for cand in candidates:
        c = canon(cand)
        # exact
        if c in cmap:
            return cmap[c]
        # partial
        for key, real in cmap.items():
            if c and c in key:
                return real
    if fallback_letter:
        return get_by_excel_letter(df, fallback_letter)
    return None


def pick_cols(df: pd.DataFrame, candidates: List[str]) -> List[str]:
    cmap = {canon(c): c for c in df.columns}
    found: List[str] = []
    for cand in candidates:
        c = canon(cand)
        if not c:
            continue
        if c in cmap:
            found.append(cmap[c])
            continue
        for key, real in cmap.items():
            if c and c in key:
                found.append(real)
                break

    out: List[str] = []
    seen = set()
    for col in found:
        if col not in seen:
            out.append(col)
            seen.add(col)
    return out


def contacto_flag(df: pd.DataFrame) -> pd.Series:
    phone_cols = pick_cols(df, ["FONO_FIJO", "FONO_MOVIL","fono_fijo","fono_movil","telefono","celular", "TELEFONO", "CELULAR", "FONO", "CONTACTO_1"])
    email_cols = pick_cols(df, ["EMAIL", "MAIL", "CORREO", "CONTACTO_2", "email", "correo", "mail"])
    
    cols: List[str] = []
    for c in phone_cols + email_cols :
        if c in df.columns:
            cols.append(c)
    if len(cols) == 0:
        return pd.Series(["Sin datos de contacto"] * len(df), index=df.index)

    tmp = df[cols].fillna("").astype(str).apply(lambda s: s.str.strip())
    all_blank = tmp.eq("").all(axis=1)
    return np.where(all_blank, "Sin datos de contacto", "Posee datos de contacto")


def faltantes_report(df: pd.DataFrame, mode: str, contacto: pd.Series) -> pd.Series:
    if mode == "ingreso":
        required = [
            "SERV_SALUD", "RUN", "DV", "NOMBRES", "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "FECHA_NAC",
            "SEXO", "PREVISION", "TIPO_PREST", "PRESTA_MIN",
            "PRESTA_EST", "F_ENTRADA", "ESTAB_ORIG", "ESTAB_DEST",
            "PRAIS", "REGION", "COMUNA",
            "CIUDAD", "COND_RURALIDAD",
            "NOM_CALLE",
            "RUN_PROF_SOL", "DV_PROF_SOL",
            "ID_LOCAL",
        ]
    else:
        required = [
            "SERV_SALUD" , "RUN", "DV", "NOMBRES", "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "FECHA_NAC",
            "SEXO", "PREVISION", "TIPO_PREST", "PRESTA_MIN",
            "PRESTA_EST", "F_ENTRADA", "ESTAB_ORIG", "ESTAB_DEST", "F_SALIDA", "C_SALIDA", "E_OTOR_AT",
            "PRESTA_MIN_SALIDA",
            "PRAIS", "REGION", "COMUNA",
            "CIUDAD", "COND_RURALIDAD",
            "NOM_CALLE",
            "RUN_PROF_SOL", "DV_PROF_SOL", "RUN_PROF_RESOL", "DV_PROF_RESOL",
            "ID_LOCAL",
        ]

    cols = pick_cols(df, required)

    if not cols:
        return pd.Series(["Sin datos faltantes"] * len(df), index=df.index)

    tmp = df[cols].fillna("").astype(str).apply(lambda s: s.str.strip())
    blank = tmp.eq("").to_numpy()
    colnames = np.array(cols)
    y_cols = pick_cols(df, ["SOSPECHA_DIAG"])
    z_cols = pick_cols(df, ["CONFIR_DIAG"])
    yz_blank = None
    if y_cols and z_cols:
        y = y_cols[0]
        z = z_cols[0]
        yz_tmp = df[[y, z]].fillna("").astype(str).apply(lambda s: s.str.strip())
        yz_blank = yz_tmp.eq("").all(axis=1).to_numpy()

    out = []
    contacto_np = np.asarray(contacto)
    for i in range(blank.shape[0]):
        miss_idx = np.flatnonzero(blank[i])
        parts = [colnames[j] for j in miss_idx]
        if yz_blank is not None and yz_blank[i]:
            parts.append("Info en Y o Z")
        if contacto_np[i] != "Posee datos de contacto":
            parts.append("Medio de Contacto")

        if len(parts) == 0:
            out.append("Sin datos faltantes")
        else:
            out.append("Falta: " + ", ".join(parts))
    return pd.Series(out, index=df.index)


def process_file(
    work_path: Path,
    selected: Dict[str, bool],
    db: Optional[DBIndex] = None,
    progress_cb: Optional[Callable[[int], None]] = None,
    local_system_records: Optional[Dict[str, Dict[str, str]]] = None,
    cancel_cb: Optional[Callable[[], None]] = None,
) -> Tuple[Path, float, Dict[str, Any]]:
    process_throttle = _CpuLoadThrottle(PROCESS_MAX_CPU_PERCENT)

    def check_cancel() -> None:
        process_throttle.tick()
        if cancel_cb:
            cancel_cb()

    t0 = time.perf_counter()
    stage_timing: Dict[str, Any] = {}
    t_stage = t0
    check_cancel()
    df = load_work_df(work_path)
    n_rows = int(len(df)) if df is not None else 0
    stage_timing["carga_archivo_s"] = round(time.perf_counter() - t_stage, 3)
    t_stage = time.perf_counter()
    sql_bulk_requested = BULK_SQL_CROSSES and any(
        selected.get(k, False)
        for k in ("historico", "cgr", "defunciones", "macrored", "traslape", "duplicidad")
    )
    sql_conn: Optional[Any] = None
    if ROW_LEVEL_SQL_LOOKUPS or sql_bulk_requested:
        try:
            sql_conn = _pg_connect()
        except Exception:
            sql_conn = None
    sql_conn_guard = _PgConnGuard(sql_conn) if sql_conn is not None else None
    stage_timing["conexion_sql_s"] = round(time.perf_counter() - t_stage, 3)
    stage_timing["sql_disponible"] = bool(sql_conn is not None)
    stage_timing["bulk_sql_solicitado"] = bool(sql_bulk_requested)
    t_stage = time.perf_counter()
    sql_timeline_cache: Dict[Tuple[str, str, str, bool, bool], List[TimelineRec]] = {}
    row_level_sql_conn: Optional[Any] = sql_conn if (ROW_LEVEL_SQL_LOOKUPS and not sql_bulk_requested) else None
    total_steps = 1
    total_steps += n_rows
    if selected.get("historico", False):
        total_steps += 1
    if selected.get("cgr", False):
        total_steps += 1
    if selected.get("defunciones", False):
        total_steps += 1
    if selected.get("macrored", False):
        total_steps += 1
    if selected.get("traslape", False):
        total_steps += n_rows
    if selected.get("duplicidad", False):
        total_steps += n_rows
    if local_system_records is not None:
        total_steps += 1
    total_steps += 1  
    step = 0
    def report_progress(pct: int) -> None:
        check_cancel()
        if progress_cb:
            progress_cb(max(0, min(100, pct)))

    def bump(n: int = 1) -> None:
        nonlocal step
        step += n
        report_progress(int((step / total_steps) * 100))

    bump(1)
    progress_every = max(1, n_rows // 100)

    col_run = pick_col(df, ["RUN", "run"], fallback_letter="B")
    col_dv = pick_col(df, ["DV", "dv"], fallback_letter="C")
    col_presta = pick_col(df, ["PRESTA_MIN", "presta_min"], fallback_letter="K") 
    col_tipo = pick_col(df, ["TIPO_PREST", "tipo_prest"], fallback_letter="J")
    col_plano = pick_col(df, ["PLANO", "plano"], fallback_letter="L")
    col_ext = pick_col(df, ["EXTREMIDAD", "extremidad"], fallback_letter="M")
    col_fin = pick_col(df, ["F_ENTRADA", "f_entrada"], fallback_letter="O")
    col_fout = pick_col(df, ["F_SALIDA", "f_salida"], fallback_letter="P")
    col_est = pick_col(df, ["estab_dest", "ESTAB_DEST", "EST_DEST"], fallback_letter="Q")
    col_idlocal = pick_col(df, ["ID_LOCAL", "id_local"], fallback_letter="AO")
    col_sigte_input = pick_col(df, ["SIGTE_ID", "sigte_id"])
    col_estab_codigo = pick_col(df,["ESTAB_DEST", "ESTAB_ORIG"])
    col_serv = pick_col(df, ["SERV_SALUD", "serv_salud"])
    col_nom = pick_col(df, ["NOMBRES", "nombres", "NOMBRE", "nombre"])
    col_ap1 = pick_col(df, ["PRIMER_APELLIDO", "APELLIDO_PATERNO", "APELLIDO1", "primer_apellido", "apellido_paterno", "apellido1"])
    col_ap2 = pick_col(df, ["SEGUNDO_APELLIDO", "APELLIDO_MATERNO", "APELLIDO2", "segundo_apellido", "apellido_materno", "apellido2"])
    col_fnac = pick_col(df, ["FECHA_NAC", "NACIMIENTO", "FECHA_NACIMIENTO", "fecha_nac", "nacimiento", "fecha_nacimiento"])
    col_sexo = pick_col(df, ["SEXO", "GENERO", "sexo", "genero"])
    col_presta_est = pick_col(df, ["PRESTA_EST", "presta_est"])
    col_estab_or = pick_col(df, ["ESTAB_ORIG", "estab_orig"])
    col_estab_de = pick_col(df, ["ESTAB_DEST", "estab_dest"])
    col_csal = pick_col(df, ["C_SALIDA", "c_salida"])
    col_eotor = pick_col(df, ["E_OTOR_AT", "e_otor_at"])
    col_presta_sal = pick_col(df, ["PRESTA_MIN_SALIDA", "presta_min_salida"])
    col_prais = pick_col(df, ["PRAIS", "prais"])
    col_prevision = pick_col(df, ["PREVISION", "prevision"])
    col_region = pick_col(df, ["REGION", "region"])
    col_comuna = pick_col(df, ["COMUNA", "comuna"])
    col_ciudad = pick_col(df, ["CIUDAD", "ciudad"])
    col_ruralidad = pick_col(df, ["COND_RURALIDAD", "cond_ruralidad"])
    col_nom_calle = pick_col(df, ["NOM_CALLE", "nom_calle"])
    col_run_ps = pick_col(df, ["RUN_PROF_SOL", "run_prof_sol"])
    col_dv_ps = pick_col(df, ["DV_PROF_SOL", "dc_prof_sol"])
    col_run_pr = pick_col(df, ["RUN_PROF_RESOL", "run_prof_resol"])
    col_dv_pr = pick_col(df, ["DV_PROF_RESOL", "dv_prof_resol"])
    col_sospecha = pick_col(df, ["SOSPECHA_DIAG", "sospecha_diag"])
    col_confirma = pick_col(df, ["CONFIR_DIAG", "confir_diag"])

    run_vals = df[col_run].map(normalize_run) if col_run else pd.Series("", index=df.index, dtype=object)
    dv_vals = df[col_dv].map(normalize_dv) if col_dv else pd.Series("", index=df.index, dtype=object)
    rut_concat_vals = run_vals + "-" + dv_vals
    df["RUT CONCATENADO"] = rut_concat_vals

    fin_dt = df[col_fin].map(normalize_date) if col_fin else pd.Series([None] * n_rows, index=df.index, dtype=object)
    fout_dt = df[col_fout].map(normalize_date) if col_fout else pd.Series([None] * n_rows, index=df.index, dtype=object)
    df["ALERTA FECHAS"] = np.where(
        fin_dt.notna() & fout_dt.notna() & (fout_dt < fin_dt),
        "Alerta: Incongruencia en fechas",
        "Sin problemas"
    )
    presta_vals = df[col_presta].fillna("").astype(str).str.strip() if col_presta else pd.Series("", index=df.index, dtype=object)
    presta_norm_vals = presta_vals.map(normalize_presta)
    idlocal_vals = df[col_idlocal].fillna("").astype(str).str.strip() if col_idlocal else pd.Series("", index=df.index, dtype=object)
    idlocal_norm = idlocal_vals.map(normalize_id)
    sigte_input_vals = (
        df[col_sigte_input].fillna("").astype(str).str.strip()
        if col_sigte_input else pd.Series("", index=df.index, dtype=object)
    )
    sigte_input_norm = sigte_input_vals.map(normalize_id)
    ext_vals = df[col_ext].fillna("").astype(str).map(normalize_text) if col_ext else pd.Series("", index=df.index, dtype=object)

    def _series(col: Optional[str], field: str) -> pd.Series:
        if col and col in df.columns:
            return df[col].map(lambda v: normalize_compare_value(field, v))
        return pd.Series("", index=df.index, dtype=object)

    ver_series = {
        "SERV_SALUD": _series(col_serv, "SERV_SALUD"),
        "RUN": _series(col_run, "RUN"),
        "DV": _series(col_dv, "DV"),
        "NOMBRES": _series(col_nom, "NOMBRES"),
        "PRIMER_APELLIDO": _series(col_ap1, "PRIMER_APELLIDO"),
        "SEGUNDO_APELLIDO": _series(col_ap2, "SEGUNDO_APELLIDO"),
        "FECHA_NAC": _series(col_fnac, "FECHA_NAC"),
        "SEXO": _series(col_sexo, "SEXO"),
        "TIPO_PREST": _series(col_tipo, "TIPO_PREST"),
        "PRESTA_MIN": _series(col_presta, "PRESTA_MIN"),
        "PLANO": _series(col_plano, "PLANO"),
        "EXTREMIDAD": _series(col_ext, "EXTREMIDAD"),
        "PRESTA_EST": _series(col_presta_est, "PRESTA_EST"),
        "F_ENTRADA": _series(col_fin, "F_ENTRADA"),
        "ESTAB_ORIG": _series(col_estab_or, "ESTAB_ORIG"),
        "ESTAB_DEST": _series(col_estab_de, "ESTAB_DEST"),
        "F_SALIDA": _series(col_fout, "F_SALIDA"),
        "C_SALIDA": _series(col_csal, "C_SALIDA"),
        "E_OTOR_AT": _series(col_eotor, "E_OTOR_AT"),
        "PRESTA_MIN_SALIDA": _series(col_presta_sal, "PRESTA_MIN_SALIDA"),
        "PRAIS": _series(col_prais, "PRAIS"),
        "RUN_PROF_SOL": _series(col_run_ps, "RUN_PROF_SOL"),
        "DV_PROF_SOL": _series(col_dv_ps, "DV_PROF_SOL"),
        "RUN_PROF_RESOL": _series(col_run_pr, "RUN_PROF_RESOL"),
        "DV_PROF_RESOL": _series(col_dv_pr, "DV_PROF_RESOL"),
    }

    ingreso_required = [
        "SERV_SALUD", "RUN", "DV", "NOMBRES", "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "FECHA_NAC",
        "SEXO", "PREVISION", "TIPO_PREST", "PRESTA_MIN", "PRESTA_EST", "F_ENTRADA", "ESTAB_ORIG",
        "ESTAB_DEST", "PRAIS", "REGION", "COMUNA", "CIUDAD", "COND_RURALIDAD", "NOM_CALLE",
        "RUN_PROF_SOL", "DV_PROF_SOL", "ID_LOCAL",
    ]
    egreso_required = [
        "SERV_SALUD", "RUN", "DV", "NOMBRES", "PRIMER_APELLIDO", "SEGUNDO_APELLIDO", "FECHA_NAC",
        "SEXO", "PREVISION", "TIPO_PREST", "PRESTA_MIN", "PRESTA_EST", "F_ENTRADA", "ESTAB_ORIG",
        "ESTAB_DEST", "F_SALIDA", "C_SALIDA", "E_OTOR_AT", "PRESTA_MIN_SALIDA", "PRAIS", "REGION",
        "COMUNA", "CIUDAD", "COND_RURALIDAD", "NOM_CALLE", "RUN_PROF_SOL", "DV_PROF_SOL",
        "RUN_PROF_RESOL", "DV_PROF_RESOL", "ID_LOCAL",
    ]

    stage_timing["normalizacion_s"] = round(time.perf_counter() - t_stage, 3)
    t_stage = time.perf_counter()
    sql_bulk_indexed: Optional[pd.DataFrame] = None
    if sql_conn is not None and sql_bulk_requested and n_rows > 0:
        try:
            def _date_to_iso(v: Any) -> str:
                if v is None:
                    return ""
                try:
                    if pd.isna(v):
                        return ""
                except Exception:
                    pass
                if isinstance(v, datetime):
                    return v.date().isoformat()
                dt = parse_excel_date(v)
                return dt.date().isoformat() if dt is not None else ""

            col_phone = pick_cols(
                df,
                ["FONO_FIJO", "FONO_MOVIL", "fono_fijo", "fono_movil", "telefono", "celular", "TELEFONO", "CELULAR", "FONO", "CONTACTO_1"],
            )
            col_email = pick_cols(df, ["EMAIL", "MAIL", "CORREO", "CONTACTO_2", "email", "correo", "mail"])
            phone_src = col_phone[:5]
            email_src = col_email[:3]

            present_stage_cols: set = set()

            def _stage_col(
                work: pd.DataFrame,
                stage_name: str,
                src_col: Optional[str],
                normalize: Optional[Callable[[Any], Any]] = None,
            ) -> None:
                if src_col and src_col in df.columns:
                    vals = _sql_work_series(df, src_col)
                    if normalize is not None:
                        vals = vals.map(normalize)
                    work[stage_name] = vals
                    present_stage_cols.add(stage_name)
                else:
                    work[stage_name] = ""

            work_sql = pd.DataFrame(index=df.index)
            work_sql["row_id"] = np.arange(1, n_rows + 1, dtype=np.int32)
            work_sql["run_norm"] = run_vals
            work_sql["dv_norm"] = dv_vals
            work_sql["presta_norm"] = presta_norm_vals
            _stage_col(work_sql, "tipo_norm", col_tipo, normalize_text)
            _stage_col(work_sql, "plano_norm", col_plano, normalize_text)
            work_sql["ext_norm"] = ext_vals
            work_sql["f_in"] = fin_dt.map(_date_to_iso)
            work_sql["f_out"] = fout_dt.map(_date_to_iso)
            _stage_col(work_sql, "estab_norm", col_est, normalize_id)
            work_sql["id_local_norm"] = idlocal_norm
            work_sql["rut_norm"] = rut_concat_vals.map(normalize_rut_concat)
            if col_estab_codigo and col_estab_codigo in df.columns:
                work_sql["estab_codigo"] = df[col_estab_codigo].fillna("").astype(str).map(lambda v: str(v).strip())
                present_stage_cols.add("estab_codigo")
            else:
                work_sql["estab_codigo"] = ""
            _stage_col(work_sql, "serv_salud", col_serv)
            _stage_col(work_sql, "run_txt", col_run)
            _stage_col(work_sql, "dv_txt", col_dv)
            _stage_col(work_sql, "nombres", col_nom)
            _stage_col(work_sql, "primer_apellido", col_ap1)
            _stage_col(work_sql, "segundo_apellido", col_ap2)
            _stage_col(work_sql, "fecha_nac", col_fnac)
            _stage_col(work_sql, "sexo", col_sexo)
            _stage_col(work_sql, "prevision", col_prevision)
            _stage_col(work_sql, "presta_est", col_presta_est)
            _stage_col(work_sql, "estab_orig", col_estab_or)
            _stage_col(work_sql, "estab_dest", col_estab_de)
            _stage_col(work_sql, "c_salida", col_csal)
            _stage_col(work_sql, "e_otor_at", col_eotor)
            _stage_col(work_sql, "presta_min_salida", col_presta_sal)
            _stage_col(work_sql, "prais", col_prais)
            _stage_col(work_sql, "region", col_region)
            _stage_col(work_sql, "comuna", col_comuna)
            _stage_col(work_sql, "ciudad", col_ciudad)
            _stage_col(work_sql, "cond_ruralidad", col_ruralidad)
            _stage_col(work_sql, "nom_calle", col_nom_calle)
            _stage_col(work_sql, "run_prof_sol", col_run_ps)
            _stage_col(work_sql, "dv_prof_sol", col_dv_ps)
            _stage_col(work_sql, "run_prof_resol", col_run_pr)
            _stage_col(work_sql, "dv_prof_resol", col_dv_pr)
            _stage_col(work_sql, "sospecha_diag", col_sospecha)
            _stage_col(work_sql, "confir_diag", col_confirma)
            for i in range(5):
                src = phone_src[i] if i < len(phone_src) else None
                _stage_col(work_sql, f"phone_{chr(ord('a') + i)}", src)
            for i in range(3):
                src = email_src[i] if i < len(email_src) else None
                _stage_col(work_sql, f"email_{chr(ord('a') + i)}", src)

            required_to_stage = {
                "SERV_SALUD": "serv_salud",
                "RUN": "run_txt",
                "DV": "dv_txt",
                "NOMBRES": "nombres",
                "PRIMER_APELLIDO": "primer_apellido",
                "SEGUNDO_APELLIDO": "segundo_apellido",
                "FECHA_NAC": "fecha_nac",
                "SEXO": "sexo",
                "PREVISION": "prevision",
                "TIPO_PREST": "tipo_norm",
                "PRESTA_MIN": "presta_norm",
                "PRESTA_EST": "presta_est",
                "F_ENTRADA": "f_in",
                "ESTAB_ORIG": "estab_orig",
                "ESTAB_DEST": "estab_dest",
                "F_SALIDA": "f_out",
                "C_SALIDA": "c_salida",
                "E_OTOR_AT": "e_otor_at",
                "PRESTA_MIN_SALIDA": "presta_min_salida",
                "PRAIS": "prais",
                "REGION": "region",
                "COMUNA": "comuna",
                "CIUDAD": "ciudad",
                "COND_RURALIDAD": "cond_ruralidad",
                "NOM_CALLE": "nom_calle",
                "RUN_PROF_SOL": "run_prof_sol",
                "DV_PROF_SOL": "dv_prof_sol",
                "RUN_PROF_RESOL": "run_prof_resol",
                "DV_PROF_RESOL": "dv_prof_resol",
                "ID_LOCAL": "id_local_norm",
            }
            ingreso_fields_present = [
                (label, required_to_stage[label])
                for label in ingreso_required
                if required_to_stage.get(label, "") in present_stage_cols
            ]
            egreso_fields_present = [
                (label, required_to_stage[label])
                for label in egreso_required
                if required_to_stage.get(label, "") in present_stage_cols
            ]
            contact_columns_present = [
                c for c in [
                    "phone_a", "phone_b", "phone_c", "phone_d", "phone_e",
                    "email_a", "email_b", "email_c",
                ] if c in present_stage_cols
            ]
            include_yz = ("sospecha_diag" in present_stage_cols) and ("confir_diag" in present_stage_cols)

            bulk_df = _sql_bulk_enrich_work_df(
                conn=sql_conn,
                work_df=work_sql,
                selected=selected,
                ingreso_fields_present=ingreso_fields_present,
                egreso_fields_present=egreso_fields_present,
                contact_columns_present=contact_columns_present,
                include_yz=include_yz,
            )
            if bulk_df is not None and not bulk_df.empty and "row_id" in bulk_df.columns:
                row_id_series = pd.to_numeric(bulk_df["row_id"], errors="coerce")
                valid_mask = row_id_series.notna()
                if not bool(valid_mask.all()):
                    bulk_df = bulk_df.loc[valid_mask]
                    row_id_series = row_id_series.loc[valid_mask]
                bulk_df = bulk_df.assign(row_id=row_id_series.astype(np.int32, copy=False).to_numpy())
                sql_bulk_indexed = bulk_df.set_index("row_id", drop=True)
        except Exception as e:
            sql_bulk_indexed = None
            stage_timing["enriquecimiento_sql_error"] = str(e)
    stage_timing["enriquecimiento_sql_s"] = round(time.perf_counter() - t_stage, 3)
    stage_timing["enriquecimiento_sql_activo"] = bool(sql_bulk_indexed is not None)
    if sql_bulk_requested and sql_bulk_indexed is None:
        row_level_sql_conn = None
    stage_timing["row_level_sql_activo"] = bool(row_level_sql_conn is not None)
    t_stage = time.perf_counter()

    def _apply_sql_output(out_col: str, sql_col: str) -> bool:
        if sql_bulk_indexed is None:
            return False
        if sql_col not in sql_bulk_indexed.columns:
            return False
        ordered = sql_bulk_indexed.reindex(np.arange(1, n_rows + 1))[sql_col]
        df[out_col] = ordered.fillna("").astype(str).tolist()
        return True

    def _ensure_db_for_fallback() -> None:
        nonlocal db
        if db is None:
            try:
                db = get_db()
            except Exception:
                db = None
            return
        try:
            maybe_empty = (
                not db.historico_unico
                and not db.cgr_399
                and not db.cgr_84
                and not db.defunciones_rut
                and not db.establecimientos
                and not db.has_nomina_ids()
                and not db.has_nomina_records()
                and not db.nomina_exact_keys_td
            )
        except Exception:
            maybe_empty = False
        if not maybe_empty:
            return
        try:
            db = get_db()
        except Exception:
            db = None

    sql_nomina_by_id: Dict[str, str] = {}
    sql_nomina_by_source: Dict[str, str] = {}
    sql_nomina_by_record: Dict[str, Dict[str, str]] = {}
    sql_historico_by_id: Dict[str, str] = {}
    if sql_conn is not None and n_rows > 0:
        lookup_ids = list(idlocal_norm.tolist()) + list(sigte_input_norm.tolist())
        need_sql_records = bool(selected.get("verificacion", False))
        try:
            sql_nomina_by_id, sql_nomina_by_source, sql_nomina_by_record = _sql_fetch_nomina_lookup_for_ids(
                sql_conn,
                lookup_ids,
                include_records=need_sql_records,
            )
        except Exception:
            sql_nomina_by_id = {}
            sql_nomina_by_source = {}
            sql_nomina_by_record = {}
        try:
            sql_historico_by_id = _sql_fetch_historico_lookup_for_ids(sql_conn, lookup_ids)
        except Exception:
            sql_historico_by_id = {}

    if not _apply_sql_output("¿POSEE ALGUN MEDIO DE CONTACTO?", "contacto"):
        contacto = contacto_flag(df)
        df["¿POSEE ALGUN MEDIO DE CONTACTO?"] = contacto
    else:
        contacto = df["¿POSEE ALGUN MEDIO DE CONTACTO?"]
    if not _apply_sql_output("¿POSEE DATOS FALTANTES? (CARGA INGRESO)", "falt_ingreso"):
        df["¿POSEE DATOS FALTANTES? (CARGA INGRESO)"] = faltantes_report(df, "ingreso", contacto)
    if not _apply_sql_output("¿POSEE DATOS FALTANTES? (CARGA EGRESO)", "falt_egreso"):
        df["¿POSEE DATOS FALTANTES? (CARGA EGRESO)"] = faltantes_report(df, "egreso", contacto)
    prevision_values = (
        df[col_prevision].fillna("").astype(str)
        if col_prevision and col_prevision in df.columns
        else pd.Series([""] * len(df), index=df.index, dtype=object)
    )
    prevision_problem_mask = ~prevision_values.map(_normalize_prevision_value).isin({"1", "2"})
    if bool(prevision_problem_mask.any()):
        falt_ing_col = next(
            (
                c
                for c in df.columns
                if ("poseedatosfaltantes" in canon(c)) and ("cargaingreso" in canon(c))
            ),
            None,
        )
        falt_egr_col = next(
            (
                c
                for c in df.columns
                if ("poseedatosfaltantes" in canon(c)) and ("cargaegreso" in canon(c))
            ),
            None,
        )
        if falt_ing_col:
            df[falt_ing_col] = _append_issue_to_missing_report(
                df[falt_ing_col],
                prevision_problem_mask,
                "Prevision con problemas",
            )
        if falt_egr_col:
            df[falt_egr_col] = _append_issue_to_missing_report(
                df[falt_egr_col],
                prevision_problem_mask,
                "Prevision con problemas",
            )

    if not _apply_sql_output("ALERTA CASO CERCANO (< 1 año)", "alerta_cercano"):
        _ensure_db_for_fallback()
        alerta_cercano = []
        for i in range(n_rows):
            check_cancel()
            alerta_cercano.append(
                compute_caso_cercano(
                    run=run_vals.iat[i],
                    dv=dv_vals.iat[i],
                    presta=presta_norm_vals.iat[i],
                    f_in=fin_dt.iat[i],
                    f_out=fout_dt.iat[i],
                    id_local=idlocal_vals.iat[i],
                    db=db,
                    sql_conn=row_level_sql_conn,
                    sql_cache=sql_timeline_cache,
                )
            )
            if progress_cb and (i % progress_every == 0 or i == n_rows - 1):
                report_progress(int(((step + i + 1) / total_steps) * 100))
        df["ALERTA CASO CERCANO (< 1 año)"] = alerta_cercano
    step += n_rows
    report_progress(int((step / total_steps) * 100))

    if selected.get("historico", False):
        if not _apply_sql_output("CRUCE CERRADAS HISTORICAS", "cruce_historico"):
            _ensure_db_for_fallback()
            tipo_vals = df[col_tipo].fillna("").astype(str).map(normalize_text) if col_tipo else pd.Series("", index=df.index, dtype=object)
            plano_vals = df[col_plano].fillna("").astype(str).map(normalize_text) if col_plano else pd.Series("", index=df.index, dtype=object)
            ext_vals = df[col_ext].fillna("").astype(str).map(normalize_text) if col_ext else pd.Series("", index=df.index, dtype=object)
            est_vals = df[col_est].fillna("").astype(str).map(normalize_text) if col_est else pd.Series("", index=df.index, dtype=object)

            fin_serial = fin_dt.apply(to_excel_serial)
            unico = run_vals + dv_vals + tipo_vals + presta_vals + plano_vals + ext_vals + fin_serial + est_vals
            found_by_id = idlocal_norm.map(lambda v: bool(v) and v in db.historico_by_id)
            found_by_unico = unico.isin(db.historico_unico)
            presta_norm = presta_vals.map(normalize_presta)
            core_key = run_vals + dv_vals + presta_norm + fin_serial
            found_by_core = core_key.isin(db.historico_core)
            df["CRUCE CERRADAS HISTORICAS"] = np.where(
                found_by_id | found_by_unico | found_by_core,
                "Se encuentra en historico",
                "No se encuentra en historico"
            )
        bump(1)

    if selected.get("cgr", False):
        if not _apply_sql_output("CRUCE CGR 399", "cruce_cgr_399"):
            _ensure_db_for_fallback()
            key = idlocal_norm
            df["CRUCE CGR 399"] = key.map(lambda k: db.cgr_399.get(k, "No se encuentra en CGR 399"))
        if not _apply_sql_output("CRUCE CGR 84", "cruce_cgr_84"):
            _ensure_db_for_fallback()
            key = idlocal_norm
            df["CRUCE CGR 84"] = key.map(lambda k: db.cgr_84.get(k, "No se encuentra en CGR 84"))
        bump(1)

    if selected.get("defunciones", False):
        got_def = _apply_sql_output("CRUCE DEFUNCIONES", "cruce_defunciones")
        got_fecha = _apply_sql_output("FECHA DEFUNCION", "fecha_defuncion")
        got_alerta_fall = _apply_sql_output("ALERTA FALLECIMIENTO", "alerta_fallecimiento")
        got_alerta_fecha = _apply_sql_output("ALERTA FECHA DEFUNCION", "alerta_fecha_defuncion")
        if not (got_def and got_fecha and got_alerta_fall and got_alerta_fecha):
            _ensure_db_for_fallback()
            rut = df["RUT CONCATENADO"].map(normalize_rut_concat)
            in_def = rut.isin(db.defunciones_rut)
            df["CRUCE DEFUNCIONES"] = np.where(
                in_def,
                "Paciente fallecido",
                "Paciente vivo"
            )
            fecha_def_vals = rut.map(lambda r: db.defunciones_fecha_dt.get(r) if r else None)
            df["FECHA DEFUNCION"] = fecha_def_vals.map(lambda d: d.strftime("%d-%m-%Y") if d else "")
            alerta_fallecimiento: List[str] = []
            alerta_fechas_def: List[str] = []
            for i in range(n_rows):
                check_cancel()
                f_out = fout_dt.iat[i]
                f_in = fin_dt.iat[i]
                f_def = fecha_def_vals.iat[i]
                if bool(in_def.iat[i]) and f_out is not None and f_def is not None:
                    try:
                        if f_out.date() > f_def:
                            alerta_fallecimiento.append("Alerta: paciente con egreso posterior a la fecha de fallecimiento")
                        else:
                            alerta_fallecimiento.append("Sin alertas")
                    except Exception:
                        alerta_fallecimiento.append("Sin alertas")
                else:
                    alerta_fallecimiento.append("Sin alertas")

                if bool(in_def.iat[i]) and f_def is not None:
                    try:
                        fd = f_def
                        fi = f_in.date() if f_in is not None else None
                        fo = f_out.date() if f_out is not None else None
                        if (fi is not None and fd < fi) or (fo is not None and fd < fo):
                            alerta_fechas_def.append("Alerta: fecha de defuncion anterior a F_ENTRADA/F_SALIDA")
                        else:
                            alerta_fechas_def.append("Sin alertas")
                    except Exception:
                        alerta_fechas_def.append("Sin alertas")
                else:
                    alerta_fechas_def.append("Sin alertas")
            df["ALERTA FALLECIMIENTO"] = alerta_fallecimiento
            df["ALERTA FECHA DEFUNCION"] = alerta_fechas_def
        bump(1)

    # Macro red / establecimientos
    if selected.get("macrored", False):
        if not _apply_sql_output("CRUCE ESTABLECIMIENTOS", "cruce_establecimientos"):
            _ensure_db_for_fallback()
            if col_estab_codigo and col_estab_codigo in df.columns:
                cod = df[col_estab_codigo].fillna("").astype(str).str.strip()
                df["CRUCE ESTABLECIMIENTOS"] = np.where(
                    cod.isin(db.establecimientos),
                    "Corresponde establecimiento",
                    "Macro red"
                )
            else:
                df["CRUCE ESTABLECIMIENTOS"] = "Macro red"
        bump(1)

    if selected.get("verificacion", False):
        out_ver = []
        fields = list(VERIFY_FIELDS.keys())
        use_sql_records = bool(sql_nomina_by_record)
        if not use_sql_records:
            _ensure_db_for_fallback()
            if db is not None and not db.has_nomina_records():
                try:
                    db.ensure_nomina_records_loaded()
                except Exception:
                    pass
        for i in range(n_rows):
            check_cancel()
            id_key = normalize_id(idlocal_vals.iat[i])
            if not id_key:
                out_ver.append("Caso no encontrado en Nominas")
                continue
            nom_rec = sql_nomina_by_record.get(id_key) if use_sql_records else db.get_nomina_record(id_key)
            if not nom_rec:
                out_ver.append("Caso no encontrado en Nominas")
                continue

            diffs = []
            matches = 0
            for f in fields:
                w = ver_series[f].iat[i] if f in ver_series else ""
                n = nom_rec.get(f, "")
                if (not w) and (not n):
                    continue
                if w != n:
                    diffs.append(f)
                else:
                    matches += 1

            if matches == 0 and len(diffs) > 0:
                out_ver.append("ID_LOCAL no pertenece a paciente, Revisar caso")
            elif len(diffs) == 0:
                out_ver.append("Información Sin Problemas")
            else:
                out_ver.append("No coincide: " + ", ".join(diffs))

        df["VERIFICACION DE DATOS"] = out_ver
        bump(1)

    if local_system_records is not None:
        out_exists_local: List[str] = []
        out_ver_local: List[str] = []
        fields = list(VERIFY_FIELDS.keys())
        for i in range(n_rows):
            check_cancel()
            id_key = normalize_id(idlocal_vals.iat[i])
            if not id_key:
                out_exists_local.append("Sin ID_LOCAL/SIGTE_ID")
                out_ver_local.append("Caso no encontrado en Sistema local")
                continue

            local_rec = local_system_records.get(id_key)
            if not local_rec:
                out_exists_local.append("No se encuentra en Sistema local")
                out_ver_local.append("Caso no encontrado en Sistema local")
                continue

            out_exists_local.append("Se encuentra en Sistema local")
            diffs = []
            matches = 0
            for f in fields:
                w = ver_series[f].iat[i] if f in ver_series else ""
                n = local_rec.get(f, "")
                if (not w) and (not n):
                    continue
                if w != n:
                    diffs.append(f)
                else:
                    matches += 1

            if matches == 0 and len(diffs) > 0:
                out_ver_local.append("ID_LOCAL no pertenece a paciente, Revisar caso")
            elif len(diffs) == 0:
                out_ver_local.append("InformaciÃ³n Sin Problemas")
            else:
                out_ver_local.append("No coincide: " + ", ".join(diffs))

        df["CRUCE SISTEMA LOCAL"] = out_exists_local
        df["VERIFICACION SISTEMA LOCAL"] = out_ver_local
        bump(1)

    has_lookup_key = (idlocal_norm != "") | (sigte_input_norm != "")
    resolved_sigte_vals: List[str] = []
    resolved_source_vals: List[str] = []
    db_checked_for_lookup = False
    for i in range(n_rows):
        check_cancel()
        keys: List[str] = []
        id_key = normalize_id(idlocal_norm.iat[i])
        hint_key = normalize_id(sigte_input_norm.iat[i])
        if id_key:
            keys.append(id_key)
        if hint_key and hint_key not in keys:
            keys.append(hint_key)

        sigte_id_found = ""
        source_found = ""
        for key in keys:
            sigte_sql = sql_nomina_by_id.get(key, "")
            if sigte_sql:
                sigte_id_found = sigte_sql
                source_found = sql_nomina_by_source.get(key, "NOMINAS")
                break
        if not sigte_id_found:
            for key in keys:
                sigte_hist_sql = sql_historico_by_id.get(key, "")
                if sigte_hist_sql:
                    sigte_id_found = sigte_hist_sql
                    source_found = "HISTORICO"
                    break
        if not sigte_id_found and keys:
            if (db is None) and (not db_checked_for_lookup):
                _ensure_db_for_fallback()
                db_checked_for_lookup = True
            if db is not None:
                for key in keys:
                    sigte_mem = db.get_nomina_sigte(key)
                    if sigte_mem:
                        sigte_id_found = sigte_mem
                        source_found = db.get_nomina_source(key, "NOMINAS")
                        break
            if not sigte_id_found and db is not None:
                for key in keys:
                    if key in db.historico_by_id:
                        sigte_id_found = db.historico_by_id_map.get(key, key)
                        source_found = "HISTORICO"
                        break
        resolved_sigte_vals.append(sigte_id_found)
        resolved_source_vals.append(source_found)

    resolved_sigte_series = pd.Series(resolved_sigte_vals, index=df.index, dtype=object)
    resolved_source_series = pd.Series(resolved_source_vals, index=df.index, dtype=object)

    sql_comges_special_ids: set = set()
    sql_comges_available = False
    if sql_conn is not None:
        try:
            sql_comges_special_ids = _sql_fetch_comges_special_ids(sql_conn, resolved_sigte_series.tolist())
            sql_comges_available = True
        except Exception:
            sql_comges_special_ids = set()
            sql_comges_available = False

    if not sql_comges_available and db is None:
        _ensure_db_for_fallback()

    if sql_comges_available:
        check_cancel()
        special_mask = resolved_sigte_series.ne("") & resolved_sigte_series.isin(sql_comges_special_ids)
        mental_mask = special_mask & presta_norm_vals.isin(_OBSERVACION_SALUD_MENTAL_PRESTA_CODES_NORM)
        observacion_arr = np.full(n_rows, "Sin observaciones", dtype=object)
        special_np = special_mask.to_numpy()
        mental_np = mental_mask.to_numpy()
        observacion_arr[special_np] = "Sename"
        observacion_arr[mental_np] = "Salud mental"
        df["Observacion caso"] = observacion_arr
    else:
        observacion_vals: List[str] = []
        for i in range(n_rows):
            check_cancel()
            sigte_val = normalize_id(resolved_sigte_series.iat[i])
            presta_val = presta_vals.iat[i]
            observacion_vals.append(
                classify_observacion_caso(
                    sigte_val,
                    presta_val,
                    db,
                )
            )
        df["Observacion caso"] = observacion_vals

    if selected.get("nominas", False):
        df["CRUCE NOMINAS (SIGTE_ID)"] = np.where(
            ~has_lookup_key,
            "Sin ID_LOCAL/SIGTE_ID",
            np.where(resolved_sigte_series != "", resolved_sigte_series, "No se encuentra"),
        )
        df["ORIGEN SIGTE_ID"] = np.where(
            ~has_lookup_key,
            "Sin ID_LOCAL/SIGTE_ID",
            np.where(resolved_source_series != "", resolved_source_series, "No se encuentra"),
        )

    work_seen: Dict[str, List[TimelineRec]] = defaultdict(list)

    if selected.get("traslape", False):
        if not _apply_sql_output("TRASLAPE", "traslape"):
            _ensure_db_for_fallback()
            out_tras = []
            for i in range(n_rows):
                check_cancel()
                out_tras.append(
                    compute_traslape(
                        run=run_vals.iat[i],
                        dv=dv_vals.iat[i],
                        presta=presta_norm_vals.iat[i],
                        f_in=fin_dt.iat[i],
                        db=db,
                        sql_conn=row_level_sql_conn,
                        sql_cache=sql_timeline_cache,
                    )
                )
                if progress_cb and (i % progress_every == 0 or i == n_rows - 1):
                    report_progress(int(((step + i + 1) / total_steps) * 100))
            df["TRASLAPE"] = out_tras
        step += n_rows
        report_progress(int((step / total_steps) * 100))

    if selected.get("duplicidad", False):
        if not _apply_sql_output("DUPLICIDAD", "duplicidad"):
            _ensure_db_for_fallback()
            out_dup = []
            for i in range(n_rows):
                check_cancel()
                run = run_vals.iat[i]
                dv = dv_vals.iat[i]
                presta = presta_norm_vals.iat[i]
                fi = fin_dt.iat[i]
                fo = fout_dt.iat[i]
                idl = idlocal_vals.iat[i] or ""

                msg = compute_duplicidad(
                    run,
                    dv,
                    presta,
                    fi,
                    fo,
                    idl,
                    ext_vals.iat[i],
                    db,
                    work_seen,
                    sql_conn=row_level_sql_conn,
                    sql_cache=sql_timeline_cache,
                )
                out_dup.append(msg)
                key_pp = f"{run}|{dv}|{presta}"
                if run and dv and presta and (fi or fo):
                    work_seen[key_pp].append(
                        TimelineRec(fi, fo, sigte_id=idl, id_local=idl, source="ARCHIVO_TRABAJO", extremidad=ext_vals.iat[i])
                    )
                if progress_cb and (i % progress_every == 0 or i == n_rows - 1):
                    report_progress(int(((step + i + 1) / total_steps) * 100))

            df["DUPLICIDAD"] = out_dup
        step += n_rows
        report_progress(int((step / total_steps) * 100))

    bump(1)
    stage_timing["cruces_y_reglas_s"] = round(time.perf_counter() - t_stage, 3)
    t_stage = time.perf_counter()
    out_name = f"LE_NOGES_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}.xlsx"
    out_path = OUTPUT_DIR / out_name
    check_cancel()
    export_df = format_dates_for_export(df)
    with pd.ExcelWriter(
        out_path,
        engine="openpyxl",
        date_format="DD-MM-YYYY",
        datetime_format="DD-MM-YYYY",
    ) as writer:
        export_df.to_excel(writer, index=False)
        apply_short_date_format_to_workbook(writer.book)
    stage_timing["export_excel_s"] = round(time.perf_counter() - t_stage, 3)

    elapsed = time.perf_counter() - t0
    stage_timing["total_s"] = round(elapsed, 3)
    report_progress(100)
    if sql_conn_guard is not None:
        sql_conn_guard.close()
    return out_path, elapsed, stage_timing


def _to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, (int, np.integer)):
        return int(v)
    if isinstance(v, (float, np.floating)):
        if float(v).is_integer():
            return int(v)
        return int(float(v))
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return int(float(s))
    except Exception:
        return None


def _parse_excel_dates_series(values: pd.Series) -> pd.Series:
    out = pd.Series(pd.NaT, index=values.index, dtype="datetime64[ns]")
    if pd.api.types.is_datetime64_any_dtype(values):
        return pd.to_datetime(values, errors="coerce")

    numeric_vals = pd.to_numeric(values, errors="coerce")
    mask_excel_serial = numeric_vals.notna() & numeric_vals.between(1, 200000)
    if mask_excel_serial.any():
        out.loc[mask_excel_serial] = pd.to_datetime(
            numeric_vals.loc[mask_excel_serial],
            unit="D",
            origin="1899-12-30",
            errors="coerce",
        )

    mask_remaining = out.isna()
    if mask_remaining.any():
        out.loc[mask_remaining] = pd.to_datetime(
            values.loc[mask_remaining],
            errors="coerce",
            dayfirst=True,
        )
    return out


def process_mediana_file(
    work_path: Path,
    fechas_corte: Dict[str, datetime],
    fechas_Percentil: Dict[str, datetime],
    ideales: Dict[str, int],
    active_classes: List[str],
    db: Optional[DBIndex] = None,
    progress_cb: Optional[Callable[[int], None]] = None,
    cancel_cb: Optional[Callable[[], None]] = None,
) -> Tuple[Path, Dict[str, Any], float]:
    process_throttle = _CpuLoadThrottle(PROCESS_MAX_CPU_PERCENT)

    def check_cancel() -> None:
        process_throttle.tick()
        if cancel_cb:
            cancel_cb()

    def report(pct: int) -> None:
        check_cancel()
        if progress_cb:
            progress_cb(max(0, min(100, pct)))

    def ensure_db_loaded() -> Optional["DBIndex"]:
        nonlocal db
        if db is None:
            try:
                db = get_db()
            except Exception:
                db = None
        return db

    t0 = time.perf_counter()
    class_order_all = ["IC", "Dental", "IQ", "PROC"]
    active_order = [c for c in class_order_all if c in set(active_classes or [])]
    if not active_order:
        raise RuntimeError("Debes seleccionar al menos una clasificacion para calcular.")
    report(5)
    check_cancel()
    df = load_work_df(work_path)
    if df is None or df.empty:
        raise RuntimeError("El archivo cargado esta vacio.")
    n_rows = int(len(df))

    col_sigte = pick_col(df, ["SIGTE_ID", "sigte_id"])
    col_presta_min = pick_col(df, ["PRESTA_MIN", "presta_min"])
    col_tipo_prest = pick_col(df, ["TIPO_PREST", "tipo_prest"])
    col_presta_est = pick_col(df, ["PRESTA_EST", "presta_est"])
    col_fentrada = pick_col(df, ["F_ENTRADA", "f_entrada"])
    col_estab_dest = pick_col(df, ["ESTAB_DEST", "estab_dest", "EST_DEST", "est_dest"])

    missing: List[str] = []
    if not col_sigte:
        missing.append("SIGTE_ID")
    if not col_presta_min:
        missing.append("PRESTA_MIN")
    if not col_tipo_prest:
        missing.append("TIPO_PREST")
    if not col_presta_est:
        missing.append("PRESTA_EST")
    if not col_fentrada:
        missing.append("F_ENTRADA")
    if not col_estab_dest:
        missing.append("ESTAB_DEST")
    if missing:
        raise RuntimeError(f"Faltan columnas obligatorias: {', '.join(missing)}")
    report(20)
    check_cancel()

    df_work = df
    df_work["_SIGTE_ID_NORM"] = df_work[col_sigte].map(normalize_id)
    presta_norm_work = (
        df_work[col_presta_min].fillna("").astype(str).map(normalize_presta)
        if col_presta_min else pd.Series("", index=df_work.index, dtype=object)
    )
    special_sigte_ids: set = set()
    special_ids_loaded_from_sql = False
    try:
        with _pg_connect() as sql_conn_med:
            special_sigte_ids = _sql_fetch_comges_special_ids(
                sql_conn_med,
                df_work["_SIGTE_ID_NORM"].tolist(),
            )
            special_ids_loaded_from_sql = True
    except Exception:
        special_sigte_ids = set()
        special_ids_loaded_from_sql = False
    if not special_ids_loaded_from_sql:
        db_obj = ensure_db_loaded()
        try:
            special_sigte_ids = db_obj.comges_especiales_sigte_ids if db_obj is not None else set()
        except Exception:
            special_sigte_ids = set()
    mask_special_case = df_work["_SIGTE_ID_NORM"].map(
        lambda sid: bool(sid) and sid in special_sigte_ids
    )
    mask_salud_mental = mask_special_case & presta_norm_work.isin(
        _OBSERVACION_SALUD_MENTAL_PRESTA_CODES_NORM
    )
    df_work["Observacion caso"] = np.where(
        mask_special_case,
        np.where(mask_salud_mental, "Salud mental", "Sename"),
        "Sin observaciones",
    )
    df_work["MOTIVO_EXCLUSION_MEDIANA"] = ""
    total_input = int(len(df_work))
    mask_has_sigte = df_work["_SIGTE_ID_NORM"] != ""
    df_work.loc[~mask_has_sigte, "MOTIVO_EXCLUSION_MEDIANA"] = "Sin SIGTE_ID"
    with_sigte = int(mask_has_sigte.sum())

    estab_dest_vals = df_work[col_estab_dest].map(_to_int)
    mask_estab = mask_has_sigte & (estab_dest_vals == 106100)
    df_work.loc[mask_has_sigte & ~mask_estab, "MOTIVO_EXCLUSION_MEDIANA"] = "ESTAB_DEST distinto de 106100 o vacio"
    df_base = df_work.loc[mask_estab].copy()

    presta_vals = df_base[col_presta_min].fillna("").astype(str).str.strip()
    presta_cmp = presta_vals.str.replace(r"\s+", "", regex=True).str.upper()
    tipo_vals = df_base[col_tipo_prest].map(_to_int)
    presta_est_vals = df_base[col_presta_est].map(_to_int)

    clasif = pd.Series(index=df_base.index, dtype=object)
    mask_dental = presta_cmp.str.startswith("09-")
    clasif.loc[mask_dental] = "Dental"
    mask_ic = clasif.isna() & (tipo_vals == 1)
    clasif.loc[mask_ic] = "IC"
    mask_iq = clasif.isna() & (tipo_vals == 4)
    clasif.loc[mask_iq] = "IQ"
    mask_proc = clasif.isna() & (presta_est_vals == 3)
    clasif.loc[mask_proc] = "PROC"

    df_base["CLASIFICACION"] = clasif
    mask_classified = df_base["CLASIFICACION"].notna()
    df_base.loc[~mask_classified, "MOTIVO_EXCLUSION_MEDIANA"] = "No clasifica en IC/Dental/IQ/PROC"

    df_class = df_base[mask_classified].copy()
    classified = int(len(df_class))

    entrada_dt = _parse_excel_dates_series(df_class[col_fentrada])
    df_class["_F_ENTRADA_DT"] = entrada_dt
    df_class["ANIO_ENTRADA"] = pd.to_numeric(entrada_dt.dt.year, errors="coerce")
    report(45)
    check_cancel()

    col_run = pick_col(df_class, ["RUN", "run"])
    col_dv = pick_col(df_class, ["DV", "dv"])
    if col_run and col_dv:
        run_vals = df_class[col_run].map(normalize_run)
        dv_vals = df_class[col_dv].map(normalize_dv)
        rut_vals = np.where((run_vals != "") & (dv_vals != ""), run_vals + "-" + dv_vals, "")
        rut_vals = pd.Series(rut_vals, index=df_class.index)
    else:
        col_rut = pick_col(df_class, ["RUT CONCATENADO", "rut_concatenado", "RUT", "rut"])
        if col_rut and col_rut in df_class.columns:
            rut_vals = df_class[col_rut].fillna("").astype(str).str.strip()
        else:
            rut_vals = pd.Series("", index=df_class.index, dtype=object)
    rut_norm_vals = rut_vals.map(normalize_rut_concat)
    rut_compact_vals = rut_norm_vals.map(normalize_rut_compact)
    sql_defunciones_lookup: Dict[str, date] = {}
    sql_defunciones_loaded = False
    try:
        with _pg_connect() as sql_conn_med:
            sql_defunciones_lookup = _sql_fetch_defunciones_lookup_for_ruts(
                sql_conn_med,
                rut_compact_vals.tolist(),
            )
            sql_defunciones_loaded = True
    except Exception:
        sql_defunciones_lookup = {}
        sql_defunciones_loaded = False

    has_candidate_ruts = bool(rut_compact_vals.fillna("").astype(str).str.strip().ne("").any())
    if sql_defunciones_loaded and (not sql_defunciones_lookup) and has_candidate_ruts:
        sql_defunciones_loaded = False

    if sql_defunciones_loaded:
        df_class["_FALLECIDO"] = rut_compact_vals.map(
            lambda rut: bool(rut) and (rut in sql_defunciones_lookup)
        )
    else:
        db_obj = ensure_db_loaded()
        if db_obj is not None:
            df_class["_FALLECIDO"] = rut_norm_vals.isin(db_obj.defunciones_rut)
        else:
            df_class["_FALLECIDO"] = False
    df_class["_FALLECIDO"] = df_class["_FALLECIDO"].fillna(False).astype(bool)

    mask_sin_fecha = df_class["_F_ENTRADA_DT"].isna()
    df_class["INCLUIDO_EN_MEDIANA"] = True
    df_class["INCLUIDO_EN_MEDIANA_CON_FALLECIDOS"] = True
    df_class.loc[df_class["_FALLECIDO"], "MOTIVO_EXCLUSION_MEDIANA"] = "Paciente fallecido"
    df_class.loc[~df_class["_FALLECIDO"] & mask_sin_fecha, "MOTIVO_EXCLUSION_MEDIANA"] = "F_ENTRADA invalida o vacia"
    df_class.loc[df_class["_FALLECIDO"] | mask_sin_fecha, "INCLUIDO_EN_MEDIANA"] = False
    df_class.loc[mask_sin_fecha, "INCLUIDO_EN_MEDIANA_CON_FALLECIDOS"] = False

    active_set = set(active_order)
    mask_not_selected = ~df_class["CLASIFICACION"].isin(active_set)
    df_class.loc[mask_not_selected, "MOTIVO_EXCLUSION_MEDIANA"] = "Clasificacion no seleccionada"
    df_class.loc[mask_not_selected, "INCLUIDO_EN_MEDIANA"] = False
    df_class.loc[mask_not_selected, "INCLUIDO_EN_MEDIANA_CON_FALLECIDOS"] = False

    excluded_parts: List[pd.DataFrame] = []
    excl_sigte = df_work.loc[~mask_has_sigte]
    if not excl_sigte.empty:
        excluded_parts.append(excl_sigte)
    excl_estab = df_work.loc[mask_has_sigte & ~mask_estab]
    if not excl_estab.empty:
        excluded_parts.append(excl_estab)
    excl_clasif = df_base.loc[~mask_classified]
    if not excl_clasif.empty:
        excluded_parts.append(excl_clasif)
    excl_mediana = df_class.loc[~df_class["INCLUIDO_EN_MEDIANA"]]
    if not excl_mediana.empty:
        excluded_parts.append(excl_mediana)
    excluded_df = pd.concat(excluded_parts, ignore_index=True, sort=False, copy=False) if excluded_parts else pd.DataFrame()

    class_order = active_order
    class_sheets: Dict[str, pd.DataFrame] = {}
    for cls in class_order:
        class_sheets[cls] = df_class.loc[df_class["CLASIFICACION"] == cls]

    class_tables: Dict[str, Dict[str, Any]] = {}
    summary_rows: List[Dict[str, Any]] = []
    totals_by_class: List[Dict[str, Any]] = []
    table2_rows: List[Dict[str, Any]] = []

    for cls in class_order:
        check_cancel()
        ideal = int(ideales.get(cls, 0))
        cdf = class_sheets[cls].copy()
        fecha_corte_cls = fechas_corte.get(cls)
        fecha_Percentil_cls = fechas_Percentil.get(cls)
        if not fecha_corte_cls or not fecha_Percentil_cls:
            raise RuntimeError(f"Faltan fechas para la clasificacion {cls}.")
        fecha_corte_date = fecha_corte_cls.date()
        fecha_Percentil_date = fecha_Percentil_cls.date()
        cdf["FECHA_CORTE"] = fecha_corte_date.isoformat()
        cdf["FECHA_Percentil"] = fecha_Percentil_date.isoformat()
        cdf["DIAS_EN_LISTA"] = (pd.Timestamp(fecha_corte_date) - cdf["_F_ENTRADA_DT"]).dt.days
        included_mask = cdf.get("INCLUIDO_EN_MEDIANA", pd.Series([False] * len(cdf), index=cdf.index)).fillna(False).astype(bool)
        included_with_dead_mask = cdf.get(
            "INCLUIDO_EN_MEDIANA_CON_FALLECIDOS",
            pd.Series([False] * len(cdf), index=cdf.index),
        ).fillna(False).astype(bool)
        dias_all = pd.to_numeric(cdf.get("DIAS_EN_LISTA"), errors="coerce")
        dias_vals = dias_all[included_mask].dropna()
        dias_vals_with_dead = dias_all[included_with_dead_mask].dropna()
        mediana_general = round(float(dias_vals.median()), 1) if not dias_vals.empty else None
        mediana_general_con_fallecidos = round(float(dias_vals_with_dead.median()), 1) if not dias_vals_with_dead.empty else None
        diferencia_general = round(mediana_general - ideal, 1) if mediana_general is not None else None
        diferencia_general_con_fallecidos = (
            round(mediana_general_con_fallecidos - ideal, 1)
            if mediana_general_con_fallecidos is not None
            else None
        )
        Percentil_count = int((included_mask & (cdf["_F_ENTRADA_DT"] <= pd.Timestamp(fecha_Percentil_date))).sum()) if not cdf.empty else 0

    
        cdf["ESTADO_MEDIANA_Percentil"] = "Excluido del calculo"
        if not cdf.empty:
            Percentil_mask = included_mask & (cdf["_F_ENTRADA_DT"] <= pd.Timestamp(fecha_Percentil_date))
            cdf.loc[Percentil_mask, "ESTADO_MEDIANA_Percentil"] = "Caso Obligatorio a Egresar"

            remaining_mask = included_mask & ~Percentil_mask
            if mediana_general is not None:
                dias_all = pd.to_numeric(cdf.get("DIAS_EN_LISTA"), errors="coerce")
                sup_mask = remaining_mask & dias_all.ge(float(mediana_general))
                inf_mask = remaining_mask & dias_all.lt(float(mediana_general))
                cdf.loc[sup_mask, "ESTADO_MEDIANA_Percentil"] = "Superior a la mediana"
                cdf.loc[inf_mask, "ESTADO_MEDIANA_Percentil"] = "Inferior a la mediana"
            else:
                cdf.loc[remaining_mask, "ESTADO_MEDIANA_Percentil"] = "Sin referencia de mediana"

        by_year_rows: List[Dict[str, Any]] = []
        if not cdf.empty:
            tmp_year = cdf[cdf["ANIO_ENTRADA"].notna()]
            if not tmp_year.empty:
                grp = (
                    tmp_year.groupby("ANIO_ENTRADA")
                    .agg(Casos=("ANIO_ENTRADA", "size"), Fallecidos=("_FALLECIDO", "sum"))
                    .reset_index()
                    .sort_values("ANIO_ENTRADA")
                )
                for _, row in grp.iterrows():
                    yrow = {
                        "Anio": int(row["ANIO_ENTRADA"]),
                        "Casos": int(row["Casos"]),
                        "Fallecidos": int(row["Fallecidos"]),
                    }
                    by_year_rows.append(yrow)
                    summary_rows.append({
                        "Clasificacion": cls,
                        "Anio": yrow["Anio"],
                        "Casos": yrow["Casos"],
                        "Fallecidos": yrow["Fallecidos"],
                    })

        general = {
            "Clasificacion": cls,
            "Casos": int(len(cdf)),
            "Fallecidos": int(cdf["_FALLECIDO"].sum()) if not cdf.empty else 0,
            "Mediana": mediana_general,
            "Mediana_con_fallecidos": mediana_general_con_fallecidos,
            "Ideal": ideal,
            "Diferencia": diferencia_general,
            "Diferencia_con_fallecidos": diferencia_general_con_fallecidos,
            "Fecha_corte": fecha_corte_date.isoformat(),
            "Fecha_Percentil": fecha_Percentil_date.isoformat(),
            "Casos_Percentil": Percentil_count,
        }
        class_tables[cls] = {
            "by_year": by_year_rows,
            "general": general,
        }
        class_sheets[cls] = cdf
        totals_by_class.append(general)
        table2_rows.append({
            "Clasificacion": cls,
            "Casos": general["Casos"],
            "Fallecidos": general["Fallecidos"],
            "Mediana": general["Mediana"],
            "Mediana_con_fallecidos": general["Mediana_con_fallecidos"],
            "Ideal": general["Ideal"],
            "Diferencia": general["Diferencia"],
            "Diferencia_con_fallecidos": general["Diferencia_con_fallecidos"],
            "Fecha_corte": general["Fecha_corte"],
            "Fecha_Percentil": general["Fecha_Percentil"],
            "Casos_Percentil": general["Casos_Percentil"],
        })
        summary_rows.append({
            "Clasificacion": cls,
            "Anio": "TOTAL",
            "Casos": general["Casos"],
            "Fallecidos": general["Fallecidos"],
        })
    report(70)
    check_cancel()

    out_name = f"LE_NOGES_mediana_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}.xlsx"
    out_path = OUTPUT_DIR / out_name
    check_cancel()
    with pd.ExcelWriter(
        out_path,
        engine="openpyxl",
        date_format="DD-MM-YYYY",
        datetime_format="DD-MM-YYYY",
    ) as writer:
        for cls in class_order:
            export_df = class_sheets[cls].drop(
                columns=[
                    "_SIGTE_ID_NORM",
                    "SIGTE_ID_NORMALIZADO",
                    "_FALLECIDO",
                    "INCLUIDO_EN_MEDIANA",
                    "INCLUIDO_EN_MEDIANA_CON_FALLECIDOS",
                    "_F_ENTRADA_DT",
                ],
                errors="ignore",
            )
            export_df = format_dates_for_export(export_df)
            export_df.to_excel(writer, sheet_name=cls, index=False)
        excluded_export = excluded_df.drop(
            columns=[
                "_SIGTE_ID_NORM",
                "SIGTE_ID_NORMALIZADO",
                "_FALLECIDO",
                "INCLUIDO_EN_MEDIANA",
                "INCLUIDO_EN_MEDIANA_CON_FALLECIDOS",
                "_F_ENTRADA_DT",
            ],
            errors="ignore",
        )
        excluded_export = format_dates_for_export(excluded_export)
        excluded_export.to_excel(writer, sheet_name="Excluidos", index=False)
        format_dates_for_export(pd.DataFrame(table2_rows)).to_excel(writer, sheet_name="Resumen", index=False)
        apply_short_date_format_to_workbook(writer.book)
    report(95)

    elapsed = time.perf_counter() - t0
    included_mediana = int(df_class["INCLUIDO_EN_MEDIANA"].sum()) if not df_class.empty else 0
    included_mediana_con_fallecidos = (
        int(df_class["INCLUIDO_EN_MEDIANA_CON_FALLECIDOS"].sum()) if not df_class.empty else 0
    )
    excluded_total = int(total_input - included_mediana)

    stats = {
        "fechas_corte": {k: v.date().isoformat() for k, v in fechas_corte.items()},
        "fechas_Percentil": {k: v.date().isoformat() for k, v in fechas_Percentil.items()},
        "total_input": total_input,
        "with_sigte": with_sigte,
        "classified": classified,
        "excluded": excluded_total,
        "included_mediana": included_mediana,
        "included_mediana_con_fallecidos": included_mediana_con_fallecidos,
        "summary": summary_rows,
        "summary_total": totals_by_class,
        "class_order": class_order,
        "class_tables": class_tables,
    }
    report(100)
    return out_path, stats, elapsed


def _map_tipo_prest(v: Any) -> str:
    val = _to_int(v)
    if val in (1, 2):
        return "IC"
    if val == 3:
        return "PROC"
    if val in (4, 5):
        return "IQ"
    return "OTRO"


def _calc_age(dt: Optional[datetime], ref: datetime) -> Optional[int]:
    if dt is None:
        return None
    try:
        if pd.isna(dt):
            return None
    except Exception:
        pass
    try:
        years = ref.year - dt.year - ((ref.month, ref.day) < (dt.month, dt.day))
        return years if years >= 0 else None
    except Exception:
        return None


def _age_range(age: Optional[int]) -> str:
    if age is None:
        return "Sin edad"
    if age <= 17:
        return "0-17"
    if age <= 64:
        return "18-64"
    return "65+"


def _load_df_any(path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df_raw = _read_csv_raw_df(path)
        if df_raw is None or df_raw.empty:
            return pd.DataFrame()
        header_row = detect_header_row_df(df_raw)
        header = [str(c).strip() if str(c).strip() else f"COL_{i}" for i, c in enumerate(df_raw.iloc[header_row - 1].tolist())]
        df = df_raw.iloc[header_row:].copy()
        df.columns = header
        return df

    engine = "pyxlsb" if path.suffix.lower() == ".xlsb" else "openpyxl"
    try:
        df_raw = pd.read_excel(path, engine=engine, header=None, sheet_name=sheet_name, dtype=object)
    except ImportError as e:
        if engine == "pyxlsb":
            raise RuntimeError(
                "Para leer archivos .xlsb, instala la dependencia 'pyxlsb' (pip install pyxlsb)."
            ) from e
        raise
    if isinstance(df_raw, dict):
        if not df_raw:
            return pd.DataFrame()
        df_raw = next(iter(df_raw.values()))
    if df_raw is None or df_raw.empty:
        return pd.DataFrame()
    header_row = detect_header_row_df(df_raw)
    header = [str(c).strip() if str(c).strip() else f"COL_{i}" for i, c in enumerate(df_raw.iloc[header_row - 1].tolist())]
    df = df_raw.iloc[header_row:].copy()
    df.columns = header
    return df


def _kpi_by_group(df: pd.DataFrame, id_col: pd.Series, group_col: pd.Series, is_closed: pd.Series) -> List[Dict[str, Any]]:
    tmp = pd.DataFrame({
        "id": id_col,
        "group": group_col,
        "closed": is_closed
    })
    tmp = tmp[tmp["id"] != ""]
    out: List[Dict[str, Any]] = []
    for group in sorted(tmp["group"].dropna().unique()):
        open_count = int(tmp[(tmp["group"] == group) & (~tmp["closed"])].shape[0])
        closed_count = int(tmp[(tmp["group"] == group) & (tmp["closed"])].shape[0])
        total = int(open_count + closed_count)
        out.append({
            "group": str(group),
            "open": int(open_count),
            "closed": int(closed_count),
            "total": total,
            "open_pct": round((open_count / total) * 100, 1) if total else 0.0,
            "closed_pct": round((closed_count / total) * 100, 1) if total else 0.0,
        })
    return out


def _kpi_closed_by_group(
    group_col: pd.Series,
    is_closed: pd.Series,
    allowed_values: Optional[Iterable[str]] = None
) -> List[Dict[str, Any]]:
    tmp = pd.DataFrame({
        "group": group_col,
        "closed": is_closed
    })
    closed_mask = tmp["closed"]
    try:
        closed_mask = closed_mask.fillna(False)
    except Exception:
        pass
    total_closed = int(closed_mask.sum())
    tmp = tmp[closed_mask].copy()
    if tmp.empty:
        return []
    tmp["group"] = tmp["group"].fillna("").astype(str).str.strip()
    tmp = tmp[tmp["group"] != ""]
    allowed_list: Optional[List[str]] = None
    if allowed_values is not None:
        allowed_list = [str(v) for v in allowed_values]
        allowed_set = set(allowed_list)
        tmp = tmp[tmp["group"].isin(allowed_set)]
        grouped = tmp.groupby("group").size()
        grouped = grouped.reindex(allowed_list, fill_value=0)
        grouped = grouped.reset_index(name="count")
    else:
        if tmp.empty:
            return []
        grouped = tmp.groupby("group").size().reset_index(name="count")
        grouped = grouped.sort_values("count", ascending=False)
    out: List[Dict[str, Any]] = []
    for _, row in grouped.iterrows():
        group = row["group"]
        count = int(row["count"])
        pct = round((count / total_closed) * 100, 1) if total_closed else 0.0
        out.append({
            "group": str(group),
            "count": count,
            "total": total_closed,
            "pct": pct
        })
    return out


def _slice_preview_rows(rows: List[Dict[str, Any]], limit: int = TABLE_PREVIEW_LIMIT) -> Tuple[List[Dict[str, Any]], int]:
    total = len(rows)
    if total <= limit:
        return rows, total
    return rows[:limit], total


def _unique_rows(rows: List[Dict[str, Any]], keys: List[str]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    seen: set = set()
    for row in rows:
        key = tuple(str(row.get(k, "")) for k in keys)
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def _is_nomina_found(value: Any) -> bool:
    c = canon(value)
    if not c:
        return False
    if "noseencuentra" in c:
        return False
    if "sinidlocalsigteid" in c:
        return False
    return True


def _map_nomina_source_label(source: Any) -> str:
    if source is None:
        return "Otro"
    src = str(source).strip()
    if not src:
        return "Otro"
    csrc = canon(src)
    if "historico" in csrc:
        return "Otro"

    file_part = src
    sheet_part = ""
    if ":" in src:
        file_part, sheet_part = src.split(":", 1)

    tipo: Optional[str] = None
    estado: Optional[str] = None

    parsed = _parse_nomina_filename(Path(file_part))
    if parsed:
        tipo = parsed[0]
        estado = parsed[1]

    if not tipo:
        stem = canon(Path(file_part).stem)
        if "proc" in stem:
            tipo = "proc"
        elif "iq" in stem:
            tipo = "iq"
        elif "cne" in stem or "ic" in stem:
            tipo = "cne"

    if not estado and sheet_part:
        if _sheet_matches_estado(sheet_part, "abierto"):
            estado = "abierto"
        elif _sheet_matches_estado(sheet_part, "cerrado"):
            estado = "cerrado"
        else:
            sheet_norm = canon(sheet_part)
            if "abierto" in sheet_norm:
                estado = "abierto"
            elif "cerrado" in sheet_norm:
                estado = "cerrado"

    if tipo not in ("cne", "iq", "proc") or estado not in ("abierto", "cerrado"):
        return "Otro"

    tipo_label = "IC" if tipo == "cne" else tipo.upper()
    estado_label = "abierta" if estado == "abierto" else "cerrada"
    return f"Nomina {tipo_label} {estado_label}"


def _format_timeline_date(value: Optional[datetime]) -> str:
    if value is None:
        return ""
    try:
        return value.strftime("%d-%m-%Y")
    except Exception:
        return ""


def _search_timeline_matches(
    timeline_map: Dict[str, List[TimelineRec]],
    rut_filter: str = "",
    id_filter: str = "",
    source_label_fn: Optional[Callable[[str], str]] = None,
) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    seen: set = set()

    run_filter = ""
    dv_filter = ""
    if rut_filter and "-" in rut_filter:
        run_raw, dv_raw = rut_filter.split("-", 1)
        run_filter = normalize_run(run_raw)
        dv_filter = normalize_dv(dv_raw[:1])

    for key, recs in timeline_map.items():
        parts = str(key).split("|", 2)
        if len(parts) < 3:
            continue
        run, dv, presta = parts[0], parts[1], parts[2]
        row_rut = f"{run}-{dv}" if run and dv else ""

        if run_filter and dv_filter and (run != run_filter or dv != dv_filter):
            continue
        if rut_filter and (not run_filter or not dv_filter):
            if normalize_rut_concat(row_rut) != rut_filter:
                continue

        for rec in recs:
            id_local = normalize_id(rec.id_local)
            sigte_id = normalize_id(rec.sigte_id)
            if id_filter and (id_filter not in {id_local, sigte_id}):
                continue

            source_raw = str(rec.source or "").strip()
            source_label = source_raw
            if source_label_fn:
                source_label = source_label_fn(source_raw)
            table_name = ""
            if source_raw:
                file_part = source_raw.split(":", 1)[0].strip()
                if file_part.lower().endswith(".sql"):
                    table_name = _pg_table_from_pseudo_path(Path(file_part))
                else:
                    table_name = file_part

            row = {
                "rut": row_rut,
                "id_local": id_local,
                "sigte_id": sigte_id,
                "presta_min": str(presta or ""),
                "f_entrada": _format_timeline_date(rec.f_in),
                "f_salida": _format_timeline_date(rec.f_out),
                "extremidad": str(rec.extremidad or ""),
                "origen": str(source_label or ""),
                "tabla": str(table_name or ""),
            }
            uniq = (
                row["rut"],
                row["id_local"],
                row["sigte_id"],
                row["presta_min"],
                row["f_entrada"],
                row["f_salida"],
                row["extremidad"],
                row["origen"],
                row["tabla"],
            )
            if uniq in seen:
                continue
            seen.add(uniq)
            rows.append(row)

    rows.sort(
        key=lambda r: (
            r["rut"],
            r["id_local"],
            r["sigte_id"],
            r["presta_min"],
            r["f_entrada"],
            r["f_salida"],
            r["origen"],
        )
    )
    return rows


def _search_safe_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def _search_format_date(value: Any) -> str:
    dt = parse_excel_date(value)
    if dt is not None:
        return dt.strftime("%d-%m-%Y")
    return _search_safe_text(value)


def _search_rut_norm_series(df: pd.DataFrame) -> pd.Series:
    col_rut = pick_col(df, ["RUT CONCATENADO", "rut_concatenado", "RUTCONCATENADO", "RUT", "rut"])
    if col_rut and col_rut in df.columns:
        rut_vals = df[col_rut].fillna("").astype(str).str.strip()
    else:
        col_run = pick_col(df, ["RUN", "run"])
        col_dv = pick_col(df, ["DV", "dv"])
        if col_run and col_dv and col_run in df.columns and col_dv in df.columns:
            run_vals = df[col_run].map(normalize_run)
            dv_vals = df[col_dv].map(normalize_dv)
            rut_vals = pd.Series(
                np.where((run_vals != "") & (dv_vals != ""), run_vals + "-" + dv_vals, ""),
                index=df.index
            )
        else:
            rut_vals = pd.Series([""] * len(df), index=df.index)
    return rut_vals.map(normalize_rut_concat)


def _search_id_norm_series(df: pd.DataFrame, candidates: List[str]) -> pd.Series:
    col = pick_col(df, candidates)
    if not col or col not in df.columns:
        return pd.Series([""] * len(df), index=df.index)
    return df[col].map(normalize_id)


def _nomina_search_label(tipo: str, estado: str) -> str:
    tipo_label = "IC" if tipo == "cne" else str(tipo).upper()
    estado_label = "abierta" if estado == "abierto" else "cerrada"
    return f"Nomina {tipo_label} {estado_label}"


def _search_cases_full_rows(
    rut_filter: str = "",
    id_filter: str = "",
    build_preview: bool = True,
) -> Tuple[List[Dict[str, str]], pd.DataFrame, Dict[str, int]]:
    rows: List[Dict[str, str]] = []
    export_frames: List[pd.DataFrame] = []
    counts = {"historico": 0, "nominas": 0}

    sources: List[Tuple[str, str, str]] = [
        ("historico", PG_BASE_TABLES["historico"], "Historico"),
    ]
    for (tipo, estado), table_name in PG_NOMINA_TABLES.items():
        sources.append(("nominas", table_name, _nomina_search_label(tipo, estado)))

    with _pg_connect() as conn:
        for source_kind, table_name, source_label in sources:
            df = _sql_fetch_filtered_table_df(
                conn,
                table_name=table_name,
                rut_filter=rut_filter,
                id_filter=id_filter,
            )
            if df is None or df.empty:
                continue

            rut_norm = _search_rut_norm_series(df)
            id_local_norm = _search_id_norm_series(df, ["ID_LOCAL", "id_local"])
            sigte_norm = _search_id_norm_series(df, ["SIGTE_ID", "sigte_id"])

            matched = df.copy()
            if matched.empty:
                continue

            matched_count = int(len(matched))
            if source_kind == "historico":
                counts["historico"] += matched_count
            else:
                counts["nominas"] += matched_count

            matched["ORIGEN_BUSQUEDA"] = source_label
            matched["TABLA_ORIGEN"] = table_name
            matched["RUT_NORMALIZADO"] = rut_norm.values
            matched["ID_LOCAL_NORMALIZADO"] = id_local_norm.values
            matched["SIGTE_ID_NORMALIZADO"] = sigte_norm.values

            meta_cols = [
                "ORIGEN_BUSQUEDA",
                "TABLA_ORIGEN",
                "RUT_NORMALIZADO",
                "ID_LOCAL_NORMALIZADO",
                "SIGTE_ID_NORMALIZADO",
            ]
            ordered_cols = meta_cols + [c for c in matched.columns if c not in meta_cols]
            matched = matched[ordered_cols]
            export_frames.append(matched)

            if build_preview:
                presta_col = pick_col(matched, ["PRESTA_MIN", "presta_min"])
                fin_col = pick_col(matched, ["F_ENTRADA", "f_entrada"])
                fout_col = pick_col(matched, ["F_SALIDA", "f_salida"])
                ext_col = pick_col(matched, ["EXTREMIDAD", "extremidad"])

                preview_df = pd.DataFrame(index=matched.index)
                preview_df["rut"] = matched["RUT_NORMALIZADO"].map(_search_safe_text)
                preview_df["id_local"] = matched["ID_LOCAL_NORMALIZADO"].map(_search_safe_text)
                preview_df["sigte_id"] = matched["SIGTE_ID_NORMALIZADO"].map(_search_safe_text)
                preview_df["presta_min"] = matched[presta_col].map(_search_safe_text) if presta_col else ""
                preview_df["f_entrada"] = matched[fin_col].map(_search_format_date) if fin_col else ""
                preview_df["f_salida"] = matched[fout_col].map(_search_format_date) if fout_col else ""
                preview_df["extremidad"] = matched[ext_col].map(_search_safe_text) if ext_col else ""
                preview_df["origen"] = source_label
                preview_df["tabla"] = table_name
                rows.extend(preview_df.to_dict(orient="records"))

    if build_preview:
        rows.sort(
            key=lambda r: (
                str(r.get("rut", "")),
                str(r.get("id_local", "")),
                str(r.get("sigte_id", "")),
                str(r.get("origen", "")),
                str(r.get("f_entrada", "")),
            )
        )

    export_df = pd.concat(export_frames, ignore_index=True, sort=False) if export_frames else pd.DataFrame()
    counts["total"] = int(counts["historico"] + counts["nominas"])
    return rows, export_df, counts


def build_cross_statistics(
    work_path: Optional[Path],
    db: DBIndex,
    progress_cb: Optional[Callable[[int], None]] = None,
    cancel_cb: Optional[Callable[[], None]] = None,
) -> Dict[str, Any]:
    def check_cancel() -> None:
        if cancel_cb:
            cancel_cb()

    def report(pct: int) -> None:
        check_cancel()
        if progress_cb:
            progress_cb(max(0, min(100, pct)))

    report(2)
    if not work_path or not work_path.exists():
        return {"error": "Debes cargar un archivo para generar estadisticas de cruces."}

    selected_all = {
        "historico": True,
        "nominas": True,
        "verificacion": True,
        "traslape": True,
        "duplicidad": True,
        "cgr": True,
        "defunciones": True,
        "macrored": True,
    }

    def cb_process(p: int) -> None:
        report(5 + int(p * 0.75))

    processed_path, elapsed, _stage_timing = process_file(
        work_path,
        selected_all,
        db,
        progress_cb=cb_process,
        cancel_cb=check_cancel,
    )
    report(82)
    df = load_work_df(processed_path)
    if df is None or df.empty:
        return {"error": "No se generaron resultados para el archivo cargado."}

    total_records = int(len(df))

    def text_series(candidates: List[str]) -> pd.Series:
        col = pick_col(df, candidates)
        if not col or col not in df.columns:
            return pd.Series([""] * total_records)
        return df[col].fillna("").astype(str).str.strip()

    id_vals = text_series(["ID_LOCAL", "id_local"]).map(normalize_id)
    rut_vals = text_series(["RUT CONCATENADO", "rut_concatenado"])
    if rut_vals.eq("").all():
        run_vals = text_series(["RUN", "run"]).map(normalize_run)
        dv_vals = text_series(["DV", "dv"]).map(normalize_dv)
        rut_vals = run_vals + "-" + dv_vals
    rut_norm_vals = rut_vals.map(normalize_rut_concat)

    nom_sigte_vals = text_series(["CRUCE NOMINAS (SIGTE_ID)"])
    nom_source_vals = text_series(["ORIGEN SIGTE_ID"])

    # 1) Historico
    hist_vals = text_series(["CRUCE CERRADAS HISTORICAS"])
    hist_found_mask = hist_vals.map(lambda v: canon(v) == "seencuentraenhistorico")
    hist_found = int(hist_found_mask.sum())
    hist_rows: List[Dict[str, Any]] = []
    for i in range(total_records):
        check_cancel()
        if not bool(hist_found_mask.iat[i]):
            continue
        id_local = str(id_vals.iat[i])
        sigte_id = db.historico_by_id_map.get(normalize_id(id_local), "") if id_local else ""
        if not sigte_id:
            sigte_nom = str(nom_sigte_vals.iat[i]) if i < len(nom_sigte_vals) else ""
            if _is_nomina_found(sigte_nom):
                sigte_id = sigte_nom
        hist_rows.append({
            "rut": str(rut_norm_vals.iat[i] or rut_vals.iat[i]),
            "id_local": id_local,
            "sigte_id": sigte_id
        })
    hist_rows = _unique_rows(hist_rows, ["rut", "id_local", "sigte_id"])
    hist_rows_preview, hist_rows_total = _slice_preview_rows(hist_rows)

    # 2) Nominas
    nom_found_mask = nom_sigte_vals.map(_is_nomina_found)
    nom_found = int(nom_found_mask.sum())
    nom_rows: List[Dict[str, Any]] = []
    origin_labels = [
        "Nomina IC abierta",
        "Nomina IQ abierta",
        "Nomina PROC abierta",
        "Nomina IC cerrada",
        "Nomina IQ cerrada",
        "Nomina PROC cerrada",
    ]
    origin_counts: Dict[str, int] = {k: 0 for k in origin_labels}
    for i in range(total_records):
        check_cancel()
        if not bool(nom_found_mask.iat[i]):
            continue
        src_label = _map_nomina_source_label(nom_source_vals.iat[i])
        if src_label in origin_counts:
            origin_counts[src_label] += 1
        nom_rows.append({
            "rut": str(rut_norm_vals.iat[i] or rut_vals.iat[i]),
            "id_local": str(id_vals.iat[i]),
            "sigte_id": str(nom_sigte_vals.iat[i]),
            "origen": src_label
        })
    nom_rows = _unique_rows(nom_rows, ["rut", "id_local", "sigte_id", "origen"])
    nom_rows_preview, nom_rows_total = _slice_preview_rows(nom_rows)
    nom_origen = [{"label": lbl, "count": int(origin_counts[lbl])} for lbl in origin_labels]

    # 3) Verificacion de datos
    ver_vals = text_series(["VERIFICACION DE DATOS"])
    ver_no_found = 0
    ver_ok = 0
    ver_problem = 0
    ver_rows: List[Dict[str, Any]] = []
    for i in range(total_records):
        check_cancel()
        msg = str(ver_vals.iat[i])
        cm = canon(msg)
        if "noencontrado" in cm:
            ver_no_found += 1
            continue
        if "sinproblemas" in cm:
            ver_ok += 1
            continue
        if cm:
            ver_problem += 1
            id_local = str(id_vals.iat[i])
            sigte = str(nom_sigte_vals.iat[i]) if i < len(nom_sigte_vals) else ""
            if not _is_nomina_found(sigte):
                sigte = db.get_nomina_sigte(normalize_id(id_local))
            ver_rows.append({
                "rut": str(rut_norm_vals.iat[i] or rut_vals.iat[i]),
                "id_local": id_local,
                "sigte_id": sigte,
                "detalle": msg
            })
    ver_rows = _unique_rows(ver_rows, ["rut", "id_local", "sigte_id", "detalle"])
    ver_rows_preview, ver_rows_total = _slice_preview_rows(ver_rows)

    # 4) Traslape + Duplicidad + Alerta cercano
    tras_vals = text_series(["TRASLAPE"])
    dup_vals = text_series(["DUPLICIDAD"])
    alerta_vals = text_series([
        "ALERTA CASO CERCANO (< 1 ano)",
        "ALERTA CASO CERCANO (< 1 aÃ±o)",
        "ALERTA CASO CERCANO (< 1 aÃƒÂ±o)"
    ])

    dup_mask = dup_vals.map(lambda v: "casoduplicado" in canon(v))
    dup_count = int(dup_mask.sum())
    dup_ok = int(total_records - dup_count)

    alerta_mask = alerta_vals.map(lambda v: canon(v).startswith("alerta"))
    alerta_count = int(alerta_mask.sum())
    alerta_ok = int(total_records - alerta_count)

    tras_issue_mask = tras_vals.map(lambda v: "casotraslape" in canon(v))
    td_rows: List[Dict[str, Any]] = []
    for i in range(total_records):
        check_cancel()
        if not bool(tras_issue_mask.iat[i] or dup_mask.iat[i]):
            continue
        td_rows.append({
            "rut": str(rut_norm_vals.iat[i] or rut_vals.iat[i]),
            "id_local": str(id_vals.iat[i]),
            "traslape": str(tras_vals.iat[i]),
            "duplicidad": str(dup_vals.iat[i]),
        })
    td_rows = _unique_rows(td_rows, ["rut", "id_local", "traslape", "duplicidad"])
    td_rows_preview, td_rows_total = _slice_preview_rows(td_rows)

    # 5) Macro red
    macro_vals = text_series(["CRUCE ESTABLECIMIENTOS"])
    macro_mask = macro_vals.map(lambda v: canon(v) == "correspondeamacrored")
    macro_count = int(macro_mask.sum())
    macro_ok = int(total_records - macro_count)

    # 6) CGR
    cgr399_vals = text_series(["CRUCE CGR 399"])
    cgr84_vals = text_series(["CRUCE CGR 84"])
    cgr_rows: List[Dict[str, Any]] = []
    cgr_labels = [
        "CGR 399-ANEXO 9",
        "CGR 399-ANEXO 11",
        "CGR 84-ANEXO 13",
        "CGR 84-ANEXO 14",
        "CGR 84-ANEXO 17",
        "CGR 84-ANEXO 19",
    ]
    cgr_counts: Dict[str, int] = {k: 0 for k in cgr_labels}
    for i in range(total_records):
        check_cancel()
        raw_399 = str(cgr399_vals.iat[i])
        raw_84 = str(cgr84_vals.iat[i])
        c399 = canon(raw_399)
        c84 = canon(raw_84)
        found_399 = bool(c399) and ("noseencuentraencgr399" not in c399)
        found_84 = bool(c84) and ("noseencuentraencgr84" not in c84)
        if not (found_399 or found_84):
            continue

        combined = f"{c399} {c84}"
        if "399" in combined and "anexo9" in combined:
            cgr_counts["CGR 399-ANEXO 9"] += 1
        if "399" in combined and "anexo11" in combined:
            cgr_counts["CGR 399-ANEXO 11"] += 1
        if "84" in combined and "anexo13" in combined:
            cgr_counts["CGR 84-ANEXO 13"] += 1
        if "84" in combined and "anexo14" in combined:
            cgr_counts["CGR 84-ANEXO 14"] += 1
        if "84" in combined and "anexo17" in combined:
            cgr_counts["CGR 84-ANEXO 17"] += 1
        if "84" in combined and "anexo19" in combined:
            cgr_counts["CGR 84-ANEXO 19"] += 1

        cgr_rows.append({
            "rut": str(rut_vals.iat[i]),
            "id_local": str(id_vals.iat[i]),
            "cgr_399": raw_399,
            "cgr_84": raw_84,
        })

    cgr_rows = _unique_rows(cgr_rows, ["rut", "id_local", "cgr_399", "cgr_84"])
    cgr_rows_preview, cgr_rows_total = _slice_preview_rows(cgr_rows)
    cgr_anexos = [{"label": lbl, "count": int(cgr_counts[lbl])} for lbl in cgr_labels]

    # 7) Defunciones
    def_vals = text_series(["CRUCE DEFUNCIONES"])
    dead_mask = def_vals.map(lambda v: "fallecido" in canon(v))
    dead_count = int(dead_mask.sum())
    alive_count = int(total_records - dead_count)
    dead_rows: List[Dict[str, Any]] = []
    for i in range(total_records):
        check_cancel()
        if not bool(dead_mask.iat[i]):
            continue
        rut_norm = str(rut_norm_vals.iat[i])
        dead_rows.append({
            "rut": rut_norm or str(rut_vals.iat[i]),
            "fecha_def": db.defunciones_fecha.get(rut_norm, "")
        })
    dead_rows = _unique_rows(dead_rows, ["rut", "fecha_def"])
    dead_rows_preview, dead_rows_total = _slice_preview_rows(dead_rows)

    report(100)
    return {
        "source_label": work_path.name,
        "processed_file": processed_path.name,
        "elapsed_display": format_duration(elapsed),
        "total_records": total_records,
        "historico": {
            "found": hist_found,
            "not_found": int(total_records - hist_found),
            "rows": hist_rows_preview,
            "rows_total": hist_rows_total,
        },
        "nominas": {
            "found": nom_found,
            "not_found": int(total_records - nom_found),
            "rows": nom_rows_preview,
            "rows_total": nom_rows_total,
            "origen": nom_origen,
        },
        "verificacion": {
            "no_encontrado": int(ver_no_found),
            "ok": int(ver_ok),
            "problema": int(ver_problem),
            "rows": ver_rows_preview,
            "rows_total": ver_rows_total,
        },
        "traslape_duplicidad": {
            "duplicidad": int(dup_count),
            "ok_duplicidad": int(dup_ok),
            "alerta": int(alerta_count),
            "ok_alerta": int(alerta_ok),
            "rows": td_rows_preview,
            "rows_total": td_rows_total,
        },
        "macrored": {
            "ok": int(macro_ok),
            "macrored": int(macro_count),
        },
        "cgr": {
            "rows": cgr_rows_preview,
            "rows_total": cgr_rows_total,
            "anexos": cgr_anexos,
        },
        "defunciones": {
            "vivos": int(alive_count),
            "fallecidos": int(dead_count),
            "rows": dead_rows_preview,
            "rows_total": dead_rows_total,
        },
    }


def build_statistics(
    source: str,
    work_path: Optional[Path],
    db: DBIndex,
    progress_cb: Optional[Callable[[int], None]] = None,
    cancel_cb: Optional[Callable[[], None]] = None,
) -> Dict[str, Any]:
    process_throttle = _CpuLoadThrottle(PROCESS_MAX_CPU_PERCENT)

    def check_cancel() -> None:
        process_throttle.tick()
        if cancel_cb:
            cancel_cb()

    def report(pct: int) -> None:
        check_cancel()
        if progress_cb:
            progress_cb(max(0, min(100, pct)))

    report(3)
    if source != "archivo":
        return {"error": "Fuente de datos invalida. Solo se permite archivo cargado."}
    if not work_path or not work_path.exists():
        return {"error": "Debes cargar un archivo para generar estadísticas."}
    df = _load_df_any(work_path)
    source_label = work_path.name

    if df is None or df.empty:
        return {"error": "No se encontraron registros para la fuente seleccionada."}
    report(12)

    col_idlocal = pick_col(df, ["ID_LOCAL", "id_local"])
    col_tipo = pick_col(df, ["TIPO_PREST", "tipo_prest"])
    col_fsalida = pick_col(df, ["F_SALIDA", "f_salida"])
    col_sexo = pick_col(df, ["SEXO", "sexo"])
    col_fnac = pick_col(df, ["FECHA_NAC", "fecha_nac"])
    col_presta_est = pick_col(df, ["PRESTA_EST", "presta_est"])
    col_run = pick_col(df, ["RUN", "run"])
    col_dv = pick_col(df, ["DV", "dv"])
    col_estab = pick_col(df, ["ESTAB_DEST", "estab_dest", "ESTAB_ORIG", "estab_orig"])
    col_csal = pick_col(df, ["C_SALIDA", "c_salida"])
    col_presta_min = pick_col(df, ["PRESTA_MIN", "presta_min"])
    col_fin = pick_col(df, ["F_ENTRADA", "f_entrada"])
    col_ext = pick_col(df, ["EXTREMIDAD", "extremidad"])

    missing = []
    for label, col in [("ID_LOCAL", col_idlocal), ("TIPO_PREST", col_tipo), ("F_SALIDA", col_fsalida)]:
        if not col:
            missing.append(label)

    if col_idlocal:
        idlocal_vals_full = df[col_idlocal].fillna("").map(normalize_id)
        base_mask = idlocal_vals_full != ""
        try:
            base_mask = base_mask.fillna(False)
        except Exception:
            pass
        df = df[base_mask].copy()
        if df.empty:
            return {"error": "No hay registros con ID_LOCAL para generar estadísticas."}
    else:
        idlocal_vals_full = pd.Series([""] * len(df))

    idlocal_vals = df[col_idlocal].fillna("").map(normalize_id) if col_idlocal else pd.Series([""] * len(df))
    tipo_vals = df[col_tipo].fillna("").map(_map_tipo_prest) if col_tipo else pd.Series(["OTRO"] * len(df))
    fsalida_vals = df[col_fsalida].map(normalize_date) if col_fsalida else pd.Series([None] * len(df))
    is_closed = fsalida_vals.notna()

    summary = _kpi_by_group(df, idlocal_vals, tipo_vals, is_closed) if not missing else []
    report(25)

    sexo_vals = df[col_sexo].fillna("").map(normalize_text) if col_sexo else pd.Series(["Sin dato"] * len(df))
    sexo_vals = sexo_vals.replace("", "Sin dato")
    kpi_sexo = _kpi_by_group(df, idlocal_vals, sexo_vals, is_closed)
    report(35)

    ref_date = datetime.now()
    fnac_vals = df[col_fnac].map(normalize_date) if col_fnac else pd.Series([None] * len(df))
    age_vals = fnac_vals.map(lambda d: _calc_age(d, ref_date))
    age_group = age_vals.map(_age_range)
    kpi_edad = _kpi_by_group(df, idlocal_vals, age_group, is_closed)
    report(45)

    presta_est_vals = df[col_presta_est].fillna("").map(normalize_text) if col_presta_est else pd.Series(["Sin dato"] * len(df))
    presta_est_vals = presta_est_vals.replace("", "Sin dato")
    kpi_presta_est = _kpi_by_group(df, idlocal_vals, presta_est_vals, is_closed)
    report(55)

    csal_vals = df[col_csal].map(normalize_id) if col_csal else pd.Series([""] * len(df))
    kpi_csalida = _kpi_closed_by_group(csal_vals, is_closed, allowed_values=ALLOWED_C_SALIDA_VALUES)
    if kpi_csalida:
        sorted_top = sorted(
            kpi_csalida,
            key=lambda r: (r.get("count", 0), str(r.get("group", ""))),
            reverse=True
        )
        top_codes = {str(row.get("group", "")) for row in sorted_top[:5]}
        for row in kpi_csalida:
            row["top"] = str(row.get("group", "")) in top_codes

    open_mask = ~is_closed
    try:
        open_mask = open_mask.fillna(False)
    except Exception:
        pass
    open_total = int(open_mask.sum())

    wait_sem = {
        "green": 0,
        "yellow": 0,
        "red": 0,
        "total": open_total,
        "green_pct": 0.0,
        "yellow_pct": 0.0,
        "red_pct": 0.0
    }
    if col_fin and open_total > 0:
        fin_vals = df[col_fin].map(normalize_date)
        today = datetime.now().date()
        def _days_since(d: Any) -> Optional[int]:
            if d is None:
                return None
            try:
                if pd.isna(d):
                    return None
            except Exception:
                pass
            try:
                return (today - d.date()).days
            except Exception:
                return None
        days = fin_vals.map(_days_since)
        open_days = days[open_mask]
        valid_days = open_days.dropna()
        if not valid_days.empty:
            green = int((valid_days <= 180).sum())
            yellow = int(((valid_days >= 181) & (valid_days <= 364)).sum())
            red = int((valid_days >= 365).sum())
            total_valid = int(valid_days.shape[0])
            wait_sem = {
                "green": green,
                "yellow": yellow,
                "red": red,
                "total": total_valid,
                "green_pct": round((green / total_valid) * 100, 1) if total_valid else 0.0,
                "yellow_pct": round((yellow / total_valid) * 100, 1) if total_valid else 0.0,
                "red_pct": round((red / total_valid) * 100, 1) if total_valid else 0.0
            }

    death_sem = {
        "alive": open_total,
        "dead": 0,
        "total": open_total,
        "alive_pct": 0.0,
        "dead_pct": 0.0
    }
    if col_run and col_dv and open_total > 0:
        rut = df[col_run].map(normalize_run) + "-" + df[col_dv].map(normalize_dv)
        open_rut = rut[open_mask]
        dead_mask = open_rut.isin(db.defunciones_rut)
        dead = int(dead_mask.sum())
        alive = int(open_total - dead)
        death_sem = {
            "alive": alive,
            "dead": dead,
            "total": open_total,
            "alive_pct": round((alive / open_total) * 100, 1) if open_total else 0.0,
            "dead_pct": round((dead / open_total) * 100, 1) if open_total else 0.0
        }

    total_records = len(df)
    total_ids = int(idlocal_vals[idlocal_vals != ""].nunique())

    filters: List[Dict[str, Any]] = []

    if col_presta_min and col_run and col_dv and col_fin:
        run_vals = df[col_run].map(normalize_run)
        dv_vals = df[col_dv].map(normalize_dv)
        presta_vals = df[col_presta_min].map(normalize_presta)
        fin_vals = df[col_fin].map(normalize_date)
        fout_vals = df[col_fsalida].map(normalize_date) if col_fsalida else pd.Series([None] * len(df))
        ext_vals = df[col_ext].fillna("").map(normalize_text) if col_ext else pd.Series([""] * len(df))
        work_seen: Dict[str, List[TimelineRec]] = defaultdict(list)
        dup_count = 0
        progress_every = max(1, len(df) // 30)
        sql_conn_stats: Optional[Any] = None
        sql_conn_stats_guard: Optional[_PgConnGuard] = None
        sql_timeline_cache: Dict[Tuple[str, str, str, bool, bool], List[TimelineRec]] = {}
        if ROW_LEVEL_SQL_LOOKUPS:
            try:
                sql_conn_stats = _pg_connect()
                sql_conn_stats_guard = _PgConnGuard(sql_conn_stats)
            except Exception:
                sql_conn_stats = None
        try:
            for i in range(len(df)):
                check_cancel()
                msg = compute_duplicidad(
                    run=run_vals.iat[i],
                    dv=dv_vals.iat[i],
                    presta=presta_vals.iat[i],
                    f_in=fin_vals.iat[i],
                    f_out=fout_vals.iat[i],
                    id_local=idlocal_vals.iat[i] or "",
                    extremidad=ext_vals.iat[i],
                    db=db,
                    work_seen=work_seen,
                    sql_conn=sql_conn_stats,
                    sql_cache=sql_timeline_cache,
                )
                if msg.startswith("Caso duplicado"):
                    dup_count += 1
                key_pp = f"{run_vals.iat[i]}|{dv_vals.iat[i]}|{presta_vals.iat[i]}"
                if run_vals.iat[i] and dv_vals.iat[i] and presta_vals.iat[i] and (fin_vals.iat[i] or fout_vals.iat[i]):
                    work_seen[key_pp].append(
                        TimelineRec(
                            f_in=fin_vals.iat[i],
                            f_out=fout_vals.iat[i],
                            sigte_id=idlocal_vals.iat[i] or "",
                            id_local=idlocal_vals.iat[i] or "",
                            source="ESTADISTICAS",
                            extremidad=ext_vals.iat[i]
                        )
                    )
                if progress_cb and (i % progress_every == 0 or i == len(df) - 1):
                    report(55 + int(((i + 1) / max(1, len(df))) * 25))
        finally:
            if sql_conn_stats_guard is not None:
                sql_conn_stats_guard.close()
        filters.append({
            "name": "Duplicidad",
            "count": dup_count,
            "total": total_records
        })
    report(85)

    if col_idlocal:
        in_cgr = idlocal_vals.map(lambda k: bool(k) and (k in db.cgr_399 or k in db.cgr_84))
        filters.append({
            "name": "Cruce CGR",
            "count": int(in_cgr.sum()),
            "total": total_records
        })

    if col_run and col_dv:
        rut = df[col_run].map(normalize_run) + "-" + df[col_dv].map(normalize_dv)
        in_def = rut.isin(db.defunciones_rut)
        filters.append({
            "name": "Cruce Defunciones",
            "count": int(in_def.sum()),
            "total": total_records
        })

    if col_estab:
        cod = df[col_estab].fillna("").astype(str).str.strip()
        macro = ~cod.isin(db.establecimientos)
        filters.append({
            "name": "Cruce Macro red",
            "count": int(macro.sum()),
            "total": total_records
        })
    report(98)

    out = {
        "source_label": source_label,
        "summary": summary,
        "kpi_sexo": kpi_sexo,
        "kpi_edad": kpi_edad,
        "kpi_presta_est": kpi_presta_est,
        "kpi_csalida": kpi_csalida,
        "csalida_labels": C_SALIDA_LABELS,
        "wait_sem": wait_sem,
        "death_sem": death_sem,
        "filters": filters,
        "total_records": total_records,
        "total_ids": total_ids,
        "missing": missing,
    }
    report(100)
    return out



_REDIS_CLIENT: Any = None
if REDIS_URL and redis is not None:
    try:
        candidate = redis.from_url(REDIS_URL, decode_responses=True)
        candidate.ping()
        _REDIS_CLIENT = candidate
    except Exception:
        _REDIS_CLIENT = None

_BACKEND_WARNINGS: List[str] = []
_BACKEND_WARNINGS_EMITTED = False
if REQUESTED_SESSION_BACKEND == "redis" and SESSION_BACKEND != "redis":
    _BACKEND_WARNINGS.append(
        "SESSION_BACKEND=redis pero REDIS_URL esta vacio; se forzo filesystem para sesiones."
    )
if REQUESTED_JOBS_BACKEND == "redis" and JOBS_BACKEND != "redis":
    _BACKEND_WARNINGS.append(
        "JOBS_BACKEND=redis pero REDIS_URL esta vacio; se usara PostgreSQL compartido para jobs si esta disponible."
    )
if SESSION_BACKEND == "redis":
    if Session is None:
        _BACKEND_WARNINGS.append(
            "SESSION_BACKEND=redis pero falta Flask-Session; se usara filesystem para sesiones."
        )
    if _REDIS_CLIENT is None:
        _BACKEND_WARNINGS.append(
            "SESSION_BACKEND=redis pero Redis no esta disponible; se usara filesystem para sesiones."
        )
if JOBS_BACKEND == "redis" and _REDIS_CLIENT is None:
    _BACKEND_WARNINGS.append(
        "JOBS_BACKEND=redis pero Redis no esta disponible; se usara PostgreSQL compartido para jobs si esta disponible."
    )


def validate_runtime_backends() -> None:
    global _BACKEND_WARNINGS_EMITTED
    if not _BACKEND_WARNINGS:
        return
    if STRICT_REDIS_BACKEND:
        raise RuntimeError(" | ".join(_BACKEND_WARNINGS))
    if _BACKEND_WARNINGS_EMITTED:
        return
    for msg in _BACKEND_WARNINGS:
        print(f"Advertencia: {msg}")
    _BACKEND_WARNINGS_EMITTED = True

_DB: Optional[DBIndex] = None
_DB_LOCK = threading.Lock()
_DB_STATUS_LOCK = threading.Lock()
_DB_UPDATING = False
_JOBS: Dict[str, Dict[str, Any]] = {}
_JOBS_LOCK = threading.Lock()
_JOB_DISK_SYNC_MIN_INTERVAL_S = 5.0
_JOB_DISK_SYNC_MIN_PROGRESS_DELTA = 10
_JOB_DISK_SYNC_STATE: Dict[str, Tuple[float, int]] = {}
_CLEANUP_LOCK = threading.Lock()
_LAST_CLEANUP_TS = 0.0
_MAINTENANCE_THREAD_LOCK = threading.Lock()
_MAINTENANCE_THREAD_STARTED = False
_REDIS_JOBS_ENABLED = JOBS_BACKEND == "redis" and _REDIS_CLIENT is not None
_REDIS_JOBS_INDEX_KEY = "{}:jobs:index".format(JOB_STORE_PREFIX)
_REDIS_JOB_KEY_PREFIX = "{}:job:".format(JOB_STORE_PREFIX)
_PG_JOBS_ENABLED = (not _REDIS_JOBS_ENABLED) and (psycopg2 is not None)
_PG_JOBS_TABLE_NAME = "le_noges_jobs_runtime"
_PG_JOBS_INDEX_UPDATED = "idx_le_noges_jobs_runtime_updated"
_PG_JOBS_TABLE_READY = False
_PG_JOBS_WARNED = False
_PG_JOBS_LOCK = threading.Lock()
_PG_JOBS_CONN_LOCAL = threading.local()


def _is_terminal_status(status: str) -> bool:
    return status in {"done", "error", "canceled"}


def _redis_job_key(job_id: str) -> str:
    return "{}{}".format(_REDIS_JOB_KEY_PREFIX, job_id)


def _json_safe_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, dict):
        out: Dict[str, Any] = {}
        for k, v in value.items():
            out[str(k)] = _json_safe_value(v)
        return out
    if isinstance(value, (list, tuple, set)):
        return [_json_safe_value(v) for v in value]
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def _job_progress_int(job: Dict[str, Any]) -> int:
    try:
        return int(float(job.get("progress", 0) or 0))
    except Exception:
        return 0


def _pg_jobs_warn_once(message: str) -> None:
    global _PG_JOBS_WARNED
    if _PG_JOBS_WARNED:
        return
    print(f"Advertencia: {message}")
    _PG_JOBS_WARNED = True


def _pg_jobs_table_qualified() -> str:
    schema = PG_SCHEMA if PG_SCHEMA else "public"
    return f"{_pg_quote_ident(schema)}.{_pg_quote_ident(_PG_JOBS_TABLE_NAME)}"


def _pg_jobs_ensure_table() -> bool:
    global _PG_JOBS_TABLE_READY
    if not _PG_JOBS_ENABLED:
        return False
    if _PG_JOBS_TABLE_READY:
        return True
    with _PG_JOBS_LOCK:
        if _PG_JOBS_TABLE_READY:
            return True
        try:
            table_q = _pg_jobs_table_qualified()
            with _pg_connect() as conn:
                conn.autocommit = True
                with conn.cursor() as cur:
                    cur.execute(
                        f"""
                        CREATE TABLE IF NOT EXISTS {table_q} (
                            job_id TEXT PRIMARY KEY,
                            payload JSONB NOT NULL,
                            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                        );
                        """
                    )
                    cur.execute(
                        f"""
                        CREATE INDEX IF NOT EXISTS {_pg_quote_ident(_PG_JOBS_INDEX_UPDATED)}
                        ON {table_q} (updated_at);
                        """
                    )
            _PG_JOBS_TABLE_READY = True
            return True
        except Exception as e:
            _pg_jobs_warn_once(f"No se pudo inicializar almacenamiento compartido de jobs en PostgreSQL: {e}")
            return False


def _pg_jobs_reset_connection() -> None:
    conn = getattr(_PG_JOBS_CONN_LOCAL, "conn", None)
    if conn is None:
        return
    try:
        conn.close()
    except Exception:
        pass
    _PG_JOBS_CONN_LOCAL.conn = None


def _pg_jobs_connection():
    conn = getattr(_PG_JOBS_CONN_LOCAL, "conn", None)
    try:
        if conn is not None and getattr(conn, "closed", 1) == 0:
            return conn
    except Exception:
        pass
    conn = _pg_connect()
    conn.autocommit = True
    _PG_JOBS_CONN_LOCAL.conn = conn
    return conn


def _pg_jobs_parse_payload(payload: Any) -> Dict[str, Any]:
    if isinstance(payload, dict):
        return payload
    if payload is None:
        return {}
    if isinstance(payload, (bytes, bytearray)):
        try:
            payload = payload.decode("utf-8", errors="replace")
        except Exception:
            return {}
    if isinstance(payload, str):
        try:
            data = json.loads(payload)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    try:
        data = json.loads(str(payload))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _disk_read_job(job_id: str) -> Dict[str, Any]:
    jid = str(job_id or "").strip()
    if not jid:
        return {}
    if not _pg_jobs_ensure_table():
        return {}
    try:
        table_q = _pg_jobs_table_qualified()
        conn = _pg_jobs_connection()
        with conn.cursor() as cur:
            cur.execute(f"SELECT payload FROM {table_q} WHERE job_id=%s;", (jid,))
            row = cur.fetchone()
    except Exception as e:
        _pg_jobs_reset_connection()
        _pg_jobs_warn_once(f"No se pudo leer job '{jid}' desde PostgreSQL: {e}")
        return {}
    if not row:
        return {}
    return _pg_jobs_parse_payload(row[0])


def _disk_write_job(job_id: str, job: Dict[str, Any]) -> None:
    jid = str(job_id or "").strip()
    if not jid:
        return
    if not _pg_jobs_ensure_table():
        return
    try:
        table_q = _pg_jobs_table_qualified()
        payload = json.dumps(_json_safe_value(job), ensure_ascii=False, separators=(",", ":"))
        conn = _pg_jobs_connection()
        with conn.cursor() as cur:
            cur.execute(
                f"""
                INSERT INTO {table_q} (job_id, payload, updated_at)
                VALUES (%s, %s::jsonb, NOW())
                ON CONFLICT (job_id)
                DO UPDATE SET payload = EXCLUDED.payload, updated_at = NOW();
                """,
                (jid, payload),
            )
    except Exception as e:
        _pg_jobs_reset_connection()
        _pg_jobs_warn_once(f"No se pudo escribir job '{jid}' en PostgreSQL: {e}")


def _disk_delete_jobs(job_ids: Iterable[str]) -> None:
    ids = [str(raw_id or "").strip() for raw_id in job_ids]
    ids = [jid for jid in ids if jid]
    if not ids:
        return
    for jid in ids:
        _JOB_DISK_SYNC_STATE.pop(jid, None)
    if not _pg_jobs_ensure_table():
        return
    try:
        table_q = _pg_jobs_table_qualified()
        conn = _pg_jobs_connection()
        with conn.cursor() as cur:
            cur.execute(f"DELETE FROM {table_q} WHERE job_id = ANY(%s);", (ids,))
    except Exception as e:
        _pg_jobs_reset_connection()
        _pg_jobs_warn_once(f"No se pudo eliminar jobs en PostgreSQL: {e}")


def _disk_jobs_snapshot(now_ts: Optional[float] = None, purge: bool = True) -> Dict[str, Dict[str, Any]]:
    now = now_ts if now_ts is not None else time.time()
    if not _pg_jobs_ensure_table():
        return {}

    jobs: Dict[str, Dict[str, Any]] = {}
    stale_ids: List[str] = []
    try:
        table_q = _pg_jobs_table_qualified()
        conn = _pg_jobs_connection()
        with conn.cursor() as cur:
            cur.execute(f"SELECT job_id, payload FROM {table_q};")
            rows = cur.fetchall()
    except Exception as e:
        _pg_jobs_reset_connection()
        _pg_jobs_warn_once(f"No se pudo listar jobs desde PostgreSQL: {e}")
        return {}

    for row in rows:
        try:
            job_id = str(row[0] or "").strip()
        except Exception:
            continue
        if not job_id:
            continue
        job = _pg_jobs_parse_payload(row[1])
        if not job:
            stale_ids.append(job_id)
            continue
        if purge and _job_is_expired(job, now):
            stale_ids.append(job_id)
            continue
        jobs[job_id] = job

    if purge and len(jobs) >= MAX_STORED_JOBS:
        removable = _terminal_jobs_sorted_by_age(jobs.items())
        keep_count = AGGRESSIVE_TERMINAL_JOBS_KEEP
        if len(removable) > keep_count:
            drop_count = len(removable) - keep_count
            for job_id, _ in removable[:drop_count]:
                stale_ids.append(job_id)
    if purge and len(jobs) > MAX_STORED_JOBS:
        removable = _terminal_jobs_sorted_by_age(jobs.items())
        overflow = len(jobs) - MAX_STORED_JOBS
        for job_id, _ in removable[:overflow]:
            stale_ids.append(job_id)

    if stale_ids:
        stale_ids = list(dict.fromkeys(stale_ids))
        _disk_delete_jobs(stale_ids)
        for job_id in stale_ids:
            jobs.pop(job_id, None)

    return jobs


def _redis_read_job(job_id: str) -> Dict[str, Any]:
    if not _REDIS_JOBS_ENABLED:
        return {}
    try:
        payload = _REDIS_CLIENT.get(_redis_job_key(job_id))
    except Exception:
        return {}
    if not payload:
        try:
            _REDIS_CLIENT.srem(_REDIS_JOBS_INDEX_KEY, job_id)
        except Exception:
            pass
        return {}
    try:
        data = json.loads(payload)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _redis_write_job(job_id: str, job: Dict[str, Any]) -> None:
    if not _REDIS_JOBS_ENABLED:
        return
    status = str(job.get("status", ""))
    ttl = JOB_RETENTION_SECONDS if _is_terminal_status(status) else JOB_RUNNING_TTL_SECONDS
    payload = json.dumps(_json_safe_value(job), ensure_ascii=False, separators=(",", ":"))
    try:
        pipe = _REDIS_CLIENT.pipeline()
        pipe.set(_redis_job_key(job_id), payload, ex=int(ttl))
        pipe.sadd(_REDIS_JOBS_INDEX_KEY, job_id)
        pipe.execute()
    except Exception:
        pass


def _redis_delete_jobs(job_ids: Iterable[str]) -> None:
    if not _REDIS_JOBS_ENABLED:
        return
    ids = [str(jid) for jid in job_ids if str(jid).strip()]
    if not ids:
        return
    try:
        pipe = _REDIS_CLIENT.pipeline()
        for jid in ids:
            pipe.delete(_redis_job_key(jid))
            pipe.srem(_REDIS_JOBS_INDEX_KEY, jid)
        pipe.execute()
    except Exception:
        pass


def _job_is_expired(job: Dict[str, Any], now_ts: float) -> bool:
    status = str(job.get("status", ""))
    finished_at = float(job.get("finished_at", 0) or 0)
    if _is_terminal_status(status):
        return bool(finished_at) and (now_ts - finished_at) > JOB_RETENTION_SECONDS
    created_at = float(job.get("created_at", 0) or 0)
    updated_at = float(job.get("updated_at", 0) or 0)
    ref_ts = updated_at if updated_at > 0 else created_at
    return bool(ref_ts) and (now_ts - ref_ts) > JOB_RUNNING_TTL_SECONDS


AGGRESSIVE_TERMINAL_JOBS_KEEP = max(1, min(5, MAX_STORED_JOBS))


def _terminal_jobs_sorted_by_age(items: Iterable[Tuple[str, Dict[str, Any]]]) -> List[Tuple[str, float]]:
    removable: List[Tuple[str, float]] = []
    for jid, job in items:
        status = str(job.get("status", ""))
        if not _is_terminal_status(status):
            continue
        updated_at = float(job.get("updated_at", 0) or 0)
        finished_at = float(job.get("finished_at", 0) or 0)
        stamp = updated_at if updated_at > 0 else finished_at
        removable.append((jid, stamp))
    removable.sort(key=lambda t: t[1])
    return removable


def _jobs_snapshot(now_ts: Optional[float] = None, purge: bool = True) -> Dict[str, Dict[str, Any]]:
    now = now_ts if now_ts is not None else time.time()

    if _REDIS_JOBS_ENABLED:
        try:
            raw_ids = _REDIS_CLIENT.smembers(_REDIS_JOBS_INDEX_KEY) or set()
        except Exception:
            return {}
        jobs: Dict[str, Dict[str, Any]] = {}
        stale_ids: List[str] = []
        for raw_jid in raw_ids:
            jid = str(raw_jid)
            job = _redis_read_job(jid)
            if not job:
                stale_ids.append(jid)
                continue
            if purge and _job_is_expired(job, now):
                stale_ids.append(jid)
                continue
            jobs[jid] = job
        if purge and len(jobs) >= MAX_STORED_JOBS:
            removable = _terminal_jobs_sorted_by_age(jobs.items())
            keep_count = AGGRESSIVE_TERMINAL_JOBS_KEEP
            if len(removable) > keep_count:
                drop_count = len(removable) - keep_count
                for jid, _ in removable[:drop_count]:
                    stale_ids.append(jid)
        if purge and len(jobs) > MAX_STORED_JOBS:
            removable = _terminal_jobs_sorted_by_age(jobs.items())
            overflow = len(jobs) - MAX_STORED_JOBS
            for jid, _ in removable[:overflow]:
                stale_ids.append(jid)
        if stale_ids:
            stale_ids = list(dict.fromkeys(stale_ids))
            _redis_delete_jobs(stale_ids)
            for jid in stale_ids:
                jobs.pop(jid, None)
        return jobs

    with _JOBS_LOCK:
        if purge:
            _purge_jobs_locked(now)
        memory_jobs = {jid: dict(job) for jid, job in _JOBS.items()}

    disk_jobs = _disk_jobs_snapshot(now_ts=now, purge=purge)
    merged: Dict[str, Dict[str, Any]] = dict(disk_jobs)
    for jid, mem_job in memory_jobs.items():
        existing = merged.get(jid)
        if not existing:
            merged[jid] = mem_job
            continue
        try:
            mem_updated = float(mem_job.get("updated_at", 0) or 0)
        except Exception:
            mem_updated = 0.0
        try:
            disk_updated = float(existing.get("updated_at", 0) or 0)
        except Exception:
            disk_updated = 0.0
        if mem_updated >= disk_updated:
            merged[jid] = mem_job
    return merged


def _set_db_updating(flag: bool) -> None:
    global _DB_UPDATING
    with _DB_STATUS_LOCK:
        _DB_UPDATING = flag


def _is_db_updating() -> bool:
    with _DB_STATUS_LOCK:
        return _DB_UPDATING


def _refresh_db_if_needed(force: bool = False) -> None:
    global _DB
    with _DB_LOCK:
        _set_db_updating(True)
        try:
            _clear_sql_timeline_columns_cache()
            if LOW_MEMORY_DB_REFRESH and _DB is not None:
                _DB.reload_all_in_place()
            else:
                old_db = _DB
                db = DBIndex(DB_DIR)
                db.load_all()
                _DB = db
                if old_db is not None:
                    del old_db
            _release_memory_pressure()
        finally:
            _set_db_updating(False)


def _ensure_db_loaded_locked() -> DBIndex:
    global _DB
    if _DB is None:
        db = DBIndex(DB_DIR)
        db.load_all()
        _DB = db
    return _DB


def get_db() -> DBIndex:
    with _DB_LOCK:
        return _ensure_db_loaded_locked()


def _safe_unlink(path: Optional[Path]) -> None:
    if not path:
        return
    try:
        if path.exists() and path.is_file():
            path.unlink()
    except Exception:
        pass


def _purge_jobs_locked(now_ts: Optional[float] = None) -> None:
    now = now_ts if now_ts is not None else time.time()
    expired_ids: List[str] = []
    for jid, job in _JOBS.items():
        if _job_is_expired(job, now):
            expired_ids.append(jid)
    for jid in expired_ids:
        _JOBS.pop(jid, None)

    if len(_JOBS) >= MAX_STORED_JOBS:
        removable = _terminal_jobs_sorted_by_age(_JOBS.items())
        keep_count = AGGRESSIVE_TERMINAL_JOBS_KEEP
        if len(removable) > keep_count:
            drop_count = len(removable) - keep_count
            for jid, _ in removable[:drop_count]:
                _JOBS.pop(jid, None)

    if len(_JOBS) <= MAX_STORED_JOBS:
        return
    removable = _terminal_jobs_sorted_by_age(_JOBS.items())
    while len(_JOBS) > MAX_STORED_JOBS and removable:
        jid, _ = removable.pop(0)
        _JOBS.pop(jid, None)


def _cleanup_app_temp_files(force: bool = False) -> None:
    global _LAST_CLEANUP_TS
    now = time.time()
    if not force:
        interval = max(0, int(CLEANUP_INTERVAL_SECONDS or 0))
        if interval > 0:
            with _CLEANUP_LOCK:
                last_ts = float(_LAST_CLEANUP_TS or 0.0)
                if (now - last_ts) < float(interval):
                    return
                _LAST_CLEANUP_TS = now
    else:
        with _CLEANUP_LOCK:
            _LAST_CLEANUP_TS = now

    protected_names: set = set()
    jobs = _jobs_snapshot(now_ts=now, purge=True)
    for job in jobs.values():
        out_name = str(job.get("out_file", "") or "").strip()
        in_name = str(job.get("input_file", "") or "").strip()
        if out_name:
            protected_names.add(out_name)
        if in_name:
            protected_names.add(Path(in_name).name)

    targets = {UPLOAD_DIR.resolve(), OUTPUT_DIR.resolve()}
    seen: set = set()
    for directory in targets:
        dkey = str(directory)
        if dkey in seen:
            continue
        seen.add(dkey)
        try:
            for path in directory.iterdir():
                if not path.is_file():
                    continue
                name = path.name
                if not name.startswith(APP_OUTPUT_PREFIXES):
                    continue
                if name in protected_names:
                    continue
                try:
                    age = now - path.stat().st_mtime
                except Exception:
                    continue
                if age > FILE_RETENTION_SECONDS:
                    _safe_unlink(path)
        except Exception:
            continue


def _maintenance_loop() -> None:
    delay = max(0, int(BACKGROUND_CLEANUP_STARTUP_DELAY_SECONDS or 0))
    if delay > 0:
        time.sleep(delay)
    interval = max(30, int(CLEANUP_INTERVAL_SECONDS or 300))
    while True:
        try:
            _cleanup_app_temp_files(force=True)
        except Exception:
            pass
        time.sleep(interval)


def start_background_maintenance() -> None:
    global _MAINTENANCE_THREAD_STARTED
    if not BACKGROUND_CLEANUP_ENABLED:
        return
    with _MAINTENANCE_THREAD_LOCK:
        if _MAINTENANCE_THREAD_STARTED:
            return
        worker = threading.Thread(
            target=_maintenance_loop,
            name="le_noges_cleanup",
            daemon=True,
        )
        worker.start()
        _MAINTENANCE_THREAD_STARTED = True


def _can_start_new_job() -> bool:
    jobs = _jobs_snapshot(purge=True)
    running = 0
    for job in jobs.values():
        if str(job.get("status", "")) == "running":
            running += 1
    return running < MAX_CONCURRENT_JOBS


def _current_user_rut() -> str:
    return str(session.get("rut", "") or "").strip().upper()


def _current_user_is_admin() -> bool:
    return _is_admin_user(_current_user_rut())


def _job_belongs_to_current_user(job: Dict[str, Any]) -> bool:
    owner = str(job.get("owner_rut", "") or "").strip().upper()
    current = _current_user_rut()
    return bool(owner) and bool(current) and owner == current


def _get_owned_job(job_id: str) -> Dict[str, Any]:
    job = _get_job(job_id)
    if not job:
        return {}
    if not _job_belongs_to_current_user(job):
        return {}
    return job


def _get_owned_or_admin_job(job_id: str) -> Dict[str, Any]:
    job = _get_job(job_id)
    if not job:
        return {}
    if _current_user_is_admin():
        return job
    if not _job_belongs_to_current_user(job):
        return {}
    return job


def _get_shared_job(job_id: str) -> Dict[str, Any]:
    now = time.time()
    jid = str(job_id or "").strip()
    if not jid:
        return {}
    if _REDIS_JOBS_ENABLED:
        job = _redis_read_job(jid)
        if not job:
            return {}
        if _job_is_expired(job, now):
            _redis_delete_jobs([jid])
            return {}
        return dict(job)

    shared_job = _disk_read_job(jid)
    if not shared_job:
        return {}
    if _job_is_expired(shared_job, now):
        _disk_delete_jobs([jid])
        return {}

    def _updated_ts(job_data: Dict[str, Any]) -> float:
        try:
            return float(job_data.get("updated_at", 0) or 0)
        except Exception:
            return 0.0

    with _JOBS_LOCK:
        existing = _JOBS.get(jid)
        if (not existing) or (_updated_ts(shared_job) >= _updated_ts(existing)):
            _JOBS[jid] = dict(shared_job)
        _JOB_DISK_SYNC_STATE.setdefault(
            jid,
            (float(shared_job.get("updated_at", now) or now), _job_progress_int(shared_job)),
        )
    return dict(shared_job)


def _get_owned_job_fresh(job_id: str) -> Dict[str, Any]:
    job = _get_shared_job(job_id)
    if not job:
        return {}
    if not _job_belongs_to_current_user(job):
        return {}
    return job


def _current_user_can_download(filename: str) -> bool:
    target = (filename or "").strip()
    if not target:
        return False
    owner = _current_user_rut()
    if not owner:
        return False
    jobs = _jobs_snapshot(purge=True)
    for job in jobs.values():
        if str(job.get("owner_rut", "")).strip().upper() != owner:
            continue
        if str(job.get("out_file", "")).strip() == target:
            return True
    return False


def _init_job(
    job_id: str,
    owner_rut: str = "",
    input_file: str = "",
    repeat_url: str = "",
) -> None:
    now = time.time()
    job_data = {
        "status": "running",
        "progress": 0,
        "out_file": "",
        "elapsed_display": "",
        "error": "",
        "cancel_requested": False,
        "owner_rut": (owner_rut or "").strip().upper(),
        "input_file": input_file,
        "repeat_url": repeat_url,
        "created_at": now,
        "updated_at": now,
        "finished_at": None,
    }
    if _REDIS_JOBS_ENABLED:
        jobs = _jobs_snapshot(now_ts=now, purge=True)
        running = 0
        for job in jobs.values():
            if str(job.get("status", "")) == "running":
                running += 1
        if running >= MAX_CONCURRENT_JOBS:
            raise RuntimeError(
                f"Hay {MAX_CONCURRENT_JOBS} procesos en ejecucion. Espera a que termine uno para iniciar otro."
            )
        _redis_write_job(job_id, job_data)
    else:
        with _JOBS_LOCK:
            _purge_jobs_locked(now)
            running = 0
            for job in _JOBS.values():
                if str(job.get("status", "")) == "running":
                    running += 1
            if running >= MAX_CONCURRENT_JOBS:
                raise RuntimeError(
                    f"Hay {MAX_CONCURRENT_JOBS} procesos en ejecucion. Espera a que termine uno para iniciar otro."
                )
            _JOBS[job_id] = job_data
            _JOB_DISK_SYNC_STATE[job_id] = (now, _job_progress_int(job_data))
        _disk_write_job(job_id, job_data)
    _cleanup_app_temp_files()


def _update_job(job_id: str, **kwargs: Any) -> None:
    now = time.time()
    must_cleanup = False
    if _REDIS_JOBS_ENABLED:
        job = _redis_read_job(job_id)
        if not job:
            return
        current_status = str(job.get("status", ""))
        incoming_status = str(kwargs.get("status", current_status))
        if current_status == "canceled" and incoming_status != "canceled":
            kwargs.pop("status", None)
        job.update(kwargs)
        job["updated_at"] = now
        status = str(job.get("status", ""))
        if _is_terminal_status(status):
            job["finished_at"] = now
            must_cleanup = True
        _redis_write_job(job_id, job)
    else:
        job_to_write: Dict[str, Any] = {}
        should_persist = False
        with _JOBS_LOCK:
            if job_id not in _JOBS:
                recovered = _disk_read_job(job_id)
                if recovered:
                    _JOBS[job_id] = recovered
                    _JOB_DISK_SYNC_STATE[job_id] = (
                        float(recovered.get("updated_at", now) or now),
                        _job_progress_int(recovered),
                    )
            if job_id not in _JOBS:
                return
            before_job = dict(_JOBS[job_id])
            current_status = str(_JOBS[job_id].get("status", ""))
            incoming_status = str(kwargs.get("status", current_status))
            if current_status == "canceled" and incoming_status != "canceled":
                kwargs.pop("status", None)
            _JOBS[job_id].update(kwargs)
            _JOBS[job_id]["updated_at"] = now
            status = str(_JOBS[job_id].get("status", ""))
            if _is_terminal_status(status):
                _JOBS[job_id]["finished_at"] = now
                must_cleanup = True
            job_to_write = dict(_JOBS[job_id])
            status_changed = str(before_job.get("status", "")) != status
            if must_cleanup or status_changed:
                should_persist = True
            elif any(
                key in kwargs
                for key in (
                    "error",
                    "out_file",
                    "elapsed_display",
                    "cancel_requested",
                    "repeat_url",
                    "owner_rut",
                    "input_file",
                    "finished_at",
                    "stats",
                    "cross_stats",
                    "mediana_stats",
                    "stage_timing",
                )
            ):
                should_persist = True
            elif "progress" in kwargs:
                last_sync = _JOB_DISK_SYNC_STATE.get(job_id)
                curr_progress = _job_progress_int(job_to_write)
                if last_sync is None:
                    should_persist = True
                else:
                    last_ts, last_progress = last_sync
                    progress_delta = abs(curr_progress - int(last_progress))
                    time_delta = max(0.0, now - float(last_ts))
                    if (
                        progress_delta >= _JOB_DISK_SYNC_MIN_PROGRESS_DELTA
                        or time_delta >= _JOB_DISK_SYNC_MIN_INTERVAL_S
                    ):
                        should_persist = True
            if should_persist:
                if must_cleanup:
                    _JOB_DISK_SYNC_STATE.pop(job_id, None)
                else:
                    _JOB_DISK_SYNC_STATE[job_id] = (now, _job_progress_int(job_to_write))
        if should_persist:
            _disk_write_job(job_id, job_to_write)
    if must_cleanup:
        _cleanup_app_temp_files()


def _get_job(job_id: str) -> Dict[str, Any]:
    now = time.time()
    if _REDIS_JOBS_ENABLED:
        job = _redis_read_job(job_id)
        if not job:
            return {}
        if _job_is_expired(job, now):
            _redis_delete_jobs([job_id])
            return {}
        return dict(job)
    with _JOBS_LOCK:
        _purge_jobs_locked(now)
        in_memory = dict(_JOBS.get(job_id, {}))
    if in_memory:
        return in_memory
    disk_job = _disk_read_job(job_id)
    if not disk_job:
        return {}
    if _job_is_expired(disk_job, now):
        _disk_delete_jobs([job_id])
        return {}
    with _JOBS_LOCK:
        if job_id not in _JOBS:
            _JOBS[job_id] = dict(disk_job)
        _JOB_DISK_SYNC_STATE.setdefault(
            job_id,
            (float(disk_job.get("updated_at", now) or now), _job_progress_int(disk_job)),
        )
    return dict(disk_job)


def _request_cancel_job(job_id: str) -> bool:
    if _REDIS_JOBS_ENABLED:
        job = _redis_read_job(job_id)
        if not job:
            return False
        status = str(job.get("status", ""))
        if _is_terminal_status(status):
            return False
        job["cancel_requested"] = True
        job["updated_at"] = time.time()
        _redis_write_job(job_id, job)
        return True
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        if not job:
            recovered = _disk_read_job(job_id)
            if recovered:
                _JOBS[job_id] = recovered
                job = _JOBS.get(job_id)
        if not job:
            return False
        status = str(job.get("status", ""))
        if _is_terminal_status(status):
            return False
        job["cancel_requested"] = True
        job["updated_at"] = time.time()
        _disk_write_job(job_id, dict(job))
        _JOB_DISK_SYNC_STATE[job_id] = (float(job.get("updated_at", time.time()) or time.time()), _job_progress_int(job))
        return True


def _is_cancel_requested(job_id: str) -> bool:
    if _REDIS_JOBS_ENABLED:
        job = _redis_read_job(job_id)
        if not job:
            return True
        return bool(job.get("cancel_requested", False))
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        if job:
            return bool(job.get("cancel_requested", False))
    disk_job = _disk_read_job(job_id)
    if not disk_job:
        return True
    if bool(disk_job.get("cancel_requested", False)):
        with _JOBS_LOCK:
            if job_id in _JOBS:
                _JOBS[job_id]["cancel_requested"] = True
    return bool(disk_job.get("cancel_requested", False))


class JobCancelledError(RuntimeError):
    pass


def _raise_if_cancel_requested(job_id: str) -> None:
    if _is_cancel_requested(job_id):
        raise JobCancelledError("Proceso cancelado por usuario.")


def _mark_job_canceled(job_id: str, message: str = "Proceso cancelado por usuario.") -> None:
    _update_job(job_id, status="canceled", error=message)


def _run_job(job_id: str, work_path: Path, selected: Dict[str, bool]) -> None:
    _update_job(job_id, status="running", progress=0, error="")
    t0 = time.perf_counter()
    out_path: Optional[Path] = None
    stage_timing: Dict[str, Any] = {}
    try:
        _raise_if_cancel_requested(job_id)
        db = get_db()

        def cb(p: int) -> None:
            _raise_if_cancel_requested(job_id)
            _update_job(job_id, progress=p)

        out_path, _elapsed, stage_timing = process_file(
            work_path,
            selected,
            db=db,
            progress_cb=cb,
            cancel_cb=lambda: _raise_if_cancel_requested(job_id),
        )
        _raise_if_cancel_requested(job_id)
        elapsed_total = time.perf_counter() - t0
        _update_job(
            job_id,
            status="done",
            progress=100,
            out_file=out_path.name,
            elapsed_display=format_duration(elapsed_total),
            stage_timing=stage_timing,
        )
    except JobCancelledError as e:
        if out_path:
            _safe_unlink(out_path)
        _mark_job_canceled(job_id, str(e))
    except Exception as e:
        _update_job(job_id, status="error", error=str(e))
    finally:
        _safe_unlink(work_path)
        _release_memory_pressure()
        _pg_jobs_reset_connection()


def _run_stats_job(job_id: str, source: str, work_path: Optional[Path]) -> None:
    _update_job(job_id, status="running", progress=0, error="", stats=None, source=source)
    try:
        _raise_if_cancel_requested(job_id)
        db = get_db()

        def cb(p: int) -> None:
            _raise_if_cancel_requested(job_id)
            _update_job(job_id, progress=p)

        stats = build_statistics(
            source,
            work_path,
            db,
            progress_cb=cb,
            cancel_cb=lambda: _raise_if_cancel_requested(job_id),
        )
        _raise_if_cancel_requested(job_id)
        _update_job(job_id, status="done", progress=100, stats=stats, source=source)
    except JobCancelledError as e:
        _mark_job_canceled(job_id, str(e))
    except Exception as e:
        _update_job(job_id, status="error", error=str(e))
    finally:
        _safe_unlink(work_path)
        _release_memory_pressure()
        _pg_jobs_reset_connection()


def _run_cross_stats_job(job_id: str, work_path: Optional[Path]) -> None:
    _update_job(job_id, status="running", progress=0, error="", cross_stats=None)
    processed_output: Optional[Path] = None
    try:
        _raise_if_cancel_requested(job_id)
        db = get_db()

        def cb(p: int) -> None:
            _raise_if_cancel_requested(job_id)
            _update_job(job_id, progress=p)

        stats = build_cross_statistics(
            work_path,
            db,
            progress_cb=cb,
            cancel_cb=lambda: _raise_if_cancel_requested(job_id),
        )
        processed_name = str(stats.get("processed_file", "") or "").strip() if isinstance(stats, dict) else ""
        if processed_name:
            processed_output = OUTPUT_DIR / processed_name
        _raise_if_cancel_requested(job_id)
        _update_job(job_id, status="done", progress=100, cross_stats=stats)
    except JobCancelledError as e:
        _mark_job_canceled(job_id, str(e))
    except Exception as e:
        _update_job(job_id, status="error", error=str(e))
    finally:
        _safe_unlink(work_path)
        _safe_unlink(processed_output)
        _release_memory_pressure()
        _pg_jobs_reset_connection()


def _run_mediana_job(
    job_id: str,
    work_path: Path,
    fechas_corte: Dict[str, datetime],
    fechas_Percentil: Dict[str, datetime],
    ideales: Dict[str, int],
    active_classes: List[str],
) -> None:
    _update_job(job_id, status="running", progress=0, error="", mediana_stats=None)
    t0 = time.perf_counter()
    out_path: Optional[Path] = None
    try:
        _raise_if_cancel_requested(job_id)
        db = get_db()

        def cb(p: int) -> None:
            _raise_if_cancel_requested(job_id)
            _update_job(job_id, progress=p)

        out_path, stats, _elapsed = process_mediana_file(
            work_path,
            fechas_corte,
            fechas_Percentil,
            ideales,
            active_classes,
            db=db,
            progress_cb=cb,
            cancel_cb=lambda: _raise_if_cancel_requested(job_id),
        )
        _raise_if_cancel_requested(job_id)
        elapsed_total = time.perf_counter() - t0
        _update_job(
            job_id,
            status="done",
            progress=100,
            mediana_stats=stats,
            out_file=out_path.name,
            elapsed_display=format_duration(elapsed_total),
        )
    except JobCancelledError as e:
        if out_path:
            _safe_unlink(out_path)
        _mark_job_canceled(job_id, str(e))
    except Exception as e:
        _update_job(job_id, status="error", error=str(e))
    finally:
        _safe_unlink(work_path)
        _release_memory_pressure()
        _pg_jobs_reset_connection()


def _admin_write_update_report(
    rows: List[Dict[str, Any]],
    owner_rut: str,
    cores_updated: bool,
    selected_tables: List[str],
    out_path: Path,
) -> None:
    meta_rows = [
        {"Campo": "Usuario", "Valor": owner_rut},
        {"Campo": "Fecha ejecucion", "Valor": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Campo": "Tablas seleccionadas", "Valor": ", ".join(selected_tables)},
        {"Campo": "Cores/indices actualizados", "Valor": "SI" if cores_updated else "NO"},
    ]
    with pd.ExcelWriter(
        out_path,
        engine="openpyxl",
        date_format="DD-MM-YYYY",
        datetime_format="DD-MM-YYYY",
    ) as writer:
        format_dates_for_export(pd.DataFrame(rows)).to_excel(writer, sheet_name="Carga_RAW", index=False)
        format_dates_for_export(pd.DataFrame(meta_rows)).to_excel(writer, sheet_name="Resumen", index=False)
        apply_short_date_format_to_workbook(writer.book)


def _run_admin_update_job(
    job_id: str,
    owner_rut: str,
    uploads: Dict[str, Path],
    selected_tables: List[str],
    run_cores: bool,
    load_modes: Optional[Dict[str, str]] = None,
) -> None:
    _update_job(job_id, status="running", progress=0, error="")
    t0 = time.perf_counter()
    out_path: Optional[Path] = None
    report_rows: List[Dict[str, Any]] = []
    cores_updated = False
    modes = load_modes or {}
    _set_db_updating(True)
    try:
        _raise_if_cancel_requested(job_id)
        with _pg_connect() as conn:
            conn.autocommit = False
            total = max(1, len(selected_tables))
            for idx, table_name in enumerate(selected_tables):
                _raise_if_cancel_requested(job_id)
                input_path = uploads[table_name]
                before_count = _admin_count_table(conn, table_name)
                table_cols = _admin_get_table_columns(conn, table_name)
                df_aligned, sheet_summary = _admin_read_input_for_table(input_path, table_name, table_cols)
                load_mode = _admin_resolve_load_mode(table_name, modes.get(table_name, "replace"))
                _admin_copy_with_mode(conn, table_name, df_aligned, load_mode)
                after_count = _admin_count_table(conn, table_name)
                sheet_text = " | ".join(
                    [
                        f"hoja={row['hoja']} header={row['header']} match={row['match']} filas={row['filas']}"
                        for row in sheet_summary
                    ]
                )
                report_rows.append(
                    {
                        "Tabla": table_name,
                        "Archivo": input_path.name,
                        "Filas antes": before_count,
                        "Filas cargadas": int(len(df_aligned)),
                        "Filas despues": after_count,
                        "Modo carga": ("Sobreescribir datos nuevos" if load_mode == "append" else "Limpiar base y cargar"),
                        "Detalle hojas": sheet_text,
                    }
                )
                progress = 10 + int(((idx + 1) * 70) / total)
                _update_job(job_id, progress=min(progress, 80))
                del df_aligned
                del sheet_summary
                _release_memory_pressure(rounds=1, sleep_ms=0)

            if run_cores and _admin_should_refresh_cores(selected_tables):
                _raise_if_cancel_requested(job_id)
                _update_job(job_id, progress=85)
                _admin_run_cores_and_indices(conn)
                cores_updated = True
                _update_job(job_id, progress=92)
            conn.commit()

        _raise_if_cancel_requested(job_id)
        _refresh_db_if_needed(force=True)
        _update_job(job_id, progress=96)

        out_name = f"LE_NOGES_admin_update_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{job_id}.xlsx"
        out_path = OUTPUT_DIR / out_name
        _admin_write_update_report(report_rows, owner_rut, cores_updated, selected_tables, out_path)
        elapsed_total = time.perf_counter() - t0
        _update_job(
            job_id,
            status="done",
            progress=100,
            out_file=out_path.name,
            elapsed_display=format_duration(elapsed_total),
            admin_summary=report_rows,
            cores_updated=cores_updated,
        )
    except JobCancelledError as e:
        if out_path:
            _safe_unlink(out_path)
        _mark_job_canceled(job_id, str(e))
    except Exception as e:
        if out_path:
            _safe_unlink(out_path)
        _update_job(job_id, status="error", error=str(e))
    finally:
        for path in uploads.values():
            _safe_unlink(path)
        _set_db_updating(False)
        _release_memory_pressure(rounds=MEMORY_TRIM_ROUNDS + 1, sleep_ms=MEMORY_TRIM_SLEEP_MS)
        _pg_jobs_reset_connection()


def _stats_to_excel(stats: Dict[str, Any], out_path: Path) -> None:
    def to_df(items: Any) -> pd.DataFrame:
        if isinstance(items, list):
            return pd.DataFrame(items)
        if isinstance(items, dict):
            return pd.DataFrame([items])
        return pd.DataFrame()

    meta_rows = [
        {"Campo": "Fuente", "Valor": stats.get("source_label", "")},
        {"Campo": "Registros", "Valor": stats.get("total_records", 0)},
        {"Campo": "ID_LOCAL", "Valor": stats.get("total_ids", 0)},
        {"Campo": "Generado", "Valor": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]

    cs_labels = stats.get("csalida_labels") or {}
    cs_rows = []
    kpi_csalida_rows = stats.get("kpi_csalida") or []
    cs_map = {str(r.get("group", "")): r for r in kpi_csalida_rows}
    total_closed = 0
    if kpi_csalida_rows:
        try:
            total_closed = int(kpi_csalida_rows[0].get("total", 0) or 0)
        except Exception:
            total_closed = 0
    for code in ALLOWED_C_SALIDA_VALUES:
        row = cs_map.get(code, {})
        count = int(row.get("count", 0) or 0)
        pct = round((count / total_closed) * 100, 1) if total_closed else 0.0
        desc = str(cs_labels.get(code, code))
        cs_rows.append({
            "C_SALIDA": code,
            "Descripcion": desc,
            "Casos": count,
            "Porcentaje": pct,
            "Total Cerrados": total_closed
        })

    wait = stats.get("wait_sem") or {}
    wait_rows = [
        {"Rango": "< 180 dias", "Casos": wait.get("green", 0), "Porcentaje": wait.get("green_pct", 0.0)},
        {"Rango": "181-364 dias", "Casos": wait.get("yellow", 0), "Porcentaje": wait.get("yellow_pct", 0.0)},
        {"Rango": ">= 365 dias", "Casos": wait.get("red", 0), "Porcentaje": wait.get("red_pct", 0.0)},
    ]

    death = stats.get("death_sem") or {}
    death_rows = [
        {"Estado": "Vivos", "Casos": death.get("alive", 0), "Porcentaje": death.get("alive_pct", 0.0)},
        {"Estado": "Fallecidos", "Casos": death.get("dead", 0), "Porcentaje": death.get("dead_pct", 0.0)},
    ]

    with pd.ExcelWriter(
        out_path,
        engine="openpyxl",
        date_format="DD-MM-YYYY",
        datetime_format="DD-MM-YYYY",
    ) as writer:
        format_dates_for_export(pd.DataFrame(meta_rows)).to_excel(writer, sheet_name="Resumen", index=False)
        format_dates_for_export(to_df(stats.get("summary"))).to_excel(writer, sheet_name="TIPO_PREST", index=False)
        format_dates_for_export(to_df(stats.get("kpi_sexo"))).to_excel(writer, sheet_name="Sexo", index=False)
        format_dates_for_export(to_df(stats.get("kpi_edad"))).to_excel(writer, sheet_name="Rango_Etario", index=False)
        format_dates_for_export(to_df(stats.get("kpi_presta_est"))).to_excel(writer, sheet_name="PRESTA_EST", index=False)
        format_dates_for_export(pd.DataFrame(cs_rows)).to_excel(writer, sheet_name="C_SALIDA", index=False)
        format_dates_for_export(to_df(stats.get("filters"))).to_excel(writer, sheet_name="Cruces", index=False)
        format_dates_for_export(pd.DataFrame(wait_rows)).to_excel(writer, sheet_name="Tiempo de Espera", index=False)
        format_dates_for_export(pd.DataFrame(death_rows)).to_excel(writer, sheet_name="Defunciones", index=False)
        apply_short_date_format_to_workbook(writer.book)

        def _col_index(ws: openpyxl.worksheet.worksheet.Worksheet, header: str) -> Optional[int]:
            for cell in ws[1]:
                if cell.value == header:
                    return cell.col_idx
            return None

        def _add_pie_chart(
            ws: openpyxl.worksheet.worksheet.Worksheet,
            title: str,
            data_col: Optional[int],
            cat_col: Optional[int],
            data_end: int,
            anchor: str
        ) -> None:
            if not ws or not data_col or not cat_col or data_end < 2:
                return
            chart = PieChart()
            data = Reference(ws, min_col=data_col, min_row=1, max_row=data_end)
            cats = Reference(ws, min_col=cat_col, min_row=2, max_row=data_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.title = title
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showPercent = True
            ws.add_chart(chart, anchor)

        def _add_bar_chart(
            ws: openpyxl.worksheet.worksheet.Worksheet,
            title: str,
            data_col: Optional[int],
            cat_col: Optional[int],
            data_end: int,
            anchor: str,
            show_labels: bool = True,
            horizontal: bool = False,
            chart_height: Optional[float] = None,
            chart_width: Optional[float] = None
        ) -> None:
            if not ws or not data_col or not cat_col or data_end < 2:
                return
            chart = BarChart()
            if horizontal:
                chart.type = "bar"
            data = Reference(ws, min_col=data_col, min_row=1, max_row=data_end)
            cats = Reference(ws, min_col=cat_col, min_row=2, max_row=data_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.title = title
            rows = max(1, data_end - 1)
            if horizontal:
                chart.height = max(7.5, rows * 0.35)
                chart.width = 13.0
            else:
                chart.height = 8.0
                chart.width = 11.0
            if chart_height is not None:
                chart.height = chart_height
            if chart_width is not None:
                chart.width = chart_width
            if horizontal:
                chart.x_axis.title = "Casos"
                try:
                    chart.y_axis.tickLblSkip = 1
                    chart.y_axis.tickMarkSkip = 1
                except Exception:
                    pass
            else:
                chart.y_axis.title = "Casos"
                try:
                    chart.x_axis.tickLblSkip = 1
                    chart.x_axis.tickMarkSkip = 1
                except Exception:
                    pass
            if show_labels:
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True
            ws.add_chart(chart, anchor)

        def _add_pie_range(
            ws: openpyxl.worksheet.worksheet.Worksheet,
            title: str,
            data_col: int,
            cat_col: int,
            header_row: int,
            data_start: int,
            data_end: int,
            anchor: str
        ) -> None:
            if data_end < data_start:
                return
            chart = PieChart()
            data = Reference(ws, min_col=data_col, min_row=header_row, max_row=data_end)
            cats = Reference(ws, min_col=cat_col, min_row=data_start, max_row=data_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.title = title
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showPercent = True
            ws.add_chart(chart, anchor)

        ws_resumen = writer.sheets.get("Resumen")
        if ws_resumen:
            summary = stats.get("summary") or []
            total_open = sum(int(r.get("open", 0) or 0) for r in summary)
            total_closed = sum(int(r.get("closed", 0) or 0) for r in summary)
            if total_open or total_closed:
                start_row = ws_resumen.max_row + 2
                ws_resumen.cell(row=start_row, column=1, value="Estado")
                ws_resumen.cell(row=start_row, column=2, value="Casos")
                ws_resumen.cell(row=start_row + 1, column=1, value="Abiertos")
                ws_resumen.cell(row=start_row + 1, column=2, value=total_open)
                ws_resumen.cell(row=start_row + 2, column=1, value="Cerrados")
                ws_resumen.cell(row=start_row + 2, column=2, value=total_closed)
                _add_pie_range(
                    ws_resumen,
                    "Estado de casos",
                    2,
                    1,
                    start_row,
                    start_row + 1,
                    start_row + 2,
                    f"D{start_row}"
                )

        ws_tipo = writer.sheets.get("TIPO_PREST")
        if ws_tipo:
            _add_pie_chart(
                ws_tipo,
                "Resumen por TIPO_PREST",
                _col_index(ws_tipo, "total"),
                _col_index(ws_tipo, "group"),
                ws_tipo.max_row,
                "H2"
            )

        ws_sexo = writer.sheets.get("Sexo")
        if ws_sexo:
            _add_pie_chart(
                ws_sexo,
                "Sexo",
                _col_index(ws_sexo, "total"),
                _col_index(ws_sexo, "group"),
                ws_sexo.max_row,
                "H2"
            )

        ws_edad = writer.sheets.get("Rango_Etario")
        if ws_edad:
            _add_pie_chart(
                ws_edad,
                "Rango Etario",
                _col_index(ws_edad, "total"),
                _col_index(ws_edad, "group"),
                ws_edad.max_row,
                "H2"
            )

        ws_csal = writer.sheets.get("C_SALIDA")
        if ws_csal:
            cs_data_end = 1 + len(cs_rows)

            legend_start = cs_data_end + 3
            ws_csal.cell(row=legend_start, column=1, value="Leyenda C_SALIDA")
            ws_csal.cell(row=legend_start + 1, column=1, value="Codigo")
            ws_csal.cell(row=legend_start + 1, column=2, value="Descripcion")
            for i, row in enumerate(cs_rows):
                ws_csal.cell(row=legend_start + 2 + i, column=1, value=row.get("C_SALIDA", ""))
                ws_csal.cell(row=legend_start + 2 + i, column=2, value=row.get("Descripcion", ""))

            cs_rows_count = max(1, len(cs_rows))
            _add_bar_chart(
                ws_csal,
                "C_SALIDA",
                _col_index(ws_csal, "Casos"),
                _col_index(ws_csal, "C_SALIDA"),
                cs_data_end,
                "G2",
                show_labels=False,
                horizontal=True,
                chart_height=max(10.0, cs_rows_count * 0.45),
                chart_width=18.0
            )

        ws_cruces = writer.sheets.get("Cruces")
        if ws_cruces:
            _add_bar_chart(
                ws_cruces,
                "Resumen de Cruces",
                _col_index(ws_cruces, "count"),
                _col_index(ws_cruces, "name"),
                ws_cruces.max_row,
                "E2"
            )

        ws_wait = writer.sheets.get("Tiempo de Espera")
        if ws_wait:
            _add_bar_chart(
                ws_wait,
                "Tiempos de espera",
                _col_index(ws_wait, "Casos"),
                _col_index(ws_wait, "Rango"),
                ws_wait.max_row,
                "E2"
            )

        ws_death = writer.sheets.get("Defunciones")
        if ws_death:
            _add_pie_chart(
                ws_death,
                "Defunciones",
                _col_index(ws_death, "Casos"),
                _col_index(ws_death, "Estado"),
                ws_death.max_row,
                "E2"
            )



